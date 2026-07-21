"""Normalise the BP hardware-development workbook into matrix.json.

The workbook (IoT BP V3.x.xlsx) is the single source of truth for the HW NPD
process. The Markdown deliverables in docs/hw-npd-process/ are derived from it,
so when the workbook version bumps, re-run this and regenerate the docs rather
than hand-patching them.

Usage:
    python extract.py [path/to/IoT BP V3.6.xlsx]
"""

import json
import sys
from pathlib import Path

import openpyxl

DEFAULT_WORKBOOK = Path.home() / "Downloads" / "IoT BP V3.6.xlsx"
OUT = Path(__file__).with_name("matrix.json")

ROLES = [
    "CEO",
    "CTO",
    "IoT Product Owner",
    "IoT Engineer",
    "Application Engineer",
    "Kernel Engineer",
    "SW Product",
    "SW Engineering",
    "QA / QC",
    "Procurement",
    "Finance",
    "Head of Service",
    "Commercial (CBO/Mktg)",
]

LANES = {
    "Full NPD": "full_npd",
    "HW Upgrade": "hw_upgrade",
    "FW Upgrade": "fw_upgrade",
    "New SW Feature": "new_sw_feature",
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def split_names(value):
    """The workbook stores R/A as a leading-comma joined list, e.g. ', A, B'."""
    return [part.strip() for part in clean(value).split(",") if part.strip()]


def extract(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["Master Matrix"]

    header = [clean(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {name: i for i, name in enumerate(header) if name}

    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        task_id = clean(row[col["Task ID"]])
        if not task_id.startswith("T"):
            continue  # skips the trailing total rows

        raci = {}
        for role in ROLES:
            marking = clean(row[col[role]]).upper()
            if marking:
                raci[role] = marking

        tasks.append(
            {
                "id": task_id,
                "stage": clean(row[col["Stage"]]),
                "task": clean(row[col["Task"]]),
                "deliverable": clean(row[col["Deliverable / Exit Criteria"]]),
                "is_gate": clean(row[col["Gate?"]]) == "Yes",
                "gate_decision": clean(row[col["Gate Decision"]]),
                "days": int(row[col["Days"]] or 0),
                "critical_path": clean(row[col["Critical Path?"]]) == "Y",
                "track": clean(row[col["Track"]]),
                "status": clean(row[col["Status"]]),
                "raci": raci,
                "responsible": split_names(row[col["Responsible"]]),
                "accountable": split_names(row[col["Accountable"]]),
                "lanes": [
                    key for label, key in LANES.items() if clean(row[col[label]]) == "Y"
                ],
                "notes": clean(row[col["Notes"]]),
            }
        )

    return {
        "source": Path(workbook_path).name,
        "roles": ROLES,
        "lanes": LANES,
        "tasks": tasks,
    }


def main():
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    data = extract(workbook_path)
    OUT.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

    tasks = data["tasks"]
    gates = [t for t in tasks if t["is_gate"]]
    ceo_gates = [t for t in gates if t["accountable"] == ["CEO"]]
    print(f"{len(tasks)} tasks, {len(gates)} gates, {len(ceo_gates)} CEO gates")
    print("CEO gates:", ", ".join(t["id"] for t in ceo_gates))
    for label, key in LANES.items():
        count = sum(1 for t in tasks if key in t["lanes"])
        crit = sum(t["days"] for t in tasks if key in t["lanes"] and t["critical_path"])
        total = sum(t["days"] for t in tasks if key in t["lanes"])
        print(f"  {label:16} {count:3} tasks  crit {crit:4}d  sum {total:4}d")
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
