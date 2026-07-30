#!/usr/bin/env python3
"""Deterministic AR reconciliation matcher.

WHY THIS EXISTS
---------------
Matches bank payment lines against ERP invoices and assigns a confidence %
per match. Following this repo's convention (see screen-cv/score.py): the
LLM never computes the confidence score by hand — this script owns the
matching arithmetic so the same two input files always produce the same
result. The ar-reconciliation agent only orchestrates (locate files, call
this script, relay its JSON summary) — it must not recompute or adjust
anything this script decides.

USAGE
-----
  python match.py <invoice_file.xlsx> <payment_file.xlsx> <output_file.xlsx>

Invoice file must have columns: Invoice Type, Number, Partner, Total
Payment file must have columns: Date, Label, Amount
(header row = row 1, first sheet of each workbook)

OUTPUT LAYOUT
-------------
The output workbook has TWO sheets — keep it that way, do not merge them back into
one sheet with a divider row:
  - "Rekonsiliasi": one row per invoice (the matched/unmatched-invoice table).
  - "Pembayaran Belum Cocok": one row per payment that never got matched to an
    invoice, with its own header row. This must be a separate sheet, not appended
    below the invoice table on the same sheet.
"""
import sys
import re
import json
import itertools
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

NUMBER_RE = re.compile(r"(?<!\d)\d{4,6}(?![\d/])")
PERIOD_REF_RE = re.compile(
    r"INV[\/\-\s]+(\d{4})[\/\-\s]+(\d{1,2})[\/\-\s]+(\d{4,6})(?!\d)",
    re.IGNORECASE,
)

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def read_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        records.append(dict(zip(header, r)))
    return records


def require_columns(records, cols, label):
    if not records:
        return
    missing = [c for c in cols if c not in records[0]]
    if missing:
        raise SystemExit(
            f"File {label} tidak punya kolom yang diharapkan: {missing}. "
            f"Kolom yang ditemukan: {list(records[0].keys())}"
        )


def normalize_partner(name):
    name = str(name or "").upper()
    name = re.sub(r"\(.*?\)", " ", name)          # drop parenthetical brand names
    name = re.sub(r"\bPT\.?\b", " ", name)          # drop PT / PT.
    name = re.sub(r"[^A-Z0-9\s]", " ", name)        # drop punctuation
    tokens = [t for t in name.split() if len(t) >= 3]
    return tokens


def name_score(partner_tokens, label_upper):
    if not partner_tokens:
        return 0.0
    hits = sum(1 for t in partner_tokens if t in label_upper)
    return hits / len(partner_tokens)


def trailing_number(number_str):
    tail = str(number_str or "").split("/")[-1]
    try:
        return int(tail)
    except ValueError:
        return None


def extract_numbers(label):
    return {int(m) for m in NUMBER_RE.findall(str(label or ""))}


def find_period_refs(text):
    """Return {(year, month, number), ...} for every INV/YYYY/MM/NNNNN-style
    reference found in text. Tolerant of '/', '-', or space separators and
    1-2 digit months. Does not attempt to handle stray whitespace inside the
    year digits (rare OCR-artifact case) - falls back to no match there."""
    refs = set()
    for y, mo, n in PERIOD_REF_RE.findall(str(text or "")):
        month = int(mo)
        if 1 <= month <= 12:
            refs.add((int(y), month, int(n)))
    return refs


def round_amount(v):
    try:
        return round(float(v))
    except (TypeError, ValueError):
        return None


def main():
    if len(sys.argv) != 4:
        raise SystemExit(f"Usage: python {sys.argv[0]} <invoice.xlsx> <payment.xlsx> <output.xlsx>")
    invoice_path, payment_path, output_path = sys.argv[1:4]

    invoices = read_rows(invoice_path)
    payments = read_rows(payment_path)
    require_columns(invoices, ["Invoice Type", "Number", "Partner", "Total"], "invoice")
    require_columns(payments, ["Date", "Label", "Amount"], "payment")

    for inv in invoices:
        inv["_tokens"] = normalize_partner(inv.get("Partner"))
        inv["_trailing"] = trailing_number(inv.get("Number"))
        inv["_amount"] = round_amount(inv.get("Total"))
        inv["_period"] = next(iter(find_period_refs(inv.get("Number"))), None)

    for pay in payments:
        pay["_label_upper"] = str(pay.get("Label") or "").upper()
        pay["_amount"] = round_amount(pay.get("Amount"))
        nums = extract_numbers(pay.get("Label"))
        nums.discard(pay["_amount"])  # the amount often appears restated in the label — not an invoice ref
        pay["_numbers"] = nums
        pay["_period_refs"] = find_period_refs(pay.get("Label"))

    all_invoice_numbers = {inv["_trailing"] for inv in invoices if inv["_trailing"] is not None}

    # --- index payments for fast lookup ------------------------------------
    by_amount = {}
    by_number = {}
    for j, pay in enumerate(payments):
        if pay["_amount"] is not None:
            by_amount.setdefault(pay["_amount"], []).append(j)
        for n in pay["_numbers"]:
            by_number.setdefault(n, []).append(j)

    # --- Pass A: candidates via invoice-number / exact-amount indexes ------
    candidates = []  # (invoice_idx, payment_idx, confidence, basis)
    for i, inv in enumerate(invoices):
        seen = set()
        cand_js = set()
        if inv["_trailing"] is not None:
            cand_js.update(by_number.get(inv["_trailing"], []))
        if inv["_amount"] is not None:
            cand_js.update(by_amount.get(inv["_amount"], []))
        for j in cand_js:
            pay = payments[j]
            trailing = inv["_trailing"]
            inv_num_match = trailing is not None and trailing in pay["_numbers"]

            period_note = ""
            if inv_num_match and inv["_period"] is not None and pay["_period_refs"]:
                matching_number_refs = [r for r in pay["_period_refs"] if r[2] == trailing]
                if matching_number_refs:
                    inv_y, inv_mo, _ = inv["_period"]
                    if all((r[0], r[1]) != (inv_y, inv_mo) for r in matching_number_refs):
                        # every explicit ref for this number points to a different
                        # month/year - the digit match is coincidental
                        inv_num_match = False
                    else:
                        period_note = " (bulan & tahun cocok)"

            amt_match = (
                inv["_amount"] is not None
                and pay["_amount"] is not None
                and abs(inv["_amount"] - pay["_amount"]) <= 1
            )
            nsc = name_score(inv["_tokens"], pay["_label_upper"])

            if inv_num_match and amt_match:
                conf, basis = 100, "Nomor invoice & jumlah cocok" + period_note
            elif inv_num_match:
                conf, basis = 90, "Nomor invoice cocok, jumlah berbeda" + period_note
            elif amt_match and nsc >= 0.6:
                conf, basis = 75, "Nama partner & jumlah cocok"
            elif amt_match:
                conf, basis = 50, "Hanya jumlah cocok"
            else:
                continue
            candidates.append((i, j, conf, basis))
            seen.add(j)

    matched_invoice_idx = {c[0] for c in candidates}
    unmatched_invoice_idx = [i for i in range(len(invoices)) if i not in matched_invoice_idx]

    # --- Pass B: name-only match (tier 35) for invoices still unmatched ----
    payment_idx_in_candidates = {c[1] for c in candidates}
    free_payment_idx = [j for j in range(len(payments)) if j not in payment_idx_in_candidates]
    scan_size = len(unmatched_invoice_idx) * len(free_payment_idx)
    name_only_scan_skipped = False
    if scan_size <= 3_000_000:
        for i in unmatched_invoice_idx:
            inv = invoices[i]
            if len(inv["_tokens"]) < 2:
                continue
            for j in free_payment_idx:
                pay = payments[j]
                nsc = name_score(inv["_tokens"], pay["_label_upper"])
                if nsc >= 0.8:
                    candidates.append((i, j, 35, "Kemungkinan pembayaran sebagian"))
    else:
        name_only_scan_skipped = True

    # --- Greedy assignment: highest confidence first, one payment per invoice, one invoice per payment
    candidates.sort(key=lambda c: -c[2])
    assigned_invoice = {}
    used_payment = set()
    for i, j, conf, basis in candidates:
        if i in assigned_invoice or j in used_payment:
            continue
        assigned_invoice[i] = (j, conf, basis)
        used_payment.add(j)

    # --- Bundle detection (informational only) for invoices still unmatched
    notes = {}
    still_unmatched = [i for i in range(len(invoices)) if i not in assigned_invoice]
    by_partner = {}
    for i in still_unmatched:
        by_partner.setdefault(tuple(invoices[i]["_tokens"]), []).append(i)

    for tokens, idxs in by_partner.items():
        if len(idxs) < 2:
            continue
        if len(idxs) > 10:
            for i in idxs:
                notes[i] = "Kelompok terlalu besar untuk deteksi bundel otomatis"
            continue
        candidate_payments = [
            j for j in range(len(payments))
            if j not in used_payment and name_score(list(tokens), payments[j]["_label_upper"]) >= 0.4
        ][:5]
        if not candidate_payments:
            continue
        found_for = set()
        for j in candidate_payments:
            pay_amt = payments[j]["_amount"]
            if pay_amt is None:
                continue
            for size in range(2, min(5, len(idxs)) + 1):
                for combo in itertools.combinations(idxs, size):
                    if any(c in found_for for c in combo):
                        continue
                    total = sum(invoices[c]["_amount"] or 0 for c in combo)
                    if abs(total - pay_amt) <= 1:
                        combo_numbers = ", ".join(str(invoices[c].get("Number")) for c in combo)
                        note = (
                            f"Kemungkinan bagian dari pembayaran gabungan tanggal "
                            f"{payments[j].get('Date')} sebesar {pay_amt} ({payments[j].get('Label')}) "
                            f"bersama invoice: {combo_numbers}"
                        )
                        for c in combo:
                            notes[c] = note
                            found_for.add(c)

    # --- Classify remaining unmatched payments ------------------------------
    unmatched_payment_rows = []
    other_period_count = 0
    for j, pay in enumerate(payments):
        if j in used_payment:
            continue
        nums = pay["_numbers"]
        if nums and not (nums & all_invoice_numbers):
            if pay["_period_refs"]:
                y, mo, n = sorted(pay["_period_refs"])[0]
                reason = f"Dibayar - periode lain (terdeteksi: {mo:02d}/{y} No. {n})"
            else:
                reason = f"Dibayar - periode lain (terdeteksi: {sorted(nums)[0]})"
            other_period_count += 1
        else:
            reason = "Tidak ada indikasi kecocokan"
        unmatched_payment_rows.append((pay, reason))

    # --- Write output workbook ----------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekonsiliasi"

    headers = [
        "Tipe Invoice", "Nomor Invoice", "Partner", "Total Invoice",
        "Tanggal Pembayaran", "Label Pembayaran", "Jumlah Dibayar",
        "Tingkat Keyakinan (%)", "Dasar Kecocokan", "Catatan Tambahan",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    rows_out = []
    for i, inv in enumerate(invoices):
        if i in assigned_invoice:
            j, conf, basis = assigned_invoice[i]
            pay = payments[j]
            rows_out.append((
                inv.get("Invoice Type"), inv.get("Number"), inv.get("Partner"), inv.get("Total"),
                pay.get("Date"), pay.get("Label"), pay.get("Amount"),
                conf, basis, notes.get(i, ""),
            ))
        else:
            rows_out.append((
                inv.get("Invoice Type"), inv.get("Number"), inv.get("Partner"), inv.get("Total"),
                None, None, None,
                0, "Belum cocok", notes.get(i, ""),
            ))

    rows_out.sort(key=lambda r: r[7])  # confidence ascending: needs-attention first

    for row in rows_out:
        ws.append(row)
        r = ws.max_row
        conf = row[7]
        fill = FILL_RED if conf == 0 else FILL_ORANGE if conf < 50 else FILL_YELLOW if conf < 90 else FILL_GREEN
        ws.cell(row=r, column=8).fill = fill

    for col_idx, header in enumerate(headers, start=1):
        width = max(14, min(50, len(header) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # unmatched payments go on their own sheet, never appended below the main table
    ws_unmatched = wb.create_sheet("Pembayaran Belum Cocok")
    unmatched_headers = ["Tanggal", "Label", "Jumlah", "Kemungkinan Alasan"]
    ws_unmatched.append(unmatched_headers)
    for cell in ws_unmatched[1]:
        cell.font = Font(bold=True)
    ws_unmatched.freeze_panes = "A2"

    for pay, reason in unmatched_payment_rows:
        ws_unmatched.append([pay.get("Date"), pay.get("Label"), pay.get("Amount"), reason])

    for col_idx, header in enumerate(unmatched_headers, start=1):
        width = max(14, min(60, len(header) + 4))
        ws_unmatched.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)

    # --- Summary -------------------------------------------------------------
    conf_counts = {}
    for _, (_, conf, _) in assigned_invoice.items():
        conf_counts[conf] = conf_counts.get(conf, 0) + 1

    summary = {
        "total_invoices": len(invoices),
        "total_payments": len(payments),
        "matched_invoices": len(assigned_invoice),
        "unmatched_invoices": len(invoices) - len(assigned_invoice),
        "matched_by_confidence": conf_counts,
        "unmatched_payments": len(unmatched_payment_rows),
        "unmatched_payments_other_period": other_period_count,
        "bundle_notes_added": len(notes),
        "name_only_scan_skipped": name_only_scan_skipped,
        "output_file": output_path,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
