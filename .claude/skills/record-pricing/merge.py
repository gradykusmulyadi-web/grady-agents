#!/usr/bin/env python3
"""Deterministic merge of an extracted vendor-quote JSON into the vendor pricing master workbook.

Usage:
    python merge.py <extraction.json> <master_workbook.xlsx> [--date YYYY-MM-DD]

Prints a JSON summary to stdout. Never guesses at ambiguous input -- raises on
bad category names, missing required fields, etc. rather than silently coercing.
"""

import argparse
import datetime
import json
import os
import sys
import time

import openpyxl
from openpyxl.styles import Font

CATEGORY_TABS = ["GPS", "Dashcam", "MDVR", "Other Sensors", "Memory Cards", "Software & Services"]
CURRENCY_SHEET = "Currency Master"

HEADER = [
    "Vendor Name",
    "Model/SKU",
    "Product Name",
    "Price",
    "Currency",
    "Standardized Price (USD)",
    "Date Added",
    "Latest?",
    "Peripherals?",
    "Description",
    "Notes",
]
CURRENCY_HEADER = ["Currency Code", "Rate to USD", "Last Updated"]

# Seeded from a web search performed 2026-08-20. Grady edits this tab by hand afterwards;
# this script only reads it once the sheet already exists.
SEED_RATES = {
    "USD": 1.0,
    "CNY": 0.1486,
    "EUR": 1.1580,
    "IDR": 0.00005575,
    "SGD": 0.7824,
}

COL = {name: i + 1 for i, name in enumerate(HEADER)}


def create_workbook(path, today):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in CATEGORY_TABS:
        ws = wb.create_sheet(name)
        ws.append(HEADER)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
    cur = wb.create_sheet(CURRENCY_SHEET)
    cur.append(CURRENCY_HEADER)
    for cell in cur[1]:
        cell.font = Font(bold=True)
    for code, rate in SEED_RATES.items():
        cur.append([code, rate, today.isoformat()])
    cur.freeze_panes = "A2"
    wb.save(path)
    return wb


def load_or_create(path, today):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        missing = [n for n in CATEGORY_TABS + [CURRENCY_SHEET] if n not in wb.sheetnames]
        if missing:
            raise ValueError(
                f"Workbook {path} exists but is missing expected tab(s): {missing}. "
                "Not touching it automatically -- fix the workbook or point at a fresh path."
            )
        return wb
    return create_workbook(path, today)


def load_currency_rates(wb):
    ws = wb[CURRENCY_SHEET]
    rates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rates[str(row[0]).strip().upper()] = float(row[1])
    return rates


def ensure_currency(wb, rates, code, today, flags):
    code = code.strip().upper()
    if code in rates:
        return rates[code]
    ws = wb[CURRENCY_SHEET]
    ws.append([code, 1.0, today.isoformat()])
    rates[code] = 1.0
    flags.append(code)
    return 1.0


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def build_latest_index(ws):
    """Map (vendor_key, sku_or_name_key) -> row number, for rows currently marked Latest=Yes."""
    index = {}
    for row_idx in range(2, ws.max_row + 1):
        vendor = ws.cell(row_idx, COL["Vendor Name"]).value
        if vendor is None:
            continue
        latest = ws.cell(row_idx, COL["Latest?"]).value
        if norm(latest) != "yes":
            continue
        sku = ws.cell(row_idx, COL["Model/SKU"]).value
        name = ws.cell(row_idx, COL["Product Name"]).value
        key_part = norm(sku) if sku else norm(name)
        index[(norm(vendor), key_part)] = row_idx
    return index


def merge(extraction_path, workbook_path, date_override):
    today = date_override or datetime.date.today()

    with open(extraction_path, "r", encoding="utf-8") as f:
        extraction = json.load(f)

    vendor = extraction["vendor"]
    items = extraction["items"]

    wb = load_or_create(workbook_path, today)
    rates = load_currency_rates(wb)

    currency_flags = []
    summary_tabs = {}
    superseded = []

    for item in items:
        category = item["category"]
        if category not in CATEGORY_TABS:
            raise ValueError(
                f"Item {item.get('product_name')!r} has invalid category {category!r}. "
                f"Must be one of {CATEGORY_TABS}."
            )
        ws = wb[category]
        index = build_latest_index(ws)

        model_sku = (item.get("model_sku") or "").strip()
        product_name = item["product_name"].strip()
        price = float(item["price"])
        currency = item["currency"].strip().upper()
        is_peripheral = bool(item["is_peripheral"])
        description = item.get("description", "")
        notes = item.get("notes", "")

        rate = ensure_currency(wb, rates, currency, today, currency_flags)
        std_price = round(price * rate, 4)

        key_part = norm(model_sku) if model_sku else norm(product_name)
        key = (norm(vendor), key_part)
        existing_row = index.get(key)

        tab_summary = summary_tabs.setdefault(category, {"new": 0, "superseded": 0, "skipped_duplicate": 0})

        if existing_row:
            ex_price = ws.cell(existing_row, COL["Price"]).value
            ex_currency = ws.cell(existing_row, COL["Currency"]).value
            if ex_price is not None and float(ex_price) == price and norm(ex_currency) == norm(currency):
                tab_summary["skipped_duplicate"] += 1
                continue
            ws.cell(existing_row, COL["Latest?"]).value = "No"
            superseded.append(
                {
                    "category": category,
                    "vendor": vendor,
                    "product_name": ws.cell(existing_row, COL["Product Name"]).value,
                    "old_price": ex_price,
                    "old_currency": ex_currency,
                    "new_price": price,
                    "new_currency": currency,
                }
            )
            tab_summary["superseded"] += 1
        else:
            tab_summary["new"] += 1

        ws.append(
            [
                vendor,
                model_sku,
                product_name,
                price,
                currency,
                std_price,
                today.isoformat(),
                "Yes",
                "Yes" if is_peripheral else "No",
                description,
                notes,
            ]
        )

    save_with_retry(wb, workbook_path)

    return {
        "workbook_path": os.path.abspath(workbook_path),
        "vendor": vendor,
        "tabs": summary_tabs,
        "superseded": superseded,
        "currency_flags": currency_flags,
    }


def save_with_retry(wb, path, attempts=5, delay=1.0):
    tmp_path = path + ".tmp"
    wb.save(tmp_path)
    last_err = None
    for _ in range(attempts):
        try:
            os.replace(tmp_path, path)
            return
        except PermissionError as e:
            last_err = e
            time.sleep(delay)
    raise PermissionError(
        f"Could not write {path} -- it looks like it's open/locked (e.g. in Excel via OneDrive sync). "
        f"Close the file and re-run. Underlying error: {last_err}"
    )


def parse_date(s):
    if not s:
        return None
    return datetime.date.fromisoformat(s)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("extraction_json")
    parser.add_argument("workbook_path")
    parser.add_argument("--date", default=None, help="Override Date Added (YYYY-MM-DD), defaults to today")
    args = parser.parse_args()

    result = merge(args.extraction_json, args.workbook_path, parse_date(args.date))
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
