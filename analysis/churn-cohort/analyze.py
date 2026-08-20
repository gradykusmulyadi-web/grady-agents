"""Churn cohort analysis -> segmented diagnosis workbook.

Reads the `Churn Raw` tab of a McEasy churn extract and emits a multi-tab
analysis workbook plus a reconciliation report on stdout.

Usage:
    python analyze.py <source.xlsx> <output.xlsx>

Notes:
  - `total_churn` is ANNUALIZED ARR in IDR (median ~1.08M per vehicle/year).
  - Row grain is one work-order product line, not one customer.
  - This script computes SHARE OF CHURN (mix) only. The source has no active
    customer base, so segment churn RATES cannot be derived. See 10_Data_Quality.
  - Console here is cp1252; keep stdout ASCII-only.
"""

from __future__ import annotations

import collections
import datetime as dt
import shutil
import statistics
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

SHEET = "Churn Raw"

SIZE_ORDER = ["A<=5", "B<=10", "C<=20", "D<=50", "E<=100", "F<400", "G>=400", "(unassigned)"]

REASON_BUCKETS = [
    "Payment/Credit",
    "Renewal Decision",
    "Fleet Event",
    "Value/Fit",
    "Product/Tech Failure",
    "Trial/Never Landed",
]

# Raw reason -> actionable bucket. Every one of the 22 observed reasons is mapped
# explicitly; an unmapped reason raises rather than silently landing in "Other".
REASON_MAP = {
    "Kendala Pembayaran - Customer Tidak Mau Bayar/ Menunggak": "Payment/Credit",
    "Kendala Pembayaran - Efisiensi Biaya": "Payment/Credit",
    "Tidak Ingin Melanjutkan Renewal": "Renewal Decision",
    "Tidak Ingin Harga Naik Saat Renewal": "Renewal Decision",
    "Menggunakan Produk Kompetitor - Lebih Murah": "Renewal Decision",
    "Unit Dijual": "Fleet Event",
    "Kondisi Kendaraan / Tidak Layak Operasi / Kecelakaan": "Fleet Event",
    "Pengembalian Ke Vendor / Ditarik Leasing": "Fleet Event",
    "Kendaraan Tidak Sedang Dipakai Dalam Waktu Lama": "Fleet Event",
    "Device Hilang - Gps Hilang": "Fleet Event",
    "Device Hilang - Unit (Fleet) Hilang, Kecewa Gps Tidak Dapat Membantu": "Fleet Event",
    "Kontrak Dengan Vendor Sudah Habis": "Fleet Event",
    "Tidak Ada Kebutuhan Lagi": "Value/Fit",
    "Produk Tidak Sesuai Dengan Kebutuhan - Tidak Efisien/Tidak Ada Benefit": "Value/Fit",
    "Permintaan Customer Yang Belum Dapat Dihandle Oleh Product": "Value/Fit",
    "Customer Tidak Ingin Menggunakan Aplikasi V2": "Value/Fit",
    "Device Bermasalah - Offline Terlalu Lama Tidak Segera Dilakukan Maintenance/"
    "Offline Berkali-Kali Setelah Maintenance": "Product/Tech Failure",
    "Device Atau Sensor Tidak Kompatible": "Product/Tech Failure",
    "Kesalahan Teknis / Kesalahan Pemasangan": "Product/Tech Failure",
    "Layanan Pelanggan Yang Buruk": "Product/Tech Failure",
    "Tidak Lanjut Trial": "Trial/Never Landed",
    "Hard Contact": "Trial/Never Landed",
}

# Ordered keyword rules: first match wins. Collapses ~87 product_name values.
PRODUCT_RULES = [
    ("trackvision", "TrackVision (Dashcam)"),
    ("track vision", "TrackVision (Dashcam)"),
    ("mdvr", "TrackVision (Dashcam)"),
    ("dashcam", "TrackVision (Dashcam)"),
    ("tms", "TMS"),
    ("delivery management", "TMS"),
    ("gps package", "VSMS GPS Rental"),
    ("custom dashboard", "Software-only"),
    ("quick assignment", "Software-only"),
    ("icall", "Other Sensor"),
    ("ibuzzer", "Other Sensor"),
    ("sos button", "Other Sensor"),
    ("tilt", "Other Sensor"),
    ("ifuel", "iFuel (Fuel Sensor)"),
    ("itemp", "iTemp (Temp Sensor)"),
    ("idoor", "iDoor (Door Sensor)"),
    ("temperature", "iTemp (Temp Sensor)"),
    ("enterprise", "VSMS Enterprise"),
    ("vsms gps rental", "VSMS GPS Rental"),
    ("vsms rental software", "Software-only"),
    ("software + sim", "Software+SIM"),
    ("starter", "McEasy STARTER"),
    ("plus", "McEasy PLUS"),
    ("pro -", "McEasy PRO"),
    ("pto", "Other Sensor"),
    ("relay", "Other Sensor"),
    ("rfid", "Other Sensor"),
    ("ble", "Other Sensor"),
    ("sensor", "Other Sensor"),
    ("vsms", "VSMS Other"),
]

TENURE_BUCKETS = [(0, 6), (6, 12), (12, 18), (18, 24), (24, 36), (36, 10**6)]

# Outlier accounts stripped from tab 11 ONLY. Every other tab stays inclusive so
# the Summary tie-out continues to hold. Corin Mulia Gemilang is IDR 2.42B /
# 1,346 vehicles of partial churn -- 17.8% of the file on its own.
TENURE_DIST_EXCLUDE = ["Corin Mulia Gemilang, PT"]

# Month bins above this are collapsed into one tail row.
TENURE_TAIL = 48

# 3-month windows centred on each contract anniversary, vs the months between.
ANNIV_WINDOWS = [
    ("Year 1 anniversary", 11, 14),
    ("Year 2 anniversary", 23, 26),
    ("Year 3 anniversary", 35, 38),
]
GAP_WINDOWS = [
    ("Between: mo 3-10", 3, 11),
    ("Between: mo 14-22", 14, 23),
    ("Between: mo 26-34", 26, 35),
]

# Reconciliation target from the source workbook's own Summary tab:
# State Cohort Jun'25 == Include, churn months Jul-2025..May-2026.
SUMMARY_TIEOUT = 6_922_356_431.04
SUMMARY_WINDOW = ("2025-07", "2026-05")

# --------------------------------------------------------------------------
# Styling
# --------------------------------------------------------------------------

H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="808080")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
IDR = "#,##0"
PCT = "0.0%"
NUM1 = "#,##0.0"


def title(ws, row, text, note=None):
    ws.cell(row, 1, text).font = TITLE_FONT
    row += 1
    if note:
        ws.cell(row, 1, note).font = NOTE_FONT
        row += 1
    return row + 1


def header(ws, row, labels, start_col=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row, start_col + i, lab)
        c.fill, c.font, c.border = H_FILL, H_FONT, BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return row + 1


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# --------------------------------------------------------------------------
# Load + enrich
# --------------------------------------------------------------------------


def parse_start(value):
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return dt.datetime.strptime(value.strip().split()[0], "%d/%m/%Y")
        except ValueError:
            return None
    return None


def product_family(name):
    low = (name or "").lower()
    for key, fam in PRODUCT_RULES:
        if key in low:
            return fam
    return "Unmapped"


def load_rows(src: Path):
    """Read the source through a temp copy so the original is never touched."""
    tmp = Path(tempfile.gettempdir()) / "_churn_src_readonly.xlsx"
    shutil.copy2(src, tmp)
    wb = openpyxl.load_workbook(tmp, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit("ERROR: sheet %r not found. Found: %s" % (SHEET, wb.sheetnames))
    ws = wb[SHEET]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        rec = dict(zip(hdr, vals))
        if not rec.get("wo_name"):
            continue
        rows.append(rec)
    wb.close()
    tmp.unlink(missing_ok=True)

    unknown = {r["reason_churn"] for r in rows} - set(REASON_MAP)
    if unknown:
        raise SystemExit("ERROR: unmapped reason_churn value(s): %s" % sorted(unknown))

    for d in rows:
        d["_amt"] = float(d["total_churn"] or 0)
        d["_veh"] = int(d["total_vehicle"] or 0)
        d["_start"] = parse_start(d["subscription_start_date"])
        d["_churn"] = d["wo_accounting_date"]
        d["_tenure"] = (
            (d["_churn"] - d["_start"]).days / 30.44 if d["_start"] and d["_churn"] else None
        )
        d["_churn_my"] = d["_churn"].strftime("%Y-%m") if d["_churn"] else None
        d["_fam"] = product_family(d["product_name"])
        d["_bucket"] = REASON_MAP[d["reason_churn"]]
        d["_raw_band"] = d["fleet_category"] or "(unassigned)"
        coh = d["Cohort Month - Year"]
        d["_cohort_my"] = coh.strftime("%Y-%m") if isinstance(coh, dt.datetime) else None
    return rows


def build_customers(rows):
    """Customer-level rollup. Size band resolves to the largest-ARR line."""
    by_cust = collections.defaultdict(list)
    for d in rows:
        by_cust[d["customer_id"]].append(d)

    customers = {}
    for cid, lines in by_cust.items():
        ranked = sorted(lines, key=lambda x: (-x["_amt"], -x["_veh"]))
        anchor = ranked[0]
        amt = sum(x["_amt"] for x in lines)
        veh = sum(x["_veh"] for x in lines)
        fam_amt = collections.Counter()
        bkt_amt = collections.Counter()
        for x in lines:
            fam_amt[x["_fam"]] += x["_amt"] or 0.001
            bkt_amt[x["_bucket"]] += x["_amt"] or 0.001
        tenures = [x["_tenure"] for x in lines if x["_tenure"] is not None]
        customers[cid] = {
            "customer_id": cid,
            "name": anchor["customer_wo"],
            "band": anchor["_raw_band"],
            "bands_seen": sorted({x["_raw_band"] for x in lines}),
            "status": anchor["churn_status"],
            "amount": amt,
            "vehicles": veh,
            "lines": len(lines),
            "family": fam_amt.most_common(1)[0][0],
            "families_seen": sorted({x["_fam"] for x in lines}),
            "bucket": bkt_amt.most_common(1)[0][0],
            "buckets_seen": sorted({x["_bucket"] for x in lines}),
            "reason": anchor["reason_churn"],
            "sales": anchor["sales_name"] or "(none)",
            "tenure": anchor["_tenure"],
            "tenure_min": min(tenures) if tenures else None,
            "first_churn": min(x["_churn"] for x in lines),
            "last_churn": max(x["_churn"] for x in lines),
        }
    # Propagate the resolved band back onto lines so amount/vehicle crosstabs
    # keep every one of a customer's lines inside a single size band.
    for d in rows:
        d["_band"] = customers[d["customer_id"]]["band"]
    return customers


# --------------------------------------------------------------------------
# Aggregation helpers
# --------------------------------------------------------------------------


def agg(records, keyfn):
    out = collections.defaultdict(lambda: {"amt": 0.0, "veh": 0, "cust": set(), "n": 0})
    for d in records:
        e = out[keyfn(d)]
        e["amt"] += d["_amt"]
        e["veh"] += d["_veh"]
        e["cust"].add(d["customer_id"])
        e["n"] += 1
    return out


def tenure_label(months):
    if months is None:
        return "(no start date)"
    for lo, hi in TENURE_BUCKETS:
        if lo <= months < hi:
            return "36+ mo" if hi == 10**6 else "%d-%d mo" % (lo, hi)
    return "(no start date)"


TENURE_ORDER = ["0-6 mo", "6-12 mo", "12-18 mo", "18-24 mo", "24-36 mo", "36+ mo", "(no start date)"]


def write_matrix(ws, row, records, row_keys, row_label, col_keys, col_label, measure, fmt=IDR):
    """One measure as a rows x cols matrix with margins. Returns next free row."""
    cells = agg(records, lambda d: (d["_rowk"], d["_colk"]))
    ws.cell(row, 1, "%s  (%s)" % (measure["title"], row_label + " x " + col_label)).font = BOLD
    row += 1
    row = header(ws, row, [row_label] + list(col_keys) + ["TOTAL"])
    grand = 0
    for rk in row_keys:
        ws.cell(row, 1, rk).font = BOLD
        ws.cell(row, 1).border = BOX
        rtot = 0
        for i, ck in enumerate(col_keys):
            e = cells.get((rk, ck))
            v = measure["get"](e) if e else 0
            c = ws.cell(row, 2 + i, v if v else None)
            c.number_format, c.border = fmt, BOX
            rtot += v
        c = ws.cell(row, 2 + len(col_keys), rtot or None)
        c.number_format, c.font, c.fill, c.border = fmt, BOLD, TOTAL_FILL, BOX
        grand += rtot
        row += 1
    ws.cell(row, 1, "TOTAL").font = BOLD
    ws.cell(row, 1).fill = TOTAL_FILL
    for i, ck in enumerate(col_keys):
        ctot = sum(measure["get"](e) for (rk, c2), e in cells.items() if c2 == ck and rk in row_keys)
        c = ws.cell(row, 2 + i, ctot or None)
        c.number_format, c.font, c.fill, c.border = fmt, BOLD, TOTAL_FILL, BOX
    c = ws.cell(row, 2 + len(col_keys), grand or None)
    c.number_format, c.font, c.fill, c.border = fmt, BOLD, TOTAL_FILL, BOX
    return row + 2, grand


M_AMT = {"title": "Churned ARR (IDR)", "get": lambda e: e["amt"]}
M_VEH = {"title": "Vehicles", "get": lambda e: e["veh"]}
M_CUST = {"title": "Distinct customers (margins NOT additive)", "get": lambda e: len(e["cust"])}


def matrix_block(ws, row, records, rowkey, row_keys, row_label, colkey, col_keys, col_label,
                 total_amt, total_veh, label):
    for d in records:
        d["_rowk"], d["_colk"] = rowkey(d), colkey(d)
    row, got_amt = write_matrix(ws, row, records, row_keys, row_label, col_keys, col_label, M_AMT)
    assert abs(got_amt - total_amt) < 1, "%s: ARR margin %.2f != %.2f" % (label, got_amt, total_amt)
    row, got_veh = write_matrix(ws, row, records, row_keys, row_label, col_keys, col_label,
                                M_VEH, IDR)
    assert got_veh == total_veh, "%s: vehicle margin %d != %d" % (label, got_veh, total_veh)
    row, _ = write_matrix(ws, row, records, row_keys, row_label, col_keys, col_label, M_CUST, IDR)
    ws.cell(row - 1, 1, "Note: a customer can appear in several cells, so customer "
                        "margins exceed the true distinct total.").font = NOTE_FONT
    return row + 1


# --------------------------------------------------------------------------
# Tabs
# --------------------------------------------------------------------------


def tab_readme(wb, ctx):
    ws = wb.create_sheet("00_Read_Me")
    widths(ws, {"A": 42, "B": 96})
    r = title(ws, 1, "Churn Cohort Analysis -- Read Me")
    facts = [
        ("Source file", ctx["src_name"]),
        ("Source tab", SHEET),
        ("Generated", ctx["rundate"]),
        ("", ""),
        ("SCOPE", ""),
        ("Churn window", "%s to %s (all rows, no cohort-flag filter)" % (ctx["min_my"], ctx["max_my"])),
        ("Rows (WO product lines)", ctx["n_rows"]),
        ("Work orders", ctx["n_wo"]),
        ("Customers", ctx["n_cust"]),
        ("Churned ARR (IDR)", ctx["total_amt"]),
        ("Vehicles", ctx["total_veh"]),
        ("", ""),
        ("DEFINITIONS", ""),
        ("total_churn", "ANNUALIZED ARR in IDR, not monthly. Median IDR 1,080,000 per "
                        "vehicle/year (~90,000 per vehicle/month)."),
        ("Row grain", "One work-order product line. 1 customer can hold several lines "
                      "across several WOs. Customer counts always come from the "
                      "customer-level rollup, never from row counts."),
        ("churn_status", "Verified consistent per customer -- no customer holds both "
                         "full_churn and partial_churn lines. Treated as a "
                         "customer-level label."),
        ("Size band", "fleet_category. 19 customers carry more than one band; each is "
                      "resolved to the band of its largest-ARR line so a customer's "
                      "lines never split across bands. See 10_Data_Quality."),
        ("Tenure", "wo_accounting_date minus subscription_start_date, in months "
                   "(days / 30.44). Customer-level tenure uses the largest-ARR line."),
        ("Reason bucket", "22 raw reasons collapsed into 6 buckets. Raw reason is kept "
                          "in 09_Target_List. Judgment call: 'Kendala Pembayaran - "
                          "Efisiensi Biaya' sits in Payment/Credit but is arguably "
                          "Value/Fit -- it is shown separately in 02_Size_x_Reason."),
        ("Product family", "87 product_name values collapsed into families by keyword "
                           "rule. Full mapping in 10_Data_Quality."),
        ("", ""),
        ("READ THIS BEFORE ACTING", ""),
        ("MIX, NOT RATE", "This file contains churned customers only -- there is no "
                          "active customer base. Every segment figure here is SHARE OF "
                          "CHURN, not churn rate. A<=5 shows the most churned customers "
                          "because it is the most populous band, not because it is the "
                          "riskiest. Do not size an AM program on these shares until "
                          "the base data in 10_Data_Quality is supplied."),
        ("SMALL-n WARNING", "G>=400 full churn is 4 customers, and one of them "
                            "(Koperasi Wahana Kalpika, IDR 755M) is 85% of that band's "
                            "ARR. Any 'large fleets behave differently' reading rests on "
                            "a single account."),
        ("REASON DATA IS SELF-REPORTED", "reason_churn is entered by sales. 'Unit "
                                         "Dijual' and arrears labels are both comfortable "
                                         "answers that can mask dissatisfaction. Treat "
                                         "the reason mix as a hypothesis generator."),
        ("INCOMPLETE MONTH", "Jun-2026 (%d rows, IDR %s) falls outside the window the "
                             "source Summary tab reports on and may be partial."
                             % (ctx["jun26_n"], format(int(ctx["jun26_amt"]), ","))),
    ]
    for k, v in facts:
        if k in ("SCOPE", "DEFINITIONS", "READ THIS BEFORE ACTING"):
            c = ws.cell(r, 1, k)
            c.font, c.fill = BOLD, SUB_FILL
            ws.cell(r, 2).fill = SUB_FILL
        elif k:
            ws.cell(r, 1, k).font = BOLD
            c = ws.cell(r, 2, v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if isinstance(v, (int, float)):
                c.number_format = IDR
            if k in ("MIX, NOT RATE", "SMALL-n WARNING", "REASON DATA IS SELF-REPORTED",
                     "INCOMPLETE MONTH"):
                ws.cell(r, 1).fill = WARN_FILL
                c.fill = WARN_FILL
        r += 1


def tab_headline(wb, rows, customers, ctx):
    ws = wb.create_sheet("01_Headline")
    widths(ws, {"A": 34, "B": 16, "C": 18, "D": 14, "E": 16, "F": 16, "G": 16})
    r = title(ws, 1, "Headline: full vs partial churn",
              "ARR is annualized IDR. Shares are of total churn in the window.")
    r = header(ws, r, ["Segment", "Customers", "Churned ARR", "% ARR", "Vehicles",
                       "% Vehicles", "ARR / customer", "Vehicles / customer",
                       "ARR / vehicle"])
    for st in ["full_churn", "partial_churn"]:
        cs = [c for c in customers.values() if c["status"] == st]
        amt = sum(c["amount"] for c in cs)
        veh = sum(c["vehicles"] for c in cs)
        vals = [st, len(cs), amt, amt / ctx["total_amt"], veh, veh / ctx["total_veh"],
                amt / len(cs), veh / len(cs), amt / veh if veh else 0]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v)
            c.border = BOX
            c.number_format = PCT if i in (3, 5) else (NUM1 if i == 7 else IDR)
        r += 1
    vals = ["TOTAL", len(customers), ctx["total_amt"], 1.0, ctx["total_veh"], 1.0,
            ctx["total_amt"] / len(customers), ctx["total_veh"] / len(customers),
            ctx["total_amt"] / ctx["total_veh"]]
    for i, v in enumerate(vals):
        c = ws.cell(r, 1 + i, v)
        c.font, c.fill, c.border = BOLD, TOTAL_FILL, BOX
        c.number_format = PCT if i in (3, 5) else (NUM1 if i == 7 else IDR)
    r += 3

    r = title(ws, r, "Zero-ARR churn (invisible in every amount-weighted view)",
              "These lines carry IDR 0, so they disappear from ARR cuts. Mostly trials "
              "that never converted -- an onboarding problem, not an AM one.")
    zero = [d for d in rows if d["_amt"] == 0]
    r = header(ws, r, ["Reason (raw)", "Lines", "Customers", "Vehicles"])
    zagg = agg(zero, lambda d: d["reason_churn"])
    for k in sorted(zagg, key=lambda k: -zagg[k]["n"]):
        e = zagg[k]
        for i, v in enumerate([k, e["n"], len(e["cust"]), e["veh"]]):
            c = ws.cell(r, 1 + i, v)
            c.border = BOX
            if i:
                c.number_format = IDR
        r += 1
    e = agg(zero, lambda d: "all")["all"]
    for i, v in enumerate(["TOTAL zero-ARR", e["n"], len(e["cust"]), e["veh"]]):
        c = ws.cell(r, 1 + i, v)
        c.font, c.fill, c.border = BOLD, TOTAL_FILL, BOX
        if i:
            c.number_format = IDR
    r += 3

    r = title(ws, r, "Churn by month (both statuses)", "Watch the last month for "
              "completeness before reading a trend into it.")
    r = header(ws, r, ["Churn month", "Lines", "Customers (first churn)", "Churned ARR",
                       "Vehicles"])
    magg = agg(rows, lambda d: d["_churn_my"])
    for k in sorted(magg):
        e = magg[k]
        for i, v in enumerate([k, e["n"], len(e["cust"]), e["amt"], e["veh"]]):
            c = ws.cell(r, 1 + i, v)
            c.border = BOX
            if i:
                c.number_format = IDR
        r += 1


def tab_size_reason(wb, full, ctx):
    ws = wb.create_sheet("02_Size_x_Reason")
    widths(ws, {"A": 22, **{get_column_letter(i): 15 for i in range(2, 12)}})
    r = title(ws, 1, "FULL CHURN: customer size band x reason bucket",
              "ARR and vehicle margins are exact. MIX ONLY -- these are shares of "
              "churn, not churn rates.")
    bands = [b for b in SIZE_ORDER if any(d["_band"] == b for d in full)]
    r = matrix_block(ws, r, full, lambda d: d["_band"], bands, "Size band",
                     lambda d: d["_bucket"], REASON_BUCKETS, "Reason bucket",
                     ctx["full_amt"], ctx["full_veh"], "02 size x bucket")

    r = title(ws, r, "Same view, % of each size band's churned ARR",
              "Read across a row: what kills this segment.")
    cells = agg(full, lambda d: (d["_band"], d["_bucket"]))
    r = header(ws, r, ["Size band"] + REASON_BUCKETS + ["Band ARR"])
    for b in bands:
        btot = sum(d["_amt"] for d in full if d["_band"] == b)
        ws.cell(r, 1, b).font = BOLD
        ws.cell(r, 1).border = BOX
        for i, bk in enumerate(REASON_BUCKETS):
            e = cells.get((b, bk))
            v = (e["amt"] / btot) if e and btot else 0
            c = ws.cell(r, 2 + i, v if v else None)
            c.number_format, c.border = PCT, BOX
            if v >= 0.5:
                c.fill = WARN_FILL
        c = ws.cell(r, 2 + len(REASON_BUCKETS), btot)
        c.number_format, c.font, c.fill, c.border = IDR, BOLD, TOTAL_FILL, BOX
        r += 1
    r += 2

    r = title(ws, r, "Payment/Credit split -- the judgment call, shown both ways",
              "'Efisiensi Biaya' is bucketed as Payment/Credit but reads as a value "
              "objection. Decide which it is before designing the playbook.")
    r = header(ws, r, ["Size band", "Arrears / won't pay (ARR)", "Cost efficiency (ARR)",
                       "Arrears veh", "Cost-eff veh", "Arrears cust", "Cost-eff cust"])
    arr_k = "Kendala Pembayaran - Customer Tidak Mau Bayar/ Menunggak"
    eff_k = "Kendala Pembayaran - Efisiensi Biaya"
    pcells = agg(full, lambda d: (d["_band"], d["reason_churn"]))
    for b in bands:
        a, e_ = pcells.get((b, arr_k)), pcells.get((b, eff_k))
        vals = [b, a["amt"] if a else 0, e_["amt"] if e_ else 0, a["veh"] if a else 0,
                e_["veh"] if e_ else 0, len(a["cust"]) if a else 0,
                len(e_["cust"]) if e_ else 0]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v if (i == 0 or v) else None)
            c.border = BOX
            if i:
                c.number_format = IDR
        r += 1
    r += 2

    r = title(ws, r, "FULL CHURN: raw reason detail (all 22 values)")
    r = header(ws, r, ["Reason (raw)", "Bucket", "Churned ARR", "% full ARR", "Vehicles",
                       "% full veh", "Customers", "Median tenure (mo)"])
    ragg = agg(full, lambda d: d["reason_churn"])
    for k in sorted(ragg, key=lambda k: -ragg[k]["amt"]):
        e = ragg[k]
        ten = [d["_tenure"] for d in full if d["reason_churn"] == k and d["_tenure"] is not None]
        vals = [k, REASON_MAP[k], e["amt"], e["amt"] / ctx["full_amt"], e["veh"],
                e["veh"] / ctx["full_veh"], len(e["cust"]),
                statistics.median(ten) if ten else None]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v)
            c.border = BOX
            c.number_format = PCT if i in (3, 5) else (NUM1 if i == 7 else IDR)
        r += 1


def tab_product_reason(wb, full, ctx):
    ws = wb.create_sheet("03_Product_x_Reason")
    widths(ws, {"A": 26, **{get_column_letter(i): 15 for i in range(2, 12)}})
    r = title(ws, 1, "FULL CHURN: product family x reason bucket",
              "Which product lines lose customers, and to what cause.")
    fams = [f for f, _ in sorted(agg(full, lambda d: d["_fam"]).items(),
                                 key=lambda x: -x[1]["amt"])]
    r = matrix_block(ws, r, full, lambda d: d["_fam"], fams, "Product family",
                     lambda d: d["_bucket"], REASON_BUCKETS, "Reason bucket",
                     ctx["full_amt"], ctx["full_veh"], "03 product x bucket")

    r = title(ws, r, "Product family summary (full churn)")
    r = header(ws, r, ["Product family", "Churned ARR", "% full ARR", "Vehicles",
                       "Customers", "ARR / vehicle", "Median tenure (mo)"])
    fagg = agg(full, lambda d: d["_fam"])
    for f in fams:
        e = fagg[f]
        ten = [d["_tenure"] for d in full if d["_fam"] == f and d["_tenure"] is not None]
        vals = [f, e["amt"], e["amt"] / ctx["full_amt"], e["veh"], len(e["cust"]),
                e["amt"] / e["veh"] if e["veh"] else 0,
                statistics.median(ten) if ten else None]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v)
            c.border = BOX
            c.number_format = PCT if i == 2 else (NUM1 if i == 6 else IDR)
        r += 1


def tab_size_product(wb, full, ctx):
    ws = wb.create_sheet("04_Size_x_Product")
    widths(ws, {"A": 22, **{get_column_letter(i): 15 for i in range(2, 16)}})
    r = title(ws, 1, "FULL CHURN: size band x product family",
              "Where each segment's churned ARR actually sits.")
    bands = [b for b in SIZE_ORDER if any(d["_band"] == b for d in full)]
    fams = [f for f, _ in sorted(agg(full, lambda d: d["_fam"]).items(),
                                 key=lambda x: -x[1]["amt"])]
    matrix_block(ws, r, full, lambda d: d["_band"], bands, "Size band",
                 lambda d: d["_fam"], fams, "Product family",
                 ctx["full_amt"], ctx["full_veh"], "04 size x family")


def tab_tenure(wb, full, ctx):
    ws = wb.create_sheet("05_Tenure_Curve")
    widths(ws, {"A": 22, **{get_column_letter(i): 15 for i in range(2, 12)}})
    r = title(ws, 1, "FULL CHURN: when in the lifecycle do they die?",
              "Tenure = churn date minus subscription start. This drives WHEN Account "
              "Management should intervene.")
    r = header(ws, r, ["Tenure bucket", "Customers", "Churned ARR", "% full ARR",
                       "Vehicles", "% full veh", "Cumulative % ARR"])
    tagg = agg(full, lambda d: tenure_label(d["_tenure"]))
    cum = 0.0
    for k in TENURE_ORDER:
        if k not in tagg:
            continue
        e = tagg[k]
        cum += e["amt"] / ctx["full_amt"]
        vals = [k, len(e["cust"]), e["amt"], e["amt"] / ctx["full_amt"], e["veh"],
                e["veh"] / ctx["full_veh"], cum]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v)
            c.border = BOX
            c.number_format = PCT if i in (3, 5, 6) else IDR
        r += 1
    r += 2

    for lbl, keyfn, keys in [
        ("reason bucket", lambda d: d["_bucket"], REASON_BUCKETS),
        ("size band", lambda d: d["_band"], [b for b in SIZE_ORDER
                                             if any(d["_band"] == b for d in full)]),
    ]:
        r = title(ws, r, "Tenure bucket x %s" % lbl)
        r = matrix_block(ws, r, full, lambda d: tenure_label(d["_tenure"]),
                         [k for k in TENURE_ORDER if any(tenure_label(d["_tenure"]) == k
                                                         for d in full)],
                         "Tenure bucket", keyfn, keys, lbl,
                         ctx["full_amt"], ctx["full_veh"], "05 tenure x %s" % lbl)

    r = title(ws, r, "Median tenure at churn (months)")
    r = header(ws, r, ["Cut", "Value", "Median tenure", "Lines"])
    cuts = [("status", lambda d: d["churn_status"])] + \
           [("reason bucket", lambda d: d["_bucket"])] + \
           [("size band", lambda d: d["_band"])]
    src = ctx["rows"]
    for name, fn in cuts:
        pool = src if name == "status" else full
        groups = collections.defaultdict(list)
        for d in pool:
            if d["_tenure"] is not None:
                groups[fn(d)].append(d["_tenure"])
        for k in sorted(groups, key=lambda k: -statistics.median(groups[k])):
            for i, v in enumerate([name, k, statistics.median(groups[k]), len(groups[k])]):
                c = ws.cell(r, 1 + i, v)
                c.border = BOX
                c.number_format = NUM1 if i == 2 else IDR
            r += 1


def tab_concentration(wb, customers, ctx):
    ws = wb.create_sheet("06_Concentration")
    widths(ws, {"A": 6, "B": 46, "C": 14, "D": 16, "E": 18, "F": 14, "G": 12, "H": 16,
                "I": 18, "J": 16})
    r = title(ws, 1, "Concentration: churned ARR is not evenly spread",
              "Customer-level, both statuses. If a handful of accounts carry the loss, "
              "an AM program is account-specific work, not a segment campaign.")
    ranked = sorted(customers.values(), key=lambda c: -c["amount"])
    r = header(ws, r, ["#", "Customer", "Status", "Size band", "Product family",
                       "Churned ARR", "Vehicles", "% of total ARR", "Cumulative %",
                       "Tenure (mo)"])
    cum = 0.0
    marks = {}
    for i, c in enumerate(ranked, 1):
        cum += c["amount"] / ctx["total_amt"]
        for thr in (0.5, 0.8):
            if thr not in marks and cum >= thr:
                marks[thr] = i
        vals = [i, c["name"], c["status"], c["band"], c["family"], c["amount"],
                c["vehicles"], c["amount"] / ctx["total_amt"], cum, c["tenure"]]
        for j, v in enumerate(vals):
            cell = ws.cell(r, 1 + j, v)
            cell.border = BOX
            cell.number_format = PCT if j in (7, 8) else (NUM1 if j == 9 else IDR)
            if i <= 15:
                cell.fill = SUB_FILL
        r += 1
    ws.cell(2, 12, "Top 15 = %.0f%% of churned ARR" % (
        sum(c["amount"] for c in ranked[:15]) / ctx["total_amt"] * 100)).font = BOLD
    ws.cell(3, 12, "%d customers carry 50%% of churned ARR (of %d total)"
            % (marks.get(0.5, 0), len(ranked))).font = BOLD
    ws.cell(4, 12, "%d customers carry 80%%" % marks.get(0.8, 0)).font = BOLD
    ws.cell(5, 12, "Highlighted rows are the top 15.").font = NOTE_FONT


def tab_cohort_matrix(wb, full, ctx):
    ws = wb.create_sheet("07_Cohort_Matrix")
    r = title(ws, 1, "FULL CHURN: cohort month x churn month",
              "Cohort from the source 'Cohort Month - Year' column (same basis as the "
              "source Summary tab). Three stacked matrices: ARR, vehicles, customers.")
    cohorts = sorted({d["_cohort_my"] for d in full if d["_cohort_my"]})
    months = sorted({d["_churn_my"] for d in full})
    widths(ws, {"A": 14, **{get_column_letter(i): 13 for i in range(2, len(months) + 3)}})
    pool = [d for d in full if d["_cohort_my"]]
    dropped = len(full) - len(pool)
    amt, veh = sum(d["_amt"] for d in pool), sum(d["_veh"] for d in pool)
    if dropped:
        ws.cell(r, 1, "%d full-churn line(s) have a non-date cohort value (#N/A) and are "
                      "excluded from this tab only." % dropped).font = NOTE_FONT
        r += 2
    matrix_block(ws, r, pool, lambda d: d["_cohort_my"], cohorts, "Cohort",
                 lambda d: d["_churn_my"], months, "Churn month", amt, veh,
                 "07 cohort x churn")


def tab_partial(wb, partial, customers, ctx):
    ws = wb.create_sheet("08_Partial_Churn")
    widths(ws, {"A": 24, **{get_column_letter(i): 15 for i in range(2, 12)}})
    r = title(ws, 1, "PARTIAL CHURN -- the larger bleed",
              "Partial churn is %.0f%% of all churned ARR (IDR %s across %d customers). "
              "It is vehicles walking off contracts at accounts that stay customers."
              % (ctx["part_amt"] / ctx["total_amt"] * 100,
                 format(int(ctx["part_amt"]), ","),
                 sum(1 for c in customers.values() if c["status"] == "partial_churn")))
    bands = [b for b in SIZE_ORDER if any(d["_band"] == b for d in partial)]
    r = header(ws, r, ["Size band", "Customers", "Churned ARR", "% partial ARR",
                       "Vehicles shed", "% partial veh", "Vehicles / customer",
                       "ARR / customer"])
    pagg = agg(partial, lambda d: d["_band"])
    for b in bands:
        e = pagg[b]
        n = len(e["cust"])
        vals = [b, n, e["amt"], e["amt"] / ctx["part_amt"], e["veh"],
                e["veh"] / ctx["part_veh"], e["veh"] / n, e["amt"] / n]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v)
            c.border = BOX
            c.number_format = PCT if i in (3, 5) else (NUM1 if i == 6 else IDR)
            if b == "G>=400":
                c.fill = WARN_FILL
        r += 1
    r += 2

    r = matrix_block(ws, r, partial, lambda d: d["_band"], bands, "Size band",
                     lambda d: d["_bucket"], REASON_BUCKETS, "Reason bucket",
                     ctx["part_amt"], ctx["part_veh"], "08 partial size x bucket")

    fams = [f for f, _ in sorted(agg(partial, lambda d: d["_fam"]).items(),
                                 key=lambda x: -x[1]["amt"])]
    r = title(ws, r, "Partial churn: product family x reason bucket")
    r = matrix_block(ws, r, partial, lambda d: d["_fam"], fams, "Product family",
                     lambda d: d["_bucket"], REASON_BUCKETS, "Reason bucket",
                     ctx["part_amt"], ctx["part_veh"], "08 partial fam x bucket")

    r = title(ws, r, "Largest vehicle-shedding accounts (partial churn)",
              "These accounts are still customers. Every vehicle here is recoverable "
              "revenue in a way a full churn is not.")
    r = header(ws, r, ["Customer", "Size band", "Product family", "Reason bucket",
                       "Vehicles shed", "Churned ARR", "Churn events (lines)",
                       "First churn", "Last churn"])
    top = sorted([c for c in customers.values() if c["status"] == "partial_churn"],
                 key=lambda c: -c["vehicles"])[:40]
    for c in top:
        vals = [c["name"], c["band"], c["family"], c["bucket"], c["vehicles"],
                c["amount"], c["lines"], c["first_churn"].strftime("%Y-%m-%d"),
                c["last_churn"].strftime("%Y-%m-%d")]
        for i, v in enumerate(vals):
            cell = ws.cell(r, 1 + i, v)
            cell.border = BOX
            if i in (4, 5, 6):
                cell.number_format = IDR
        r += 1


def tab_target_list(wb, customers, ctx):
    ws = wb.create_sheet("09_Target_List")
    widths(ws, {"A": 46, "B": 14, "C": 12, "D": 24, "E": 20, "F": 52, "G": 16, "H": 12,
                "I": 12, "J": 12, "K": 22, "L": 14, "M": 14})
    r = title(ws, 1, "Customer-level target list",
              "One row per churned customer. Sort or filter this to assign AM "
              "playbooks. Bands resolved to largest-ARR line; product/reason are the "
              "customer's ARR-dominant values.")
    hdr_row = r
    r = header(ws, r, ["Customer", "Status", "Size band", "Product family (dominant)",
                       "Reason bucket (dominant)", "Reason (raw, largest line)",
                       "Churned ARR", "Vehicles", "Lines", "Tenure (mo)",
                       "Sales owner", "First churn", "Last churn"])
    ws.freeze_panes = ws.cell(r, 1)
    for c in sorted(customers.values(), key=lambda c: (-c["amount"])):
        vals = [c["name"], c["status"], c["band"], c["family"], c["bucket"], c["reason"],
                c["amount"], c["vehicles"], c["lines"], c["tenure"], c["sales"],
                c["first_churn"].strftime("%Y-%m-%d"), c["last_churn"].strftime("%Y-%m-%d")]
        for i, v in enumerate(vals):
            cell = ws.cell(r, 1 + i, v)
            cell.border = BOX
            if i in (6, 7, 8):
                cell.number_format = IDR
            elif i == 9:
                cell.number_format = NUM1
        r += 1
    ws.auto_filter.ref = "A%d:M%d" % (hdr_row, r - 1)


def tab_tenure_dist(wb, rows, ctx):
    """Per-tenure-month distribution, outlier accounts removed."""
    ws = wb.create_sheet("11_Tenure_Distribution")
    widths(ws, {"A": 13, "B": 9, "C": 9, "D": 17, "E": 9, "F": 10, "G": 9, "H": 9,
                "I": 10, "J": 3, "K": 34, "L": 15})

    pool = [d for d in rows if d["customer_wo"] not in TENURE_DIST_EXCLUDE
            and d["_tenure"] is not None and d["_tenure"] >= 0]
    dropped_out = [d for d in rows if d["customer_wo"] in TENURE_DIST_EXCLUDE]
    dropped_neg = [d for d in rows if d["_tenure"] is not None and d["_tenure"] < 0]
    dropped_nul = [d for d in rows if d["_tenure"] is None]

    r = title(ws, 1, "Churn distribution by tenure month (outliers excluded)",
              "Tenure month N = churned during the Nth month after subscription start "
              "(floor of days/30.44). Months %d+ collapsed into one tail row."
              % TENURE_TAIL)
    for line in [
        "EXCLUDED: %s -- IDR %s across %d lines and %d vehicles (%.1f%% of all churned "
        "ARR in the file). Partial churn, so the full-churn columns below are unaffected."
        % (", ".join(TENURE_DIST_EXCLUDE),
           format(int(sum(d["_amt"] for d in dropped_out)), ","), len(dropped_out),
           sum(d["_veh"] for d in dropped_out),
           sum(d["_amt"] for d in dropped_out) / ctx["total_amt"] * 100),
        "ALSO EXCLUDED: %d line(s) with negative tenure (churn dated before "
        "subscription start) and %d with no parseable start date."
        % (len(dropped_neg), len(dropped_nul)),
        "This tab is the ONLY one that drops accounts. Every other tab is inclusive and "
        "ties to the source Summary tab.",
    ]:
        c = ws.cell(r, 1, line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.fill = WARN_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1

    full = [d for d in pool if d["churn_status"] == "full_churn"]
    part = [d for d in pool if d["churn_status"] == "partial_churn"]
    TA, TV = sum(d["_amt"] for d in pool), sum(d["_veh"] for d in pool)
    FA = sum(d["_amt"] for d in full)
    PA = sum(d["_amt"] for d in part)

    def bin_of(d):
        return min(int(d["_tenure"]), TENURE_TAIL)

    bins = sorted({bin_of(d) for d in pool})
    a_all, a_full, a_part = (agg(x, bin_of) for x in (pool, full, part))

    r = header(ws, r, ["Tenure month", "Lines", "Customers", "Churned ARR (all)",
                       "% ARR", "Cum % ARR", "Vehicles", "% veh", "Cum % veh",
                       "", "Largest single account in bin", "Its share of bin"])
    cum = cumv = 0.0
    for b in bins:
        e = a_all[b]
        cum += e["amt"] / TA
        cumv += e["veh"] / TV
        lines_in = [d for d in pool if bin_of(d) == b]
        by_cust = collections.defaultdict(float)
        for d in lines_in:
            by_cust[d["customer_wo"]] += d["_amt"]
        top_name, top_amt = max(by_cust.items(), key=lambda x: x[1])
        share = top_amt / e["amt"] if e["amt"] else 0
        vals = [("%d+" % TENURE_TAIL) if b == TENURE_TAIL else b, e["n"], len(e["cust"]),
                e["amt"], e["amt"] / TA, cum, e["veh"], e["veh"] / TV, cumv, None,
                top_name, share]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v)
            if i != 9:
                c.border = BOX
            c.number_format = PCT if i in (4, 5, 7, 8, 11) else IDR
        # Flag bins that are really one account wearing a distribution's clothing.
        if share >= 0.5 and e["amt"] / TA >= 0.02:
            for i in (0, 3, 4, 10, 11):
                ws.cell(r, 1 + i).fill = WARN_FILL
        r += 1
    for i, v in enumerate(["TOTAL", len(pool), len({d["customer_id"] for d in pool}),
                           TA, 1.0, None, TV, 1.0, None]):
        c = ws.cell(r, 1 + i, v)
        c.font, c.fill, c.border = BOLD, TOTAL_FILL, BOX
        c.number_format = PCT if i in (4, 7) else IDR
    ws.cell(r + 1, 1, "Shaded rows: one account is >=50% of the bin AND the bin is >=2% "
                      "of ARR -- read these as single events, not as lifecycle "
                      "patterns.").font = NOTE_FONT
    r += 4

    r = title(ws, r, "Same distribution split by churn type",
              "Excluding Corin changes the partial and total columns only -- full churn "
              "is untouched because Corin is a partial-churn account.")
    r = header(ws, r, ["Tenure month", "FULL ARR", "% full", "FULL veh", "FULL cust",
                       "PARTIAL ARR", "% partial", "PARTIAL veh", "PARTIAL cust"])
    for b in bins:
        f, p = a_full.get(b), a_part.get(b)
        vals = [("%d+" % TENURE_TAIL) if b == TENURE_TAIL else b,
                f["amt"] if f else 0, (f["amt"] / FA) if f else 0, f["veh"] if f else 0,
                len(f["cust"]) if f else 0,
                p["amt"] if p else 0, (p["amt"] / PA) if p else 0, p["veh"] if p else 0,
                len(p["cust"]) if p else 0]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v if (i == 0 or v) else None)
            c.border = BOX
            c.number_format = PCT if i in (2, 6) else IDR
        r += 1
    r += 2

    r = title(ws, r, "Contract-anniversary clustering",
              "Is churn actually clustered at renewal points, or spread evenly? "
              "Baseline = average ARR share per single tenure month over months 0-%d."
              % (TENURE_TAIL - 1))
    base_pool = [d for d in pool if d["_tenure"] < TENURE_TAIL]
    base_amt = sum(d["_amt"] for d in base_pool)
    baseline = (base_amt / TA) / TENURE_TAIL
    r = header(ws, r, ["Window", "Months", "Churned ARR", "% of ARR", "Vehicles",
                       "Lines", "Customers", "ARR share per month", "vs baseline"])
    for label, lo, hi in ANNIV_WINDOWS + GAP_WINDOWS:
        s = [d for d in pool if lo <= d["_tenure"] < hi]
        a = sum(d["_amt"] for d in s)
        per = (a / TA) / (hi - lo)
        vals = [label, "%d-%d" % (lo, hi - 1), a, a / TA, sum(d["_veh"] for d in s),
                len(s), len({d["customer_id"] for d in s}), per, per / baseline]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v)
            c.border = BOX
            c.number_format = PCT if i in (3, 7) else (NUM1 if i == 8 else IDR)
            if i == 8:
                c.font = BOLD
                if v >= 1.5:
                    c.fill = WARN_FILL
        r += 1
    ws.cell(r, 1, "Baseline ARR share per tenure month").font = BOLD
    c = ws.cell(r, 8, baseline)
    c.number_format, c.font = PCT, BOLD
    ws.cell(r, 9, 1.0).number_format = NUM1
    r += 3

    r = title(ws, r, "Coarse buckets, for reporting")
    r = header(ws, r, ["Tenure bucket", "Lines", "Customers", "Churned ARR", "% ARR",
                       "Cum % ARR", "Vehicles", "% veh"])
    bagg = agg(pool, lambda d: tenure_label(d["_tenure"]))
    cum = 0.0
    for k in TENURE_ORDER:
        if k not in bagg:
            continue
        e = bagg[k]
        cum += e["amt"] / TA
        vals = [k, e["n"], len(e["cust"]), e["amt"], e["amt"] / TA, cum, e["veh"],
                e["veh"] / TV]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v)
            c.border = BOX
            c.number_format = PCT if i in (4, 5, 7) else IDR
        r += 1

    return {
        "pool_amt": TA, "pool_veh": TV, "pool_lines": len(pool),
        "pool_cust": len({d["customer_id"] for d in pool}),
        "median": statistics.median([d["_tenure"] for d in pool]),
        "baseline": baseline,
        "dropped_amt": sum(d["_amt"] for d in dropped_out),
        "n_neg": len(dropped_neg),
    }


def tab_quality(wb, rows, customers, ctx):
    ws = wb.create_sheet("10_Data_Quality")
    widths(ws, {"A": 46, "B": 30, "C": 30, "D": 20, "E": 20})
    r = title(ws, 1, "Data quality, exclusions, and the missing denominator")

    r = title(ws, r, "1. What is needed to turn MIX into RATE",
              "Without this, no statement in this workbook can claim a segment is "
              "riskier than another.")
    for line in [
        "Active customer count, active ARR, and active vehicle count -- as at each "
        "month-end -- broken down by fleet_category and by product family.",
        "Same breakdown for the start of the analysis window, so denominators can be "
        "averaged rather than point-in-time.",
        "Contract start and contract end (or renewal) dates per subscription, so the "
        "renewal cliff can be measured against actual contract terms instead of "
        "inferred from tenure.",
        "AR ageing per customer at churn date, to test whether arrears-labelled churn "
        "was preceded by a collectible signal.",
    ]:
        c = ws.cell(r, 1, "- " + line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
    r += 2

    r = title(ws, r, "2. Row counts and exclusions")
    r = header(ws, r, ["Item", "Count", "Effect on this workbook"])
    excl_jun = sum(1 for d in rows if d["State Cohort Jun'25"] != "Include")
    excl_dec = sum(1 for d in rows if d["State Cohort Dec'25"] != "Include")
    items = [
        ("Data rows (WO product lines)", ctx["n_rows"], "All included"),
        ("Zero-ARR lines", sum(1 for d in rows if d["_amt"] == 0),
         "Kept. Invisible in ARR cuts -- see 01_Headline"),
        ("Lines with blank fleet_category", sum(1 for d in rows if not d["fleet_category"]),
         "Customer's band resolved from its largest-ARR line; '(unassigned)' only where "
         "every line is blank"),
        ("Lines with blank sales_name", sum(1 for d in rows if not d["sales_name"]),
         "Shown as '(none)' in 09_Target_List"),
        ("Lines with non-date Cohort Month - Year (#N/A)",
         sum(1 for d in rows if not d["_cohort_my"]),
         "Excluded from 07_Cohort_Matrix only"),
        ("Lines with unparseable subscription_start_date",
         sum(1 for d in rows if d["_start"] is None),
         "Excluded from tenure views only"),
        ("Lines with NEGATIVE tenure (churn dated before subscription start)",
         sum(1 for d in rows if d["_tenure"] is not None and d["_tenure"] < 0),
         "Source data error. Kept in ARR/vehicle views, excluded from "
         "11_Tenure_Distribution. Listed below."),
        ("Lines flagged State Cohort Jun'25 <> Include", excl_jun,
         "NOT excluded here -- this workbook uses the full window"),
        ("Lines flagged State Cohort Dec'25 <> Include", excl_dec,
         "NOT excluded here"),
        ("Customers whose lines span >1 fleet_category",
         sum(1 for c in customers.values() if len(c["bands_seen"]) > 1),
         "Resolved to largest-ARR band; listed below"),
        ("Customers whose lines span >1 reason bucket",
         sum(1 for c in customers.values() if len(c["buckets_seen"]) > 1),
         "Dominant bucket used at customer level; all lines counted in crosstabs"),
        ("Product lines mapped to 'Unmapped' family",
         sum(1 for d in rows if d["_fam"] == "Unmapped"), "Should be 0 -- see mapping"),
    ]
    for k, v, eff in items:
        ws.cell(r, 1, k).border = BOX
        c = ws.cell(r, 2, v)
        c.number_format, c.border, c.font = IDR, BOX, BOLD
        c = ws.cell(r, 3, eff)
        c.border, c.alignment = BOX, Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 2

    r = title(ws, r, "3. Customers with conflicting size bands")
    r = header(ws, r, ["Customer", "Bands seen in source", "Band used (largest-ARR line)",
                       "Churned ARR", "Vehicles"])
    for c in sorted([c for c in customers.values() if len(c["bands_seen"]) > 1],
                    key=lambda c: -c["amount"]):
        vals = [c["name"], " | ".join(c["bands_seen"]), c["band"], c["amount"], c["vehicles"]]
        for i, v in enumerate(vals):
            cell = ws.cell(r, 1 + i, v)
            cell.border = BOX
            if i >= 3:
                cell.number_format = IDR
        r += 1
    r += 2

    r = title(ws, r, "3b. Lines with negative tenure")
    r = header(ws, r, ["Customer", "Subscription start", "Churn (WO acc date)",
                       "Tenure (mo)", "Churned ARR"])
    for d in sorted([d for d in rows if d["_tenure"] is not None and d["_tenure"] < 0],
                    key=lambda d: d["_tenure"]):
        vals = [d["customer_wo"], d["_start"].strftime("%Y-%m-%d"),
                d["_churn"].strftime("%Y-%m-%d"), d["_tenure"], d["_amt"]]
        for i, v in enumerate(vals):
            c = ws.cell(r, 1 + i, v)
            c.border, c.fill = BOX, WARN_FILL
            c.number_format = NUM1 if i == 3 else IDR
        r += 1
    r += 2

    r = title(ws, r, "4. Cohort month vs subscription start month",
              "The source 'Cohort Month - Year' does not always equal the subscription "
              "start month. 07_Cohort_Matrix follows the source column; tenure follows "
              "the actual start date.")
    mismatch = [d for d in rows if d["_start"] and d["_cohort_my"]
                and d["_cohort_my"] != d["_start"].strftime("%Y-%m")]
    ws.cell(r, 1, "Lines where cohort month <> subscription start month:").font = BOLD
    c = ws.cell(r, 2, len(mismatch))
    c.number_format, c.font = IDR, BOLD
    r += 1
    ws.cell(r, 1, "As %% of all lines:").font = BOLD
    c = ws.cell(r, 2, len(mismatch) / ctx["n_rows"])
    c.number_format, c.font = PCT, BOLD
    r += 3

    r = title(ws, r, "5. Product family mapping (auditable -- edit PRODUCT_RULES to change)")
    r = header(ws, r, ["product_name (source)", "Mapped family", "Lines", "Churned ARR"])
    pagg = agg(rows, lambda d: (d["product_name"], d["_fam"]))
    for (pn, fam) in sorted(pagg, key=lambda k: (k[1], -pagg[k]["amt"])):
        e = pagg[(pn, fam)]
        vals = [pn, fam, e["n"], e["amt"]]
        for i, v in enumerate(vals):
            cell = ws.cell(r, 1 + i, v)
            cell.border = BOX
            if i >= 2:
                cell.number_format = IDR
            if fam == "Unmapped":
                cell.fill = WARN_FILL
        r += 1
    r += 2

    r = title(ws, r, "6. Reason bucket mapping")
    r = header(ws, r, ["reason_churn (source)", "Bucket", "Lines", "Churned ARR"])
    ragg = agg(rows, lambda d: (d["reason_churn"], d["_bucket"]))
    for (rn, bk) in sorted(ragg, key=lambda k: (k[1], -ragg[k]["amt"])):
        e = ragg[(rn, bk)]
        for i, v in enumerate([rn, bk, e["n"], e["amt"]]):
            cell = ws.cell(r, 1 + i, v)
            cell.border = BOX
            if i >= 2:
                cell.number_format = IDR
        r += 1


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------


def main():
    if len(sys.argv) != 3:
        raise SystemExit(__doc__)
    src, out = Path(sys.argv[1]), Path(sys.argv[2])
    if not src.exists():
        raise SystemExit("ERROR: source not found: %s" % src)

    rows = load_rows(src)
    customers = build_customers(rows)
    full = [d for d in rows if d["churn_status"] == "full_churn"]
    partial = [d for d in rows if d["churn_status"] == "partial_churn"]

    statuses = {d["churn_status"] for d in rows}
    if statuses - {"full_churn", "partial_churn"}:
        raise SystemExit("ERROR: unexpected churn_status: %s" % sorted(statuses))
    mixed = [c for c in customers.values()
             if len({d["churn_status"] for d in rows if d["customer_id"] == c["customer_id"]}) > 1]
    if mixed:
        raise SystemExit("ERROR: %d customer(s) hold both full and partial lines; the "
                         "customer-level status label is no longer safe." % len(mixed))

    jun26 = [d for d in rows if d["_churn_my"] == "2026-06"]
    ctx = {
        "rows": rows,
        "src_name": src.name,
        "rundate": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "n_rows": len(rows),
        "n_wo": len({d["wo_name"] for d in rows}),
        "n_cust": len(customers),
        "total_amt": sum(d["_amt"] for d in rows),
        "total_veh": sum(d["_veh"] for d in rows),
        "full_amt": sum(d["_amt"] for d in full),
        "full_veh": sum(d["_veh"] for d in full),
        "part_amt": sum(d["_amt"] for d in partial),
        "part_veh": sum(d["_veh"] for d in partial),
        "min_my": min(d["_churn_my"] for d in rows),
        "max_my": max(d["_churn_my"] for d in rows),
        "jun26_n": len(jun26),
        "jun26_amt": sum(d["_amt"] for d in jun26),
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    tab_readme(wb, ctx)
    tab_headline(wb, rows, customers, ctx)
    tab_size_reason(wb, full, ctx)
    tab_product_reason(wb, full, ctx)
    tab_size_product(wb, full, ctx)
    tab_tenure(wb, full, ctx)
    tab_concentration(wb, customers, ctx)
    tab_cohort_matrix(wb, full, ctx)
    tab_partial(wb, partial, customers, ctx)
    tab_target_list(wb, customers, ctx)
    tab_quality(wb, rows, customers, ctx)
    td = tab_tenure_dist(wb, rows, ctx)

    # Write via temp file: outputs/ sits on a OneDrive-synced path.
    tmp_out = Path(tempfile.gettempdir()) / ("_churn_out_%s" % out.name)
    wb.save(tmp_out)
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(tmp_out, out)
    except PermissionError:
        # Target open in Excel or locked by OneDrive sync. Don't discard the run.
        alt = out.with_name("%s-rerun%s" % (out.stem, out.suffix))
        shutil.copy2(tmp_out, alt)
        out = alt
        print("NOTE: target was locked (open in Excel or syncing). Wrote to %s instead."
              % alt.name)
    tmp_out.unlink(missing_ok=True)

    # ---------------- reconciliation ----------------
    print("=" * 74)
    print("RECONCILIATION")
    print("=" * 74)
    print("source                : %s" % src.name)
    print("rows (WO lines)       : %d" % ctx["n_rows"])
    print("work orders           : %d" % ctx["n_wo"])
    print("customers             : %d" % ctx["n_cust"])
    print("churned ARR (IDR)     : %s" % format(round(ctx["total_amt"], 2), ","))
    print("vehicles              : %s" % format(ctx["total_veh"], ","))
    print("churn window          : %s .. %s" % (ctx["min_my"], ctx["max_my"]))
    print("-" * 74)
    for label, subset in (("full_churn", full), ("partial_churn", partial)):
        cs = [c for c in customers.values() if c["status"] == label]
        print("%-14s: lines=%4d customers=%3d ARR=%18s veh=%6s"
              % (label, len(subset), len(cs),
                 format(round(sum(d["_amt"] for d in subset), 2), ","),
                 format(sum(d["_veh"] for d in subset), ",")))
    split = sum(d["_amt"] for d in full) + sum(d["_amt"] for d in partial)
    assert abs(split - ctx["total_amt"]) < 0.01, "status split does not sum to total"
    assert (sum(d["_veh"] for d in full) + sum(d["_veh"] for d in partial)
            == ctx["total_veh"]), "vehicle split does not sum to total"
    print("-" * 74)

    lo, hi = SUMMARY_WINDOW
    tie = sum(d["_amt"] for d in rows
              if d["State Cohort Jun'25"] == "Include" and lo <= d["_churn_my"] <= hi)
    delta = tie - SUMMARY_TIEOUT
    print("TIE-OUT vs source Summary tab (Cohort Jun'25=Include, %s..%s)" % (lo, hi))
    print("  computed : %s" % format(round(tie, 2), ","))
    print("  expected : %s" % format(SUMMARY_TIEOUT, ","))
    print("  delta    : %s  -> %s" % (format(round(delta, 2), ","),
                                      "PASS" if abs(delta) < 1 else "FAIL"))
    assert abs(delta) < 1, "tie-out FAILED: parsing does not reproduce the Summary tab"
    print("-" * 74)
    print("TENURE DISTRIBUTION tab (excludes %s)" % ", ".join(TENURE_DIST_EXCLUDE))
    print("  removed          : IDR %s (%.1f%% of churned ARR)"
          % (format(round(td["dropped_amt"]), ","),
             td["dropped_amt"] / ctx["total_amt"] * 100))
    print("  also removed     : %d negative-tenure line(s)" % td["n_neg"])
    print("  remaining        : %d lines, %d customers, IDR %s, %s vehicles"
          % (td["pool_lines"], td["pool_cust"], format(round(td["pool_amt"]), ","),
             format(td["pool_veh"], ",")))
    print("  median tenure    : %.1f months" % td["median"])
    print("-" * 74)
    print("all crosstab margins asserted against section totals: PASS")
    unmapped = [d for d in rows if d["_fam"] == "Unmapped"]
    if unmapped:
        names = sorted({d["product_name"] for d in unmapped})
        u_amt = sum(d["_amt"] for d in unmapped)
        print("WARNING: %d line(s) / IDR %s (%.2f%% of churn) have no product-family "
              "rule." % (len(unmapped), format(round(u_amt), ","),
                         u_amt / ctx["total_amt"] * 100))
        print("         add a rule to PRODUCT_RULES for: %s" % "; ".join(names))
    else:
        print("product-family mapping: 0 unmapped lines: PASS")
    print("workbook written      : %s" % out)
    print("tabs                  : %s" % ", ".join(wb.sheetnames))
    print("=" * 74)


if __name__ == "__main__":
    main()
