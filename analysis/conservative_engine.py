"""Monthly re-basing engine for McEasy's conservative 2030 cases.

Reads the source model's monthly columns, then re-bases the Jul-2026 onward path so FY2030
revenue lands on a target, under two drivers:

  volume  - fewer vehicles, price holds. COGS and capex scale with revenue.
  arpu    - vehicle count holds, price falls. SaaS COGS and capex held absolute (the devices
            are already bought); spare-part COGS scales with spare-part revenue.

Everything below gross profit is held at plan in both drivers (full operating deleverage).
Depreciation runs off a 48-month straight-line device cohort, which the actuals support: the
implied blended life over H1-2026 is 47.7 months.
"""
import datetime

import openpyxl

MODEL_PATH = (r"C:\Users\Grady Kusmulyadi\OneDrive - PT. Otto Menara Globalindo"
              r"\Documents\H2 2026\Strategy 2027\McEasy model for claude reading.xlsx")

DEVICE_LIFE = 48          # months, per Grady
VEH_2030 = 950_000        # Grady's stated 2030 vehicle count
ANCHOR = "2030-12"        # month whose ARR anchors ARR-per-vehicle


def _find(ws, label):
    for r in range(1, 120):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == label:
            return r
    raise KeyError(f"{label!r} not in {ws.title!r}")


def _months(ws):
    out = []
    for c in range(1, ws.max_column + 1):
        y = ws.cell(row=8, column=c).value
        s = ws.cell(row=7, column=c).value
        if isinstance(y, (datetime.date, datetime.datetime)):
            out.append((c, y.strftime("%Y-%m"), str(s or "")))
    return out


def load():
    """Return the plan as monthly dicts keyed 'YYYY-MM', plus the actual/projection split."""
    wb = openpyxl.load_workbook(MODEL_PATH, data_only=True)
    pl, bs = wb["PnL_Consol (USD)"], wb["BS_Consol (USD)"]
    mp, mb = _months(pl), _months(bs)
    keys = [k for _, k, _ in mp]
    status = {k: s for _, k, s in mp}

    PL_ROWS = {"arr": "ARR", "rev": "Total Revenue", "saas": "SaaS Solutions",
               "cogs_tot": "Total Cost of Revenue", "gp": "Total Gross Profit",
               "logi": "Logistic Cost", "sm": "Sales & Marketing Costs",
               "opex": "Total Operational Expenses", "ebitda": "EBITDA",
               "da": "Depreciation", "tax": "Tax & Interest", "np": "Net Profit (Loss)"}
    # "SaaS Solutions" appears under Revenue, Cost of Revenue and Gross Profit, in that order
    saas_rows = [r for r in range(1, 100)
                 if str(pl.cell(row=r, column=1).value).strip() == "SaaS Solutions"]
    spare_rows = [r for r in range(1, 100)
                  if str(pl.cell(row=r, column=1).value).strip() == "Sparepart Solutions"]

    P = {}
    for k, lab in PL_ROWS.items():
        r = saas_rows[0] if k == "saas" else _find(pl, lab)
        P[k] = {key: (pl.cell(row=r, column=c).value or 0.0) for c, key, _ in mp}
    P["cogs_saas"] = {key: (pl.cell(row=saas_rows[1], column=c).value or 0.0) for c, key, _ in mp}
    P["spare"] = {key: (pl.cell(row=spare_rows[0], column=c).value or 0.0) for c, key, _ in mp}
    P["cogs_spare"] = {key: (pl.cell(row=spare_rows[1], column=c).value or 0.0) for c, key, _ in mp}

    BS_ROWS = {"cash": "Cash and Cash Equivalent", "gfa": "Fixed Assets",
               "accdep": "Accumulated Depreciation", "ta": "TOTAL ASSETS",
               "puc": "Paid Up Capital"}
    B = {}
    for k, lab in BS_ROWS.items():
        r = _find(bs, lab)
        B[k] = {key: (bs.cell(row=r, column=c).value or 0.0) for c, key, _ in mb}
    # net working capital = total assets less cash less net fixed assets, minus equity+liab side
    # simpler and sufficient: everything on the balance sheet except cash, net PPE and equity
    nwc_assets = ["Accounts Receivable", "Other Receivables", "Inventory", "Prepaid Tax",
                  "Prepaid Expenses", "Prepaid Supplier"]
    nwc_liabs = ["Accounts Payable", "Other Payables", "Accrued Expenses", "Taxes Payable",
                 "Unearned Revenue", "Employee Benefit Liabilities"]
    nwc = {}
    for key in [k for _, k, _ in mb]:
        nwc[key] = 0.0
    for lab, sign in [(l, 1) for l in nwc_assets] + [(l, -1) for l in nwc_liabs]:
        r = _find(bs, lab)
        for c, key, _ in mb:
            nwc[key] += sign * (bs.cell(row=r, column=c).value or 0.0)
    B["nwc"] = nwc

    # monthly capex from gross fixed-asset movement
    bkeys = [k for _, k, _ in mb]
    capex = {}
    for i, k in enumerate(bkeys):
        capex[k] = (B["gfa"][k] - B["gfa"][bkeys[i - 1]]) if i else 0.0
    B["capex"] = capex

    # Effective tax-and-interest rate on EBIT, derived from the plan's own FY2030 column rather
    # than hardcoded — a rounded rate leaves the base case a couple of thousand dollars off plan.
    y30 = [k for k in keys if k.startswith("2030")]
    ebit30 = sum(P["ebitda"][k] for k in y30) - sum(P["da"][k] for k in y30)
    tax_rate = sum(P["tax"][k] for k in y30) / ebit30

    return dict(keys=keys, bkeys=bkeys, status=status, P=P, B=B, tax_rate=tax_rate)


def cohort_dep(capex_by_month, months, opening_gross, opening_remaining_life):
    """48-month straight-line depreciation, plus an opening block amortised over a stated life."""
    dep = {}
    for i, k in enumerate(months):
        d = 0.0
        for j in range(max(0, i - DEVICE_LIFE + 1), i + 1):
            d += capex_by_month.get(months[j], 0.0) / DEVICE_LIFE
        if opening_remaining_life and i < opening_remaining_life:
            d += opening_gross / opening_remaining_life
        dep[k] = d
    return dep


def calibrate_opening_block(M):
    """Solve the opening-block remaining life that reproduces accumulated depreciation at Jun-2026."""
    bkeys, B = M["bkeys"], M["B"]
    first = bkeys[0]
    opening = B["gfa"][first]
    target = -B["accdep"]["2026-06"]
    cap = {k: v for k, v in B["capex"].items()}
    cap[first] = 0.0                     # the opening stock is not a capex event
    upto = bkeys[:bkeys.index("2026-06") + 1]
    best, best_err = None, None
    for life in range(1, 121):
        dep = cohort_dep(cap, upto, opening, life)
        cum = sum(dep[k] for k in upto)
        err = abs(cum - target)
        if best_err is None or err < best_err:
            best, best_err = life, err
    dep = cohort_dep(cap, upto, opening, best)
    return best, sum(dep[k] for k in upto), target


def base_cohort(M, opening_life):
    """The 48-month cohort run on the PLAN's own capex — the reference for the delta approach."""
    cap = {k: (0.0 if k == M["bkeys"][0] else M["B"]["capex"].get(k, 0.0)) for k in M["bkeys"]}
    return cohort_dep(cap, M["bkeys"], M["B"]["gfa"][M["bkeys"][0]], opening_life)


def build(M, target_2030, driver, opening_life, ref_cohort=None):
    """Re-base from Jul-2026. Returns monthly scenario series and annual summaries."""
    keys, P, B, status = M["keys"], M["P"], M["B"], M["status"]
    split = keys.index("2026-07")
    hist, fwd = keys[:split], keys[split:]

    plan_g = {}
    for i, k in enumerate(keys):
        if i == 0 or not P["arr"][keys[i - 1]]:
            plan_g[k] = 0.0
        else:
            plan_g[k] = P["arr"][k] / P["arr"][keys[i - 1]] - 1

    def run(h):
        arr = {}
        for k in hist:
            arr[k] = P["arr"][k]
        prev = arr[hist[-1]]
        for k in fwd:
            prev = prev * (1 + plan_g[k] * h)
            arr[k] = prev
        rev = {}
        for k in keys:
            ratio = (P["rev"][k] / P["arr"][k]) if P["arr"][k] else 0.0
            rev[k] = P["rev"][k] if k in hist else arr[k] * ratio
        return arr, rev

    lo, hi = 0.0, 2.0
    for _ in range(200):
        mid = (lo + hi) / 2
        _, rev = run(mid)
        got = sum(rev[k] for k in keys if k.startswith("2030"))
        if got > target_2030:
            hi = mid
        else:
            lo = mid
    h = (lo + hi) / 2
    arr, rev = run(h)

    S = {"arr": arr, "rev": rev, "h": h}
    for name in ("saas", "spare", "cogs_saas", "cogs_spare", "logi", "sm", "opex"):
        S[name] = {}
    S["capex"] = {}
    for k in keys:
        f = (rev[k] / P["rev"][k]) if P["rev"][k] else 1.0
        if k in hist:
            for name in ("saas", "spare", "cogs_saas", "cogs_spare", "logi", "sm", "opex"):
                S[name][k] = P[name][k]
            S["capex"][k] = B["capex"].get(k, 0.0)
            continue
        mix_s = (P["saas"][k] / P["rev"][k]) if P["rev"][k] else 0.0
        S["saas"][k] = rev[k] * mix_s
        S["spare"][k] = rev[k] * (1 - mix_s)
        spare_ratio = (P["cogs_spare"][k] / P["spare"][k]) if P["spare"][k] else 0.0
        S["cogs_spare"][k] = S["spare"][k] * spare_ratio
        if driver == "volume":
            saas_ratio = (P["cogs_saas"][k] / P["saas"][k]) if P["saas"][k] else 0.0
            S["cogs_saas"][k] = S["saas"][k] * saas_ratio
            S["capex"][k] = B["capex"].get(k, 0.0) * f
        else:                                      # arpu: vehicle-driven costs unchanged
            S["cogs_saas"][k] = P["cogs_saas"][k]
            S["capex"][k] = B["capex"].get(k, 0.0)
        for name in ("logi", "sm", "opex"):        # held at plan
            S[name][k] = P[name][k]

    # Depreciation on a DELTA basis: keep the plan's own depreciation and apply only the
    # incremental 48-month cohort effect of the capex change. Base and ARPU cases then tie to
    # the plan exactly (their capex is unchanged), while the volume case picks up the reduction
    # with the correct 48-month lag rather than instantly.
    cap = {k: (0.0 if k == M["bkeys"][0] else S["capex"].get(k, B["capex"].get(k, 0.0)))
           for k in M["bkeys"]}
    coh = cohort_dep(cap, M["bkeys"], B["gfa"][M["bkeys"][0]], opening_life)
    ref = ref_cohort if ref_cohort is not None else coh
    S["cohort"] = coh
    S["dep"] = {k: P["da"].get(k, 0.0) + (coh.get(k, 0.0) - ref.get(k, 0.0)) for k in keys}

    S["gp"], S["ebitda"], S["ebit"], S["tax"], S["np"] = {}, {}, {}, {}, {}
    for k in keys:
        gp = rev[k] - S["cogs_saas"][k] - S["cogs_spare"][k]
        eb = gp - S["logi"][k] - S["sm"][k] - S["opex"][k]
        ebit = eb - S["dep"].get(k, 0.0)
        tax = M["tax_rate"] * ebit if ebit > 0 else P["tax"][k]
        S["gp"][k], S["ebitda"][k], S["ebit"][k] = gp, eb, ebit
        S["tax"][k], S["np"][k] = tax, ebit - tax

    arr_per_veh = P["arr"][ANCHOR] / VEH_2030
    S["veh"] = {k: (arr[k] / arr_per_veh if driver == "volume" else P["arr"][k] / arr_per_veh)
                for k in keys}
    S["arr_per_veh"] = arr_per_veh
    return S


def annual(series, keys, years=(2026, 2027, 2028, 2029, 2030)):
    out = {}
    for y in years:
        ks = [k for k in keys if k.startswith(str(y))]
        out[y] = sum(series.get(k, 0.0) for k in ks)
    return out


def cash_path(M, S, base_residual=None):
    """Roll cash monthly from Jun-2026. Returns (path, residual) where residual calibrates
    non-modelled cash movements so the BASE case ties exactly to the plan."""
    bkeys, B, P = M["bkeys"], M["B"], M["P"]
    start = bkeys.index("2026-06")
    fwd = bkeys[start + 1:]
    plan_rev_l = {}
    for i, k in enumerate(bkeys):
        ks = bkeys[max(0, i - 11):i + 1]
        plan_rev_l[k] = sum(P["rev"].get(x, 0.0) for x in ks)
    scen_rev_l = {}
    for i, k in enumerate(bkeys):
        ks = bkeys[max(0, i - 11):i + 1]
        scen_rev_l[k] = sum(S["rev"].get(x, 0.0) for x in ks)

    nwc = {}
    for k in bkeys:
        scale = (scen_rev_l[k] / plan_rev_l[k]) if plan_rev_l[k] else 1.0
        nwc[k] = B["nwc"][k] * scale

    residual = {} if base_residual is None else base_residual
    path, prev = {}, B["cash"]["2026-06"]
    for i, k in enumerate(fwd):
        pk = bkeys[bkeys.index(k) - 1]
        dnwc = nwc[k] - nwc[pk]
        flow = S["ebitda"].get(k, 0.0) - S["tax"].get(k, 0.0) - S["capex"].get(k, 0.0) - dnwc
        if base_residual is None:                  # calibrating on the plan
            residual[k] = (B["cash"][k] - prev) - flow
        prev = prev + flow + residual.get(k, 0.0)
        path[k] = prev
    return path, residual
