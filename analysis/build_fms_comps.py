"""Build the McEasy FMS comparables benchmark workbook."""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

OUT = "McEasy_FMS_Comparables_Benchmark.xlsx"

# ---------------------------------------------------------------- style kit
F = "Arial"
BLUE = "0000FF"      # hardcoded input / source figure
BLACK = "000000"     # formula
GREEN = "008000"     # cross-sheet link
GREY = "808080"      # note / not disclosed
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
SEC_FILL = PatternFill("solid", fgColor="F2F2F2")
YEL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM = '#,##0.0;(#,##0.0);-'
NUM0 = '#,##0;(#,##0);-'
PCT = '0.0%;(0.0%);-'
CNT = '#,##0'
RATE = '0.00'
USD2 = '$#,##0.00'
USD0 = '$#,##0'


def title(ws, text, sub=None, width=None):
    ws["A1"] = text
    ws["A1"].font = Font(F, 15, bold=True, color="1F3864")
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(F, 9, italic=True, color=GREY)
    ws.sheet_view.showGridLines = False


def header_row(ws, row, labels, widths=None, height=32):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = Font(F, 9, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = height
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=2)


def put(ws, row, col, value, *, fmt=None, color=BLACK, bold=False, italic=False,
        size=10, wrap=False, halign=None, fill=None, note=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(F, size, bold=bold, italic=italic, color=color)
    if fmt:
        c.number_format = fmt
    if wrap or halign:
        c.alignment = Alignment(wrap_text=wrap, vertical="top",
                                horizontal=halign or "general")
    if fill:
        c.fill = fill
    c.border = BOX
    if note:
        c.comment = Comment(note, "Benchmark build")
    return c


def section(ws, row, text, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = SEC_FILL
        c.border = BOX
        if col == 1:
            c.value = text
            c.font = Font(F, 10, bold=True, color="1F3864")


def notes_block(ws, row, lines, ncols=8):
    put(ws, row, 1, "Notes & definitions", bold=True, size=10, color="1F3864")
    r = row + 1
    for ln in lines:
        c = ws.cell(row=r, column=1, value=ln)
        c.font = Font(F, 8.5, color=GREY)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        ws.row_dimensions[r].height = 26
        r += 1
    return r


wb = openpyxl.Workbook()

# ============================================================== ASSUMPTIONS
ws = wb.active
ws.title = "Assumptions"
title(ws, "Assumptions & FX",
      "Blue = hardcoded input you can change. Every FX-converted figure elsewhere in this "
      "workbook points back to these cells.")
header_row(ws, 4, ["Assumption", "Value", "Unit", "Basis / source"],
           widths=[46, 14, 16, 92])

# (key, label, value, unit, basis). `key` is how the rest of the script refers to the row,
# so row positions are derived below rather than hardcoded — adding a row here is safe.
rows = [
    ("FY2018", "USD/ZAR — average, Cartrack FY2018 (Mar-17 to Feb-18)", 12.90, "ZAR per USD",
     "Period average. Subscription revenue ZAR1,165.5m / 12.90 = USD90m — the year before "
     "Cartrack crossed USD100m subscription ARR."),
    ("FY2019", "USD/ZAR — average, Cartrack FY2019 (Mar-18 to Feb-19)", 13.70, "ZAR per USD",
     "Period average. Cross-check: FY2019 subscription revenue ZAR1,520.5m / 13.70 = USD111m, "
     "consistent with Cartrack crossing ~USD100m subscription ARR during FY2019."),
    ("FY2020", "USD/ZAR — average, FY2020 (Mar-19 to Feb-20)", 14.60, "ZAR per USD", "Period average."),
    ("FY2022", "USD/ZAR — average, FY2022 (Mar-21 to Feb-22)", 14.90, "ZAR per USD", "Period average."),
    ("FY2023", "USD/ZAR — average, FY2023 (Mar-22 to Feb-23)", 17.40, "ZAR per USD", "Period average."),
    ("FY2024", "USD/ZAR — average, FY2024 (Mar-23 to Feb-24)", 18.70, "ZAR per USD", "Period average."),
    ("FY2025", "USD/ZAR — average, FY2025 (Mar-24 to Feb-25)", 18.63, "ZAR per USD",
     "Implied by Karooooo's own disclosure: FY2025 EPS ZAR29.81 = USD1.60."),
    ("FY2026", "USD/ZAR — average, FY2026 (Mar-25 to Feb-26)", 15.92, "ZAR per USD",
     "Implied by Karooooo's own disclosure: FY2026 EPS ZAR32.17 = USD2.02. Cross-check: "
     "ARR ZAR5,179m = USD325m implies 15.94."),
    ("GBPUSD", "GBP/USD — average, calendar 2025", 1.3190, "USD per GBP",
     "2025 full-year average, 1.3190. Applied to Microlise and Quartix. Note this rate MULTIPLIES "
     "GBP to reach USD, whereas the ZAR and CNY rates DIVIDE."),
    ("USDCNY", "USD/CNY — average, calendar 2025", 7.1873, "CNY per USD",
     "2025 full-year average, 7.1873. Applied to Queclink and Streamax."),
    ("TGT_ARR", "McEasy target ARR", 100.0, "USD m", "User-stated 2030 goal."),
    ("TGT_YR", "McEasy target year", 2030, "calendar year", "User-stated."),
]
r = 5
AROW = {}
for key, lab, val, unit, basis in rows:
    AROW[key] = r
    put(ws, r, 1, lab, size=9, wrap=True)
    is_rate = unit.endswith("per USD") or unit.startswith("USD per")
    put(ws, r, 2, val, fmt=RATE if is_rate else (NUM if isinstance(val, float) else CNT),
        color=BLUE, bold=True, halign="center")
    put(ws, r, 3, unit, size=9, color=GREY, halign="center")
    put(ws, r, 4, basis, size=8.5, color=GREY, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1

FX = {k: f"$B${AROW[k]}" for k in
      ("FY2018", "FY2019", "FY2020", "FY2022", "FY2023", "FY2024", "FY2025", "FY2026")}
GBP = f"Assumptions!$B${AROW['GBPUSD']}"     # multiply GBP by this
CNY = f"Assumptions!$B${AROW['USDCNY']}"     # divide CNY by this
TGT_ARR = f"Assumptions!$B${AROW['TGT_ARR']}"
TGT_YR = f"Assumptions!$B${AROW['TGT_YR']}"

r += 1
section(ws, r, "Data-confidence legend", 4)
r += 1
for code, mean in [
    ("H — High", "Taken directly from an audited filing, SEC submission, or company press release."),
    ("M — Medium", "Derived by arithmetic from disclosed figures (FX conversion, margin calculation, "
                   "growth-rate roll-forward). Stated method, no judgement calls."),
    ("L — Low", "Third-party estimate (Revelio, Growjo, Latka, PitchBook) or a private company that "
                "does not publish financials. Directional only — do not put these in a board pack "
                "without a second source."),
    ("n/d", "Not disclosed and not reliably estimable. Left blank on purpose."),
]:
    put(ws, r, 1, code, bold=True, size=9)
    put(ws, r, 4, mean, size=8.5, color=GREY, wrap=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 30
    r += 1

# ============================================================ KAROOOOO PATH
kp = wb.create_sheet("Karooooo Path")
title(kp, "Cartrack / Karooooo — full trajectory, FY2019 to FY2026",
      "The single most useful table in this workbook: Cartrack crossed ~USD100m subscription ARR in "
      "FY2019 and is at USD325m ARR today. This is the actual path, not the destination.")

YEARS = ["FY2018", "FY2019", "FY2020", "FY2022", "FY2023", "FY2024", "FY2025", "FY2026"]
COLS = {y: 2 + i for i, y in enumerate(YEARS)}  # B..I
LAST = get_column_letter(COLS["FY2026"])          # I
V19 = get_column_letter(COLS["FY2019"])           # C
header_row(kp, 4, ["Metric"] + YEARS, widths=[52] + [13] * 8)

put(kp, 5, 1, "Year ended", bold=True, size=9)
for y, ends in zip(YEARS, ["Feb-2018", "Feb-2019", "Feb-2020", "Feb-2022", "Feb-2023",
                           "Feb-2024", "Feb-2025", "Feb-2026"]):
    put(kp, 5, COLS[y], ends, size=9, color=GREY, halign="center")

put(kp, 6, 1, "USD/ZAR average rate applied", size=9)
for y in YEARS:
    put(kp, 6, COLS[y], f"=Assumptions!{FX[y]}", fmt=RATE, color=GREEN, halign="center")

R = {}
r = 7
section(kp, r, "As reported (ZAR millions)", 9); r += 1

zar_rows = [
    ("rev", "Total revenue", {"FY2018": 1324.2, "FY2019": 1692.7, "FY2020": 1946.0,
                              "FY2022": 2746.0, "FY2023": 3507.0, "FY2024": 4206.0,
                              "FY2025": 4567.0, "FY2026": 5479.0}),
    ("sub", "Subscription revenue", {"FY2018": 1165.5, "FY2019": 1520.5, "FY2020": 1888.0,
                                     "FY2025": 4068.0, "FY2026": 4844.0}),
    ("gp", "Gross profit", {"FY2022": 1824.0, "FY2023": 2272.0, "FY2024": 2691.0,
                            "FY2025": 3203.0, "FY2026": 3724.0}),
    ("op", "Operating profit", {"FY2018": 434.3, "FY2019": 499.9, "FY2022": 699.1,
                                "FY2023": 881.9, "FY2024": 1043.0, "FY2025": 1312.0,
                                "FY2026": 1415.0}),
    ("np", "Profit for the year", {"FY2018": 310.5, "FY2019": 361.0, "FY2022": 450.0,
                                   "FY2023": 597.2, "FY2024": 738.2, "FY2025": 937.0,
                                   "FY2026": 1011.0}),
    ("eb", "EBITDA / adjusted EBITDA", {"FY2018": 650.8, "FY2019": 761.4,
                                        "FY2025": 1973.0, "FY2026": 2285.0}),
    ("ocf", "Cash generated from operations", {"FY2018": 467.8, "FY2019": 543.7,
                                               "FY2025": 1933.0, "FY2026": 1967.0}),
    ("cpx", "Capex (PP&E + contract assets)", {"FY2018": 420.1, "FY2019": 507.2,
                                               "FY2025": 1022.0, "FY2026": 1158.0}),
]
for key, lab, vals in zar_rows:
    put(kp, r, 1, lab, size=9)
    for y in YEARS:
        if y in vals:
            put(kp, r, COLS[y], vals[y], fmt=NUM, color=BLUE)
        else:
            put(kp, r, COLS[y], None, fmt=NUM, color=GREY)
    R[key] = r
    r += 1

put(kp, r, 1, "Free cash flow (operating cash flow less capex)", size=9)
for y in YEARS:
    a, b = f"{get_column_letter(COLS[y])}{R['ocf']}", f"{get_column_letter(COLS[y])}{R['cpx']}"
    put(kp, r, COLS[y], f'=IF(COUNT({a},{b})=2,{a}-{b},"")', fmt=NUM)
R["fcf"] = r; r += 1

put(kp, r, 1, "Company-reported \"adjusted free cash flow\"", size=9)
for y, v in [("FY2025", 425.0), ("FY2026", 809.0)]:
    put(kp, r, COLS[y], v, fmt=NUM, color=BLUE,
        note="Karooooo reported adjusted FCF of ZAR809m in FY2026, +90% YoY, implying ZAR425m in "
             "FY2025. Operating cash flow less capex per the cash-flow statement gives ZAR911m for "
             "FY2025 and ZAR809m for FY2026. The two definitions do not reconcile in FY2025. This "
             "workbook uses OCF-less-capex consistently for cross-company comparison.")
for y in YEARS:
    if y not in ("FY2025", "FY2026"):
        put(kp, r, COLS[y], None, fmt=NUM, color=GREY)
R["afcf"] = r; r += 1

put(kp, r, 1, "Cartrack SaaS ARR (ZAR m)", size=9)
for y, v in [("FY2025", 4384.0), ("FY2026", 5179.0)]:
    put(kp, r, COLS[y], v, fmt=NUM, color=BLUE)
for y in YEARS:
    if y not in ("FY2025", "FY2026"):
        put(kp, r, COLS[y], None, fmt=NUM, color=GREY)
R["arr_zar"] = r; r += 1

put(kp, r, 1, "Subscribers (period end)", size=9)
for y, v in [("FY2018", 751380), ("FY2019", 960798), ("FY2020", 1126515),
             ("FY2025", 2302236), ("FY2026", 2662222)]:
    put(kp, r, COLS[y], v, fmt=CNT, color=BLUE)
for y in YEARS:
    if y not in ("FY2018", "FY2019", "FY2020", "FY2025", "FY2026"):
        put(kp, r, COLS[y], None, fmt=CNT, color=GREY)
R["subs"] = r; r += 1

r += 1
section(kp, r, "Converted to USD millions (at the average rate in row 6)", 9); r += 1
for key, lab, src in [("rev_u", "Total revenue", "rev"),
                      ("sub_u", "Subscription revenue", "sub"),
                      ("arr_u", "Cartrack SaaS ARR", "arr_zar"),
                      ("eb_u", "EBITDA / adjusted EBITDA", "eb"),
                      ("fcf_u", "Free cash flow (OCF less capex)", "fcf")]:
    put(kp, r, 1, lab, size=9)
    for y in YEARS:
        cl = get_column_letter(COLS[y])
        put(kp, r, COLS[y], f'=IF(COUNT({cl}{R[src]},{cl}6)=2,{cl}{R[src]}/{cl}6,"")', fmt=NUM)
    R[key] = r; r += 1

r += 1
section(kp, r, "Margins", 9); r += 1
for lab, num, den in [("Gross margin", "gp", "rev"),
                      ("Operating margin", "op", "rev"),
                      ("Net margin", "np", "rev"),
                      ("EBITDA margin", "eb", "rev"),
                      ("Free-cash-flow margin", "fcf", "rev"),
                      ("Capex as % of revenue", "cpx", "rev"),
                      ("Subscription revenue as % of total", "sub", "rev")]:
    put(kp, r, 1, lab, size=9)
    for y in YEARS:
        cl = get_column_letter(COLS[y])
        put(kp, r, COLS[y],
            f'=IF(COUNT({cl}{R[num]},{cl}{R[den]})=2,{cl}{R[num]}/{cl}{R[den]},"")', fmt=PCT)
    R["m_" + num + den] = r; r += 1

r += 1
section(kp, r, "Unit economics", 9); r += 1
put(kp, r, 1, "Revenue per subscriber (USD per year)", size=9)
for y in YEARS:
    cl = get_column_letter(COLS[y])
    put(kp, r, COLS[y], f'=IF(COUNT({cl}{R["rev_u"]},{cl}{R["subs"]})=2,'
                        f'{cl}{R["rev_u"]}*1000000/{cl}{R["subs"]},"")', fmt=USD0)
R["rps"] = r; r += 1
put(kp, r, 1, "Subscription revenue per subscriber (USD per month)", size=9)
for y in YEARS:
    cl = get_column_letter(COLS[y])
    put(kp, r, COLS[y], f'=IF(COUNT({cl}{R["sub_u"]},{cl}{R["subs"]})=2,'
                        f'{cl}{R["sub_u"]}*1000000/{cl}{R["subs"]}/12,"")', fmt=USD2)
R["arpu_m"] = r; r += 1
put(kp, r, 1, "Net subscriber additions in the year", size=9)
for i, y in enumerate(YEARS):
    cl = get_column_letter(COLS[y])
    if i == 0:
        put(kp, r, COLS[y], None, fmt=CNT, color=GREY)
    else:
        pc = get_column_letter(COLS[YEARS[i - 1]])
        put(kp, r, COLS[y], f'=IF(COUNT({cl}{R["subs"]},{pc}{R["subs"]})=2,'
                            f'{cl}{R["subs"]}-{pc}{R["subs"]},"")', fmt=CNT)
R["adds"] = r; r += 1

r += 1
notes_block(kp, r, [
    "FY2021 (year ended Feb-2021) is deliberately blank — it was not retrieved from a primary source. "
    "FY2020 total revenue is derived: subscription revenue of ZAR1,888m was disclosed as 97% of total. "
    "FY2018 is included to show the crossing: subscription revenue went from USD90m to USD111m.",
    "Karooooo Ltd. was interposed above Cartrack Holdings Ltd. ahead of the April 2021 NASDAQ listing. "
    "FY2019 and FY2020 are Cartrack Holdings (JSE-listed); FY2022 onward are Karooooo Ltd. The "
    "operating business is continuous, but the reporting entity and accounting basis are not identical.",
    "Gross profit is blank for FY2018-FY2020 because Cartrack Holdings reported under a cost "
    "classification that did not disclose it. The FY2019 SENS release also gave segment EBITDA "
    "rather than segment operating profit.",
    "Read the FY2019 column as the closest available analogue to McEasy at a USD100m ARR milestone.",
], ncols=9)

# ======================================================== REVENUE BY PRODUCT
rp = wb.create_sheet("Revenue by Product")
title(rp, "Revenue mix by product line",
      "Only Karooooo, Powerfleet and Ituran disclose a usable split. Motive discloses recurring vs "
      "other but not by product. Geotab, Lytx and Gurtam disclose nothing.")
header_row(rp, 4, ["Company", "Fiscal period", "Revenue category / product line", "Currency",
                   "Amount (m)", "% of company total", "Confidence", "Note"],
           widths=[24, 18, 42, 10, 14, 15, 12, 62])

prod = [
    ("Karooooo (Cartrack)", "FY2026 (Feb-26)", "Cartrack — subscription (SaaS telematics)", "ZAR", 4831.0, "H", "Disclosed segment figure."),
    ("Karooooo (Cartrack)", "FY2026 (Feb-26)", "Cartrack — hardware, installation & other", "ZAR", 108.0, "M", "Cartrack total revenue ZAR4,939m less subscription ZAR4,831m."),
    ("Karooooo (Cartrack)", "FY2026 (Feb-26)", "Karooooo Logistics — Delivery-as-a-Service", "ZAR", 540.0, "H", "Disclosed segment. Gross margin only 31% vs 72% for Cartrack."),
    ("Karooooo (Cartrack)", "FY2019 (Feb-19)", "Subscription (SaaS telematics)", "ZAR", 1520.5, "H", "The USD100m-ARR vintage — subscription was already ~90% of revenue."),
    ("Karooooo (Cartrack)", "FY2019 (Feb-19)", "Hardware, installation & other", "ZAR", 172.2, "M", "Total revenue ZAR1,692.7m less subscription ZAR1,520.5m."),
    ("Ituran", "FY2025 (Dec-25)", "Telematics services — subscription fees", "USD", 265.7, "M", "Company disclosed a 74% / 26% services-versus-products split of USD359.0m."),
    ("Ituran", "FY2025 (Dec-25)", "Products — hardware sales", "USD", 93.3, "M", "26% of USD359.0m."),
    ("Powerfleet", "FY2026 (Mar-26)", "Services (recurring: SaaS + connectivity)", "USD", 359.8, "H", "Grew 30% YoY."),
    ("Powerfleet", "FY2026 (Mar-26)", "Products (hardware)", "USD", 84.0, "H", "Declined slightly YoY — deliberate shift toward recurring."),
    ("Motive", "LTM to Sep-25", "Recurring subscription revenue", "USD", 416.1, "M", "S-1 states recurring revenue is 97% of total; applied to LTM revenue of USD429m."),
    ("Motive", "LTM to Sep-25", "Other (services, hardware)", "USD", 12.9, "M", "Residual 3%."),
    # --- players currently at USD80-250m revenue -------------------------------
    ("Microlise", "FY2025 (Dec-25)", "Recurring revenue (subscription / SaaS)", "GBP", 58.8, "H", "Grew 7.5%, of which direct customers +16%. 70% of revenue."),
    ("Microlise", "FY2025 (Dec-25)", "Non-recurring (hardware, professional services)", "GBP", 25.2, "H", "Declined 4.3%. Global OEM customers fell to 27% of group revenue from 33%."),
    ("Quartix", "FY2025 (Dec-25)", "Subscription — Total Fleet and Insurance", "GBP", 35.71, "H", "Effectively all revenue. Reportable segments are Insurance and Total Fleet; no hardware line is broken out."),
    ("CalAmp", "CY2024", "Software & Subscription Services", "USD", 126.9, "M", "Derived: Q3-FY2024 mix of 64.4% S&SS / 35.8% products applied to CY2024 revenue of USD197m."),
    ("CalAmp", "CY2024", "Telematics products (hardware)", "USD", 70.1, "M", "Residual 35.6%."),
    ("Queclink", "FY2025 (Dec-25)", "On-board information intelligent terminal", "CNY", 471.76, "H", "52.9% of revenue. Core vehicle tracking device line."),
    ("Queclink", "FY2025 (Dec-25)", "Asset management intelligent terminal", "CNY", 241.39, "H", "27.1%."),
    ("Queclink", "FY2025 (Dec-25)", "Other products", "CNY", 68.42, "H", "7.7%."),
    ("Queclink", "FY2025 (Dec-25)", "Video connected-vehicle products", "CNY", 48.46, "H", "5.4%. The AI-dashcam equivalent line."),
    ("Queclink", "FY2025 (Dec-25)", "Animal tracking / traceability", "CNY", 46.05, "H", "5.2%."),
    ("Queclink", "FY2025 (Dec-25)", "Two-wheeler smart terminal", "CNY", 15.02, "H", "1.7%. Directly relevant to Indonesian two-wheeler fleets."),
]
r = 5
PROD_FIRST = r
PROD_LAST_EST = r + len(prod) - 1
for co, per, cat, cur, amt, conf, note in prod:
    put(rp, r, 1, co, size=9)
    put(rp, r, 2, per, size=9, halign="center")
    put(rp, r, 3, cat, size=9, wrap=True)
    put(rp, r, 4, cur, size=9, halign="center")
    put(rp, r, 5, amt, fmt=NUM, color=BLUE)
    put(rp, r, 6, f'=IFERROR(E{r}/SUMIFS($E${PROD_FIRST}:$E${PROD_LAST_EST},'
                  f'$A${PROD_FIRST}:$A${PROD_LAST_EST},$A{r},'
                  f'$B${PROD_FIRST}:$B${PROD_LAST_EST},$B{r}),"")', fmt=PCT)
    put(rp, r, 7, conf, size=9, bold=True, halign="center")
    put(rp, r, 8, note, size=8.5, color=GREY, wrap=True)
    rp.row_dimensions[r].height = 26
    r += 1

PROD_LAST = r - 1
assert PROD_LAST == PROD_LAST_EST, (PROD_LAST, PROD_LAST_EST)
for co, per, txt in [
    ("Geotab", "CY2024", "Not disclosed. Private, bootstrapped. Third-party revenue estimate only; no product-line split published."),
    ("Lytx", "CY2025", "Not disclosed. Permira-owned; video-based safety is the core line but no split published."),
    ("Gurtam", "CY2025", "Not disclosed. Reseller/channel model on the Wialon platform; revenue is platform licence fees, not end-customer subscriptions."),
    ("Streamax", "FY2025 (Dec-25)", "Not disclosed at this granularity. AI video / AIoT fleet-safety hardware and platform; 65% of revenue is overseas."),
    ("ORBCOMM", "TTM Sep-25", "Not disclosed. Private since the 2021 GI Partners buyout."),
    ("Teletrac Navman", "CY2026", "Not disclosed. Was never separately segment-reported inside Vontier."),
    ("Netradyne", "CY2025", "Not disclosed. Single product family (AI video safety, Driveri); ARR is effectively all of revenue."),
    ("Fleetio", "CY2025", "Not disclosed. Fleet maintenance SaaS plus the acquired Auto Integrate maintenance-authorisation network."),
]:
    put(rp, r, 1, co, size=9)
    put(rp, r, 2, per, size=9, halign="center")
    put(rp, r, 3, "n/d", size=9, color=GREY, halign="center")
    for c in (4, 5, 6):
        put(rp, r, c, None, color=GREY)
    put(rp, r, 7, "n/d", size=9, bold=True, halign="center", color=GREY)
    put(rp, r, 8, txt, size=8.5, color=GREY, wrap=True)
    rp.row_dimensions[r].height = 26
    r += 1

r += 1
notes_block(rp, r, [
    "The pattern to take away: every profitable comparable is 74–98% recurring revenue, and the "
    "hardware line is flat-to-shrinking by design. Powerfleet's product revenue fell in absolute "
    "terms while services grew 30%.",
    "Karooooo's Delivery-as-a-Service line is a cautionary case: 10% of revenue at a 31% gross "
    "margin against 72% for the core telematics SaaS. Adjacent logistics services dilute the "
    "multiple even when they grow faster.",
], ncols=8)

# ======================================================== REVENUE BY COUNTRY
rc = wb.create_sheet("Revenue by Country")
title(rc, "Revenue mix by country / region",
      "Karooooo FY2019 and FY2025 are disclosed; FY2026 is rolled forward from disclosed regional "
      "growth rates. Powerfleet does not disclose geography in its results release.")
header_row(rc, 4, ["Company", "Fiscal period", "Region / country", "Currency", "Revenue (m)",
                   "% of total (computed)", "Growth applied", "% of total (as disclosed)",
                   "Confidence", "Note"],
           widths=[22, 18, 34, 10, 13, 15, 12, 14, 11, 58])

FIRST = 5
geo_actual = [
    ("Karooooo (Cartrack)", "FY2019 (Feb-19)", "South Africa", "ZAR", 1245.7, None, None, "H", "73.6% of revenue — home market dominance at the USD100m-ARR milestone."),
    ("Karooooo (Cartrack)", "FY2019 (Feb-19)", "Asia-Pacific & Middle East", "ZAR", 179.7, None, None, "H", "Segment EBITDA only ZAR16m — Asia was barely profitable at this stage."),
    ("Karooooo (Cartrack)", "FY2019 (Feb-19)", "Europe", "ZAR", 147.6, None, None, "H", ""),
    ("Karooooo (Cartrack)", "FY2019 (Feb-19)", "Africa — other", "ZAR", 115.6, None, None, "H", ""),
    ("Karooooo (Cartrack)", "FY2019 (Feb-19)", "USA", "ZAR", 4.0, None, None, "H", "Loss-making sub-scale entry."),
    ("Karooooo (Cartrack)", "FY2025 (Feb-25)", "South Africa", "ZAR", 3360.0, None, None, "H", "73.7% — essentially unchanged from FY2019 six years later."),
    ("Karooooo (Cartrack)", "FY2025 (Feb-25)", "Asia-Pacific, Middle East & USA", "ZAR", 659.6, None, None, "H", ""),
    ("Karooooo (Cartrack)", "FY2025 (Feb-25)", "Europe", "ZAR", 399.2, None, None, "H", ""),
    ("Karooooo (Cartrack)", "FY2025 (Feb-25)", "Africa — other", "ZAR", 143.8, None, None, "H", ""),
    ("Queclink", "FY2025 (Dec-25)", "Outside China", "CNY", 840.69, None, None, "H", "94.3% of revenue earned outside its home market — the mirror image of Cartrack. A device vendor exports; an FMS operator does not."),
    ("Queclink", "FY2025 (Dec-25)", "China", "CNY", 50.41, None, None, "H", "5.7%."),
]
r = FIRST
fy25_rows = {}
for co, per, reg, cur, amt, _g, _d, conf, note in geo_actual:
    put(rc, r, 1, co, size=9)
    put(rc, r, 2, per, size=9, halign="center")
    put(rc, r, 3, reg, size=9, wrap=True)
    put(rc, r, 4, cur, size=9, halign="center")
    put(rc, r, 5, amt, fmt=NUM, color=BLUE)
    put(rc, r, 7, None, color=GREY)
    put(rc, r, 8, None, fmt=PCT, color=GREY)
    put(rc, r, 9, conf, size=9, bold=True, halign="center")
    put(rc, r, 10, note, size=8.5, color=GREY, wrap=True)
    rc.row_dimensions[r].height = 26
    if per.startswith("FY2025"):
        fy25_rows[reg] = r
    r += 1

# FY2026 estimate rolled forward from FY2025 actuals
fy26_map = [("South Africa", "South Africa", 0.20,
             "Disclosed: South Africa subscription revenue grew 20% in FY2026."),
            ("Asia-Pacific, Middle East & USA", "Asia-Pacific, Middle East & USA", 0.17,
             "Disclosed: Asia & Middle East subscription revenue grew 17% (20% constant currency)."),
            ("Europe", "Europe", 0.22,
             "Disclosed: Europe subscription revenue grew 22% (19% constant currency)."),
            ("Africa — other", "Africa — other", 0.01,
             "No revenue growth disclosed; 1% subscriber growth used as a proxy. Lowest-confidence line.")]
fy26_rows = []
for reg, src, g, note in fy26_map:
    put(rc, r, 1, "Karooooo (Cartrack)", size=9)
    put(rc, r, 2, "FY2026E (Feb-26)", size=9, halign="center")
    put(rc, r, 3, reg, size=9, wrap=True)
    put(rc, r, 4, "ZAR", size=9, halign="center")
    put(rc, r, 5, f"=E{fy25_rows[src]}*(1+G{r})", fmt=NUM)
    put(rc, r, 7, g, fmt=PCT, color=BLUE, halign="center")
    put(rc, r, 8, None, fmt=PCT, color=GREY)
    put(rc, r, 9, "M", size=9, bold=True, halign="center")
    put(rc, r, 10, note, size=8.5, color=GREY, wrap=True)
    rc.row_dimensions[r].height = 26
    fy26_rows.append(r)
    r += 1

GEO_LAST = r - 1
for rr in range(FIRST, GEO_LAST + 1):
    put(rc, rr, 6, f'=IFERROR(E{rr}/SUMIFS($E${FIRST}:$E${GEO_LAST},'
                   f'$A${FIRST}:$A${GEO_LAST},$A{rr},$B${FIRST}:$B${GEO_LAST},$B{rr}),"")', fmt=PCT)

# reconciliation check
r += 1
put(rc, r, 3, "FY2026E regional total (check)", size=9, bold=True)
put(rc, r, 5, f"=SUM(E{fy26_rows[0]}:E{fy26_rows[-1]})", fmt=NUM, bold=True)
KP_REV_LAST = f"'Karooooo Path'!{LAST}{R['rev']}"
put(rc, r, 10, "Compare with reported group revenue below. A gap is expected: regional growth rates "
               "are for subscription revenue, applied here to total revenue.", size=8.5,
    color=GREY, wrap=True)
CHK1 = r; r += 1
put(rc, r, 3, "FY2026 reported group revenue", size=9, bold=True)
put(rc, r, 5, f"={KP_REV_LAST}", fmt=NUM, color=GREEN, bold=True)
CHK2 = r; r += 1
put(rc, r, 3, "Estimate as % of reported", size=9, bold=True)
put(rc, r, 5, f"=IFERROR(E{CHK1}/E{CHK2},\"\")", fmt=PCT, bold=True)
put(rc, r, 10, "Above ~97% means the roll-forward is a fair approximation of the true mix.",
    size=8.5, color=GREY, wrap=True)
r += 2

# other companies' geography
section(rc, r, "Other comparables", 10); r += 1
oth_first = r
oth = [
    ("Ituran", "FY2025 (Dec-25)", "Israel", 0.55, "H", "Q4-2025 disclosed mix, used as the full-year proxy."),
    ("Ituran", "FY2025 (Dec-25)", "Brazil", 0.23, "H", "Second home market; the LatAm stolen-vehicle-recovery engine."),
    ("Ituran", "FY2025 (Dec-25)", "Rest of world", 0.22, "H", "Argentina, Mexico, USA, Colombia, Ecuador."),
    ("Motive", "LTM to Sep-25", "United States", 0.90, "M", "S-1 states approximately 90% US revenue."),
    ("Motive", "LTM to Sep-25", "International", 0.10, "M", "Canada and Mexico principally; not broken out."),
    ("Streamax", "FY2025 (Dec-25)", "Overseas (ex-China)", 0.65, "M", "Company states 65% overseas. Large Southeast Asian installed base; no country split published."),
    ("Streamax", "FY2025 (Dec-25)", "China", 0.35, "M", "Residual."),
]
OTH_CO = []                           # company per row, so the fix-up below is not positional
for co, per, reg, share, conf, note in oth:
    OTH_CO.append(co)
    put(rc, r, 1, co, size=9)
    put(rc, r, 2, per, size=9, halign="center")
    put(rc, r, 3, reg, size=9, wrap=True)
    put(rc, r, 4, "USD", size=9, halign="center")
    put(rc, r, 8, share, fmt=PCT, color=BLUE, halign="center")
    put(rc, r, 9, conf, size=9, bold=True, halign="center")
    put(rc, r, 10, note, size=8.5, color=GREY, wrap=True)
    put(rc, r, 7, None, color=GREY)
    rc.row_dimensions[r].height = 26
    r += 1
OTH_ROWS = list(range(oth_first, r))

for co, per, txt in [
    ("Powerfleet", "FY2026 (Mar-26)", "n/d in the results release. Operates in North America, Israel, "
     "Europe, South Africa, Australia and Latin America following the MiX Telematics and Fleet "
     "Complete mergers. Geography is in the 10-K segment note only."),
    ("Geotab", "CY2024", "n/d. North-America-weighted; 55,000+ customers and 5m+ subscriptions globally."),
    ("Lytx", "CY2025", "n/d. North-America-weighted."),
    ("Gurtam", "CY2025", "n/d. Reseller network across CIS, Europe, Middle East, Africa, LatAm and Asia."),
    ("Microlise", "FY2025 (Dec-25)", "n/d as a split. Operations in the UK, France, Australia and India "
     "serving 2,500+ clients; management states growth was broad-based across all geographies. The "
     "only mix it does disclose is channel: global OEM customers fell to 27% of revenue from 33%."),
    ("Quartix", "FY2025 (Dec-25)", "n/d as percentages. Disclosed order of size only: UK largest, then "
     "France, then USA, then other European territories."),
    ("CalAmp", "CY2024", "n/d. Went private via Chapter 11 in June 2024 and stopped filing; the last "
     "10-K geographic note predates the restructuring."),
    ("Netradyne", "CY2025", "n/d. US-weighted with a large India engineering base."),
    ("ORBCOMM", "TTM Sep-25", "n/d. Private since 2021."),
    ("Teletrac Navman", "CY2026", "n/d. Historically US, Australia, New Zealand and the UK."),
    ("Fleetio", "CY2025", "n/d. North-America-centred: USA, Canada and Mexico repair-shop network, "
     "with customers in 100+ countries."),
]:
    put(rc, r, 1, co, size=9)
    put(rc, r, 2, per, size=9, halign="center")
    put(rc, r, 3, "n/d", size=9, color=GREY, halign="center")
    for c in (4, 5, 6, 7, 8):
        put(rc, r, c, None, color=GREY)
    put(rc, r, 9, "n/d", size=9, bold=True, halign="center", color=GREY)
    put(rc, r, 10, txt, size=8.5, color=GREY, wrap=True)
    rc.row_dimensions[r].height = 30
    r += 1

r += 1
section(rc, r, "Cartrack subscribers by region (disclosed — highest-confidence geographic cut)", 10)
r += 1
sub_hdr = r
for i, lab in enumerate(["Region", "FY2026 subscribers", "FY2025 subscribers", "Growth",
                         "FY2026 share of base"], start=1):
    put(rc, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL)
r += 1
sub_first = r
for reg, a, b in [("South Africa", 2005888, 1736542),
                  ("Asia-Pacific & Middle East", 335907, 273946),
                  ("Europe", 228384, 200774),
                  ("Africa — other", 92043, 90974)]:
    put(rc, r, 1, reg, size=9)
    put(rc, r, 2, a, fmt=CNT, color=BLUE)
    put(rc, r, 3, b, fmt=CNT, color=BLUE)
    put(rc, r, 4, f'=IFERROR(B{r}/C{r}-1,"")', fmt=PCT)
    r += 1
sub_last = r - 1
put(rc, r, 1, "Total", bold=True, size=9)
put(rc, r, 2, f"=SUM(B{sub_first}:B{sub_last})", fmt=CNT, bold=True)
put(rc, r, 3, f"=SUM(C{sub_first}:C{sub_last})", fmt=CNT, bold=True)
put(rc, r, 4, f'=IFERROR(B{r}/C{r}-1,"")', fmt=PCT, bold=True)
sub_tot = r
for rr in range(sub_first, sub_last + 1):
    put(rc, rr, 5, f'=IFERROR(B{rr}/$B${sub_tot},"")', fmt=PCT)
put(rc, r, 5, f'=IFERROR(B{r}/$B${sub_tot},"")', fmt=PCT, bold=True)
r += 2

notes_block(rc, r, [
    "The headline finding for McEasy: Cartrack's home market was 73.6% of revenue at USD100m ARR "
    "and is still ~74% at USD325m ARR. Geographic diversification did NOT fund the path from "
    "USD100m to USD325m — deepening penetration of the home market did. Southeast Asia is 12.6% of "
    "the subscriber base after roughly fifteen years of presence.",
    "Management explicitly flags that Southeast Asian ARPU (Indonesia, Philippines, Thailand) is "
    "materially below South African ARPU, and that a rising SEA mix will dilute group ARPU. This is "
    "the single most important structural warning in the dataset for an Indonesia-based company.",
], ncols=10)

# ======================================================== CURRENT POSITION
cp = wb.create_sheet("Current Position")
title(cp, "Comparables — current position",
      "Most recent reported full year for each company. All figures in USD millions. "
      "Karooooo cells are live links to the 'Karooooo Path' sheet.")
header_row(cp, 4, ["Company", "Listing / ownership", "Fiscal period", "ARR (USD m)",
                   "Total revenue (USD m)", "Recurring revenue (USD m)", "Recurring % of revenue",
                   "Revenue growth YoY", "Subscribers / connected assets", "Employees",
                   "Revenue per employee (USD k)", "Gross margin", "Operating margin",
                   "Net margin", "EBITDA margin", "FCF (USD m)", "FCF margin",
                   "Confidence", "Notes"],
           widths=[22, 22, 16, 12, 13, 13, 12, 11, 15, 11, 13, 11, 11, 11, 11, 11, 11, 11, 70])

CP_FIRST = 5
K = "'Karooooo Path'!"
PREV = get_column_letter(COLS["FY2025"])
# Karooooo row -- links
r = CP_FIRST
put(cp, r, 1, "Karooooo (Cartrack)", size=9, bold=True)
put(cp, r, 2, "NASDAQ: KARO", size=9)
put(cp, r, 3, "FY2026 (Feb-26)", size=9, halign="center")
put(cp, r, 4, f"={K}{LAST}{R['arr_u']}", fmt=NUM, color=GREEN)
put(cp, r, 5, f"={K}{LAST}{R['rev_u']}", fmt=NUM, color=GREEN)
put(cp, r, 6, f"={K}{LAST}{R['sub_u']}", fmt=NUM, color=GREEN)
put(cp, r, 7, f'=IF(COUNT(F{r},E{r})=2,F{r}/E{r},"")', fmt=PCT)
put(cp, r, 8, f'=IF(COUNT({K}{LAST}{R["rev"]},{K}{PREV}{R["rev"]})=2,'
              f'{K}{LAST}{R["rev"]}/{K}{PREV}{R["rev"]}-1,"")', fmt=PCT)
put(cp, r, 9, f"={K}{LAST}{R['subs']}", fmt=CNT, color=GREEN)
put(cp, r, 10, 5000, fmt=CNT, color=BLUE,
    note="Not precisely disclosed. Company transcript said >5,000 for FY2025; Revelio Labs "
         "estimates 4,550 globally as at Mar-2026. Treat as +/- 15%.")
put(cp, r, 11, f'=IF(COUNT(E{r},J{r})=2,E{r}*1000/J{r},"")', fmt=USD0)
put(cp, r, 12, f'={K}{LAST}{R["m_gprev"]}', fmt=PCT, color=GREEN)
put(cp, r, 13, f'={K}{LAST}{R["m_oprev"]}', fmt=PCT, color=GREEN)
put(cp, r, 14, f'={K}{LAST}{R["m_nprev"]}', fmt=PCT, color=GREEN)
put(cp, r, 15, f'={K}{LAST}{R["m_ebrev"]}', fmt=PCT, color=GREEN)
put(cp, r, 16, f"={K}{LAST}{R['fcf_u']}", fmt=NUM, color=GREEN)
put(cp, r, 17, f'=IF(COUNT(P{r},E{r})=2,P{r}/E{r},"")', fmt=PCT)
put(cp, r, 18, "H", size=9, bold=True, halign="center")
put(cp, r, 19, "The only comparable that is simultaneously growing ARR ~20%+ in local currency, "
               "GAAP-profitable at an 18% net margin, and free-cash-flow positive. Employees are "
               "the one soft number. Dividend-paying.", size=8.5, color=GREY, wrap=True)
cp.row_dimensions[r].height = 42
KARO_ROW = r
r += 1

others = [
    dict(name="Ituran", own="NASDAQ: ITRN", per="FY2025 (Dec-25)", arr=None, rev=359.0,
         rec=265.7, growth=0.07, subs=2630000, emp=None, gm=None, om=0.218, nm=58.0/359.0,
         eb=96.2/359.0, fcf=None, conf="H",
         note="Closest profitability profile to Karooooo. 16% net margin, 27% EBITDA margin, "
              "record USD88.6m operating cash flow. Growth is only 7% — this is a mature, "
              "cash-returning business, not a growth story. Israel + Brazil = 78% of revenue."),
    dict(name="Powerfleet", own="NASDAQ: AIOT", per="FY2026 (Mar-26)", arr=None, rev=443.8,
         rec=359.8, growth=0.22, subs=2600000, emp=None, gm=0.555, om=19.6/443.8,
         nm=-20.6/443.8, eb=0.219, fcf=-9.6, conf="H",
         note="Roll-up of I.D. Systems, Pointer, MiX Telematics and Fleet Complete. 22% revenue "
              "growth is largely acquired, not organic. Still GAAP loss-making with net debt of "
              "USD239m at 2.47x EBITDA. A warning about buying your way to scale."),
    dict(name="Motive", own="Private; S-1 filed, NYSE: MTVE pending", per="LTM to Sep-25",
         arr=501.0, rev=429.0, rec=416.1, growth=0.21, subs=None, emp=4508, gm=0.70,
         om=-0.17, nm=-138.5/327.3, eb=None, fcf=-78.7, conf="H",
         note="ARR USD501m at 28% growth but burning cash: -23% FCF margin, USD138.5m net loss "
              "over nine months. IPO filed 23-Dec-2025, still unpriced as at Jul-2026. Net dollar "
              "retention 110% core / 126% large. 90% US revenue. 78% of staff are offshore."),
    dict(name="Geotab", own="Private, bootstrapped", per="CY2024", arr=None, rev=681.0,
         rec=None, growth=None, subs=5000000, emp=2900, gm=None, om=None, nm=None, eb=None,
         fcf=None, conf="L",
         note="Largest pure-play by subscriptions (5m+) and never took outside capital. Revenue "
              "is a third-party estimate. 2,900+ staff. No margin data exists publicly."),
    dict(name="Lytx", own="Private (Permira)", per="CY2025", arr=None, rev=None, rec=None,
         growth=None, subs=5500000, emp=1082, gm=None, om=None, nm=None, eb=None, fcf=None,
         conf="L",
         note="Video-based driver safety specialist. Public revenue estimates span USD100-500m — "
              "too wide to be useful. Included for product-strategy context, not financial "
              "benchmarking."),
    dict(name="Gurtam", own="Private", per="CY2025", arr=None, rev=37.3, rec=None,
         growth=None, subs=4000000, emp=339, gm=None, om=None, nm=None, eb=None, fcf=None,
         conf="L",
         note="Platform/reseller model (Wialon), not a direct fleet operator. USD37.3m revenue on "
              "339 staff implies ~USD110k revenue per employee — the highest here. Relevant as a "
              "capital-light alternative model, not as an ARR comparable."),
]
for d in others:
    put(cp, r, 1, d["name"], size=9, bold=True)
    put(cp, r, 2, d["own"], size=9)
    put(cp, r, 3, d["per"], size=9, halign="center")
    put(cp, r, 4, d["arr"], fmt=NUM, color=BLUE if d["arr"] else GREY)
    put(cp, r, 5, d["rev"], fmt=NUM, color=BLUE if d["rev"] else GREY)
    put(cp, r, 6, d["rec"], fmt=NUM, color=BLUE if d["rec"] else GREY)
    put(cp, r, 7, f'=IF(COUNT(F{r},E{r})=2,F{r}/E{r},"")', fmt=PCT)
    put(cp, r, 8, d["growth"], fmt=PCT, color=BLUE if d["growth"] else GREY)
    put(cp, r, 9, d["subs"], fmt=CNT, color=BLUE if d["subs"] else GREY)
    put(cp, r, 10, d["emp"], fmt=CNT, color=BLUE if d["emp"] else GREY)
    put(cp, r, 11, f'=IF(COUNT(E{r},J{r})=2,E{r}*1000/J{r},"")', fmt=USD0)
    put(cp, r, 12, d["gm"], fmt=PCT, color=BLUE if d["gm"] is not None else GREY)
    put(cp, r, 13, d["om"], fmt=PCT, color=BLUE if d["om"] is not None else GREY)
    put(cp, r, 14, d["nm"], fmt=PCT, color=BLUE if d["nm"] is not None else GREY)
    put(cp, r, 15, d["eb"], fmt=PCT, color=BLUE if d["eb"] is not None else GREY)
    put(cp, r, 16, d["fcf"], fmt=NUM, color=BLUE if d["fcf"] is not None else GREY)
    put(cp, r, 17, f'=IF(COUNT(P{r},E{r})=2,P{r}/E{r},"")', fmt=PCT)
    put(cp, r, 18, d["conf"], size=9, bold=True, halign="center")
    put(cp, r, 19, d["note"], size=8.5, color=GREY, wrap=True)
    cp.row_dimensions[r].height = 46
    r += 1
CP_LAST = r - 1
ITU_ROW = CP_FIRST + 1
MOT_ROW = CP_FIRST + 3

r += 1
notes_block(cp, r, [
    "ARR is only reported by Karooooo and Motive. Ituran, Powerfleet, Geotab, Lytx and Gurtam do "
    "not publish an ARR figure, so revenue is the only like-for-like top line across the set.",
    "Ituran operating margin is the Q4-2025 rate (21.8%) used as a full-year proxy; the full-year "
    "figure was not disclosed in the results release. Ituran FCF is blank: operating cash flow was "
    "USD88.6m but capex was not disclosed, so an FCF margin cannot be computed honestly.",
    "Powerfleet FCF is the sum of disclosed halves (-USD13.7m H1, +USD4.1m H2). Motive net margin "
    "uses the nine-month loss over nine-month revenue, since no full-year figure exists yet.",
    "Recurring % of revenue is blank wherever recurring revenue is not disclosed. It is not zero.",
], ncols=19)

# fix the percentage-only geography links now that the source rows are known.
# Streamax is resolved after the 'At 100M (Today)' sheet is built, further below.
GEO_REV_REF = {"Ituran": f"'Current Position'!$E${ITU_ROW}",
               "Motive": f"'Current Position'!$E${MOT_ROW}"}
for rr, co in zip(OTH_ROWS, OTH_CO):
    if co in GEO_REV_REF:
        put(rc, rr, 5, f'=IFERROR(H{rr}*{GEO_REV_REF[co]},"")', fmt=NUM, color=GREEN)
        put(rc, rr, 6, f'=IFERROR(H{rr},"")', fmt=PCT)

# ==================================================== AT 100M (TODAY)
td = wb.create_sheet("At 100M (Today)")
title(td, "Who is at this size RIGHT NOW — FMS / telematics, revenue USD80-250m",
      "Same columns as 'Current Position' so the two read side by side. Native reporting currency is "
      "kept alongside USD so nothing is hidden inside an FX conversion.")
header_row(td, 4, ["Company", "Listing / ownership", "Fiscal period", "Currency",
                   "ARR (USD m)", "Distance from USD100m ARR", "Total revenue (USD m)",
                   "Recurring revenue (USD m)", "Recurring % of revenue", "Revenue growth YoY",
                   "Subscribers / units", "Employees", "Revenue per employee (USD k)",
                   "Gross margin", "Operating margin", "Net margin", "EBITDA margin",
                   "FCF (USD m)", "FCF margin", "Confidence", "Notes"],
           widths=[22, 26, 15, 9, 11, 13, 12, 12, 11, 11, 13, 10, 12, 10, 11, 10, 11, 10, 10, 10, 78])

TD_FIRST = 5
r = TD_FIRST
section(td, r, "Block A — subscription-FMS operators, revenue USD80-250m (the primary comparison)", 21)
r += 1


def td_row(r, d):
    """Write one company row. GBP/CNY figures are converted by formula, never pre-computed."""
    cv = d.get("cur", "USD")
    def conv(v):
        """Return a formula (or literal) that lands the native figure in USD millions."""
        if v is None:
            return None
        if cv == "GBP":
            return f"={v}*{GBP}"
        if cv == "CNY":
            return f"={v}/{CNY}"
        return v
    put(td, r, 1, d["name"], size=9, bold=True)
    put(td, r, 2, d["own"], size=9, wrap=True)
    put(td, r, 3, d["per"], size=9, halign="center")
    put(td, r, 4, cv, size=9, halign="center", color=GREY)
    arr, rev, rec, fcf = d.get("arr"), d.get("rev"), d.get("rec"), d.get("fcf")
    put(td, r, 5, conv(arr), fmt=NUM, color=(GREEN if cv != "USD" else BLUE) if arr is not None else GREY)
    put(td, r, 6, f'=IF(COUNT(E{r})=1,E{r}-{TGT_ARR},"")', fmt=NUM)
    put(td, r, 7, conv(rev), fmt=NUM, color=(GREEN if cv != "USD" else BLUE) if rev is not None else GREY)
    put(td, r, 8, conv(rec), fmt=NUM, color=(GREEN if cv != "USD" else BLUE) if rec is not None else GREY)
    put(td, r, 9, f'=IF(COUNT(H{r},G{r})=2,H{r}/G{r},"")', fmt=PCT)
    put(td, r, 10, d.get("growth"), fmt=PCT, color=BLUE if d.get("growth") is not None else GREY)
    put(td, r, 11, d.get("subs"), fmt=CNT, color=BLUE if d.get("subs") is not None else GREY)
    put(td, r, 12, d.get("emp"), fmt=CNT, color=BLUE if d.get("emp") is not None else GREY)
    put(td, r, 13, f'=IF(COUNT(G{r},L{r})=2,G{r}*1000/L{r},"")', fmt=USD0)
    for col, key in [(14, "gm"), (15, "om"), (16, "nm"), (17, "eb")]:
        v = d.get(key)
        put(td, r, col, v, fmt=PCT, color=BLUE if v is not None else GREY)
    put(td, r, 18, conv(fcf), fmt=NUM, color=(GREEN if cv != "USD" else BLUE) if fcf is not None else GREY)
    put(td, r, 19, f'=IF(COUNT(R{r},G{r})=2,R{r}/G{r},"")', fmt=PCT)
    put(td, r, 20, d["conf"], size=9, bold=True, halign="center")
    put(td, r, 21, d["note"], size=8.5, color=GREY, wrap=True)
    td.row_dimensions[r].height = 58
    TD_ROW[d["name"]] = r


TD_ROW = {}


block_a = [
    dict(name="Microlise", own="AIM: SAAS (listed UK)", per="FY2025 (Dec-25)", cur="GBP",
         arr=59.2, rev=84.03, rec=58.8, growth=0.057, subs=None, emp=615,
         gm=54.14 / 84.03, om=-2.42 / 84.03, nm=-2.17 / 84.03, eb=0.10, fcf=9.47, conf="H",
         note="The anchor of this sheet and the most uncomfortable row in the workbook. ARR GBP59.2m "
              "(+4.6%), 64% gross margin, 1.4% churn, 417 new customers — and a STATUTORY OPERATING "
              "LOSS. The 10% figure is adjusted EBITDA, down from 14%. Enterprise value is GBP32m "
              "against GBP84m of revenue: 0.39x sales. Headcount also reported as ~730 across UK, "
              "France, Australia and India before announced cuts."),
    dict(name="ORBCOMM", own="Private (GI Partners, 2021 LBO)", per="TTM to Sep-25", cur="USD",
         arr=None, rev=250.0, rec=None, growth=None, subs=None, emp=878,
         gm=None, om=None, nm=None, eb=None, fcf=None, conf="L",
         note="Top of the band at ~USD250m revenue. Taken private in April 2021 at ~USD1.1bn "
              "including net debt. No margin, ARR or geographic disclosure since. Included for scale "
              "and the buyout multiple only."),
    dict(name="CalAmp", own="Private (Lynrock Lake, post-Ch11)", per="CY2024", cur="USD",
         arr=None, rev=197.0, rec=126.9, growth=None, subs=2700000, emp=644,
         gm=None, om=None, nm=None, eb=12.7 / 197.0, fcf=None, conf="M",
         note="The cautionary row. Reached ~USD200m revenue and 2.7m subscribers, then filed a "
              "pre-packaged Chapter 11 in June 2024 that wiped out USD230m of debt AND the public "
              "equity; delisted October 2024. EBITDA margin of 6.4% could not service the balance "
              "sheet. Headcount is as at Feb-2023, the last 10-K."),
    dict(name="Teletrac Navman", own="Private (Respida Capital, Jun-2026)", per="CY2026", cur="USD",
         arr=None, rev=168.0, rec=None, growth=None, subs=None, emp=None,
         gm=None, om=None, nm=None, eb=None, fcf=None, conf="L",
         note="The valuation row. Vontier sold the majority to Respida Capital, completing 30-Jun-2026 "
              "at a total transaction value of USD220m — roughly 1.3x revenue — with only ~USD80m of "
              "cash proceeds to Vontier. A strategic owner exiting a USD168m-revenue FMS business at "
              "about one and a third times sales."),
    dict(name="Netradyne", own="Private (VC; USD1.3bn valuation)", per="CY2024-25", cur="USD",
         arr=210.0, rev=None, rec=None, growth=0.62, subs=None, emp=1000,
         gm=None, om=None, nm=None, eb=None, fcf=None, conf="L",
         note="Above the band and growing fastest here: ARR USD210m in 2024 from USD129.8m in 2023, "
              "+62%. Valued at USD1.3bn on the 2025 Series D, ~6.2x ARR. Single product family (AI "
              "video safety). No margin disclosure. US-weighted with a large India engineering base."),
]
for d in block_a:
    td_row(r, d)
    r += 1

r += 1
section(td, r, "Block B — hardware / device vendors. NOT ARR comparables: these are suppliers to "
               "companies like McEasy, shown to expose where the device margin pool sits", 21)
r += 1
block_b = [
    dict(name="Queclink", own="SHE: 300590 (listed China)", per="FY2025 (Dec-25)", cur="CNY",
         arr=None, rev=891.1, rec=None, growth=-0.078, subs=None, emp=570,
         gm=329.17 / 891.1, om=48.51 / 891.1, nm=74.56 / 891.1, eb=None, fcf=45.79, conf="H",
         note="Squarely in the band at ~USD124m revenue, and squarely a device business: 37% gross "
              "margin, and operating margin collapsed from 13.4% to 5.4% in one year on a 7.8% "
              "revenue decline. 94% of revenue is earned outside China. Discloses a two-wheeler "
              "terminal line — directly relevant to Indonesian fleet mix."),
    dict(name="Streamax", own="SZ: 002970 (listed China)", per="FY2025 (Dec-25)", cur="CNY",
         arr=None, rev=2477.0, rec=None, growth=-0.108, subs=None, emp=2231,
         gm=1111.0 / 2477.0, om=375.0 / 2477.0, nm=383.0 / 2477.0, eb=None, fcf=294.0, conf="H",
         note="ABOVE the band at ~USD345m revenue — kept because it is the most Southeast-Asia-relevant "
              "listed player at any scale, with 65% overseas revenue and 5m+ vehicles. Note the shape: "
              "revenue FELL 10.8% while gross margin rose from 35.0% to 44.8% and net margin reached "
              "15.5%. Deliberately traded volume for margin."),
]
for d in block_b:
    td_row(r, d)
    r += 1

r += 1
section(td, r, "Block C — below the band, but the closest business models to McEasy", 21)
r += 1
block_c = [
    dict(name="Quartix", own="AIM: QTX (listed UK)", per="FY2025 (Dec-25)", cur="GBP",
         arr=37.0, rev=35.71, rec=35.71, growth=0.123, subs=310701, emp=173,
         gm=26.13 / 35.71, om=8.68 / 35.71, nm=6.38 / 35.71, eb=None, fcf=4.47, conf="H",
         note="The purest SMB-subscription telematics comparable that exists, and profitable at half "
              "the target size: 73% gross margin, 24% operating margin, 18% net margin, 12.5% FCF "
              "margin on ~USD47m of revenue. 310,701 subscriptions across 31,040 customers. Proof "
              "that this model works well below USD100m — and that scale is not what creates margin. "
              "Headcount 173 is a third-party estimate."),
    dict(name="Fleetio", own="Private (VC; USD1.5bn valuation)", per="CY2025", cur="USD",
         arr=None, rev=58.0, rec=None, growth=None, subs=8500, emp=463,
         gm=None, om=None, nm=None, eb=None, fcf=None, conf="L",
         note="The valuation counterpoint to Teletrac Navman. Raised USD450m+ Series D in March 2025 "
              "and acquired Auto Integrate; combined business valued above USD1.5bn on roughly USD58m "
              "of revenue — about 26x sales, versus Teletrac's 1.3x. 8,500+ fleets, 8m+ vehicles "
              "serviced. Revenue is a third-party estimate and the weakest number on this sheet. "
              "Subscribers column is fleets, not vehicles."),
]
for d in block_c:
    td_row(r, d)
    r += 1
TD_LAST = r - 1

# Streamax's percentage-only geography rows can now point at its revenue cell
for rr, co in zip(OTH_ROWS, OTH_CO):
    if co == "Streamax":
        put(rc, rr, 5, f"=IFERROR(H{rr}*'At 100M (Today)'!$G${TD_ROW['Streamax']},\"\")",
            fmt=NUM, color=GREEN)
        put(rc, rr, 6, f'=IFERROR(H{rr},"")', fmt=PCT)

r += 1
notes_block(td, r, [
    "Read this sheet against 'At 100M (Historic)'. Cartrack at USD111m ARR in FY2019 ran a 45% EBITDA "
    "margin and a 21% net margin. Microlise at USD78m ARR today runs a 10% ADJUSTED EBITDA margin and "
    "a statutory operating loss. Same revenue scale, same industry, opposite economics — the "
    "difference is the market and the pricing power, not the size.",
    "Recurring % of revenue is blank where recurring revenue is not disclosed. It is not zero. "
    "Quartix's recurring figure equals total revenue because it reports no separate hardware line, "
    "not because a split was assumed.",
    "Currency column matters. GBP figures are multiplied by the GBP/USD rate on 'Assumptions'; CNY "
    "figures are divided by the USD/CNY rate. Every USD cell here is a live formula over the native "
    "reported number, so changing an FX assumption reflows the whole sheet.",
    "Microlise margins are STATUTORY (gross 64.4%, operating -2.9%, net -2.6%) with the company's own "
    "adjusted EBITDA margin of 10% shown in the EBITDA column. Mixing statutory and adjusted in one "
    "row is deliberate: the gap between them is the point.",
    "Blocks B and C are labelled because they are not like-for-like. Block B sells devices to "
    "operators; Block C is below the size band. Neither should be averaged into Block A.",
], ncols=21)

# ========================================================= AT 100M ARR
ha = wb.create_sheet("At 100M (Historic)")
title(ha, "Comparables — at the moment they crossed ~USD100m ARR",
      "This is the real benchmark for a 2030 target of USD100m ARR. Note how different these "
      "columns look from the 'Current Position' sheet — particularly free-cash-flow margin.")
header_row(ha, 4, ["Company", "Fiscal period at ~USD100m ARR", "Years from founding",
                   "Recurring revenue (USD m)", "Total revenue (USD m)", "Revenue growth YoY",
                   "Subscribers", "ARPU (USD per subscriber per month)", "Employees",
                   "Gross margin", "Operating margin", "Net margin", "EBITDA margin",
                   "Capex % of revenue", "FCF margin", "Home market % of revenue",
                   "Confidence", "Notes"],
           widths=[22, 20, 12, 13, 13, 11, 13, 14, 11, 11, 11, 11, 11, 12, 11, 13, 11, 74])

r = 5
put(ha, r, 1, "Karooooo (Cartrack)", size=9, bold=True)
put(ha, r, 2, "FY2019 (Feb-19)", size=9, halign="center")
put(ha, r, 3, 15, fmt=CNT, color=BLUE, halign="center", note="Cartrack founded 2004.")
put(ha, r, 4, f"={K}{V19}{R['sub_u']}", fmt=NUM, color=GREEN)
put(ha, r, 5, f"={K}{V19}{R['rev_u']}", fmt=NUM, color=GREEN)
put(ha, r, 6, f'=IF(COUNT({K}{V19}{R["rev"]},{K}B{R["rev"]})=2,'
              f'{K}{V19}{R["rev"]}/{K}B{R["rev"]}-1,"")', fmt=PCT,
    note="Versus FY2018 total revenue of ZAR1,324.2m, per the FY2019 SENS release.")
put(ha, r, 7, f"={K}{V19}{R['subs']}", fmt=CNT, color=GREEN)
put(ha, r, 8, f"={K}{V19}{R['arpu_m']}", fmt=USD2, color=GREEN)
put(ha, r, 9, None, fmt=CNT, color=GREY)
put(ha, r, 10, None, fmt=PCT, color=GREY)
put(ha, r, 11, f"={K}{V19}{R['m_oprev']}", fmt=PCT, color=GREEN)
put(ha, r, 12, f"={K}{V19}{R['m_nprev']}", fmt=PCT, color=GREEN)
put(ha, r, 13, f"={K}{V19}{R['m_ebrev']}", fmt=PCT, color=GREEN)
put(ha, r, 14, f"={K}{V19}{R['m_cpxrev']}", fmt=PCT, color=GREEN)
put(ha, r, 15, f"={K}{V19}{R['m_fcfrev']}", fmt=PCT, color=GREEN)
put(ha, r, 16, "='Revenue by Country'!F5", fmt=PCT, color=GREEN)
put(ha, r, 17, "H", size=9, bold=True, halign="center")
put(ha, r, 18, "The benchmark row. 45% EBITDA margin and 21% net margin, but free cash flow was "
               "close to nil because capex ran at 30% of revenue — the hardware-subsidy model "
               "converts almost all EBITDA into installed devices. Gross margin was not disclosed "
               "in the FY2019 release.", size=8.5, color=GREY, wrap=True)
ha.row_dimensions[r].height = 52
r += 1

ha_rows = [
    dict(name="Ituran", per="~CY2007", yrs=12, rec=None, rev=114.0, growth=0.17, subs=407000,
         arpu=None, emp=800, gm=None, om=None, nm=None, eb=None, cpx=None, fcf=None,
         home=None, conf="L",
         note="Revenue annualised from Q2-2007 of USD29.2m. ~800 staff as at 2006, 407k "
              "subscribers in 2007. Pre-2008 disclosure is thin; margins not retrieved. "
              "Directional only."),
    dict(name="Powerfleet (via MiX Telematics)", per="FY2014 (Mar-14)", yrs=18, rec=81.0,
         rev=126.0, growth=0.20, subs=448000, arpu=None, emp=None, gm=None, om=None, nm=None,
         eb=0.215, cpx=None, fcf=None, home=None, conf="M",
         note="MiX Telematics (merged into Powerfleet in 2024) is the only lineage in Powerfleet "
              "that organically crossed USD100m. Figures are FY2014 guidance at the midpoint: "
              "revenue ZAR1,270-1,300m, subscription ZAR825-833m, adjusted EBITDA ZAR270-280m. "
              "EBITDA margin of ~21.5% at this scale versus Cartrack's 45% — the clearest "
              "illustration that USD100m ARR says nothing about profitability."),
    dict(name="Motive (as KeepTruckin)", per="~CY2019", yrs=6, rec=None, rev=None, growth=None,
         subs=None, arpu=None, emp=None, gm=None, om=None, nm=None, eb=None, cpx=None,
         fcf=None, home=0.90, conf="L",
         note="Crossed ~USD100m ARR around 2019 (55k+ customers as at Apr-2019; USD150m ARR by "
              "2021). Raised a USD149m Series D in 2019. No margin or headcount disclosure exists "
              "for that year — the company was private and deeply loss-making. The fastest to "
              "USD100m in this set by a wide margin, and still unprofitable seven years later."),
    dict(name="Geotab", per="~CY2013 (est.)", yrs=13, rec=None, rev=None, growth=None, subs=None,
         arpu=None, emp=None, gm=None, om=None, nm=None, eb=None, cpx=None, fcf=None, home=None,
         conf="L", note="No historical disclosure. Founded 2000; reached USD412m by 2021. Included "
                        "for completeness only."),
    dict(name="Lytx", per="~CY2013 (est.)", yrs=15, rec=None, rev=None, growth=None, subs=None,
         arpu=None, emp=None, gm=None, om=None, nm=None, eb=None, cpx=None, fcf=None, home=None,
         conf="L", note="No historical disclosure."),
    dict(name="Gurtam", per="Not reached", yrs=None, rec=None, rev=None, growth=None, subs=None,
         arpu=None, emp=None, gm=None, om=None, nm=None, eb=None, cpx=None, fcf=None, home=None,
         conf="H", note="At USD37.3m revenue in 2025, Gurtam has never approached USD100m. It is "
                        "the useful counter-example: a profitable, capital-light platform business "
                        "that chose not to chase scale."),
]
for d in ha_rows:
    put(ha, r, 1, d["name"], size=9, bold=True)
    put(ha, r, 2, d["per"], size=9, halign="center")
    put(ha, r, 3, d["yrs"], fmt=CNT, color=BLUE if d["yrs"] else GREY, halign="center")
    for col, key, fmt in [(4, "rec", NUM), (5, "rev", NUM), (6, "growth", PCT),
                          (7, "subs", CNT), (8, "arpu", USD2), (9, "emp", CNT),
                          (10, "gm", PCT), (11, "om", PCT), (12, "nm", PCT),
                          (13, "eb", PCT), (14, "cpx", PCT), (15, "fcf", PCT),
                          (16, "home", PCT)]:
        v = d[key]
        put(ha, r, col, v, fmt=fmt, color=BLUE if v is not None else GREY)
    put(ha, r, 17, d["conf"], size=9, bold=True, halign="center")
    put(ha, r, 18, d["note"], size=8.5, color=GREY, wrap=True)
    ha.row_dimensions[r].height = 52
    r += 1

r += 1
notes_block(ha, r, [
    "Blank cells are honest. For four of the seven companies the USD100m-ARR year predates any "
    "meaningful public disclosure. Cartrack FY2019 is the only fully-populated row, which is why "
    "it should be McEasy's primary north star and the rest are context.",
    "The pattern that matters: time-to-USD100m ranged from 6 years (Motive, venture-funded, "
    "unprofitable) to 18 years (MiX Telematics). EBITDA margin at that milestone ranged from ~21% "
    "(MiX) to 45% (Cartrack) to deeply negative (Motive). USD100m ARR is not one destination — it "
    "is at least three different businesses.",
    "Cartrack's ~2% FCF margin at USD100m ARR is the number to internalise. High EBITDA plus 30% "
    "capex-to-revenue equals no free cash. Any 2030 plan that assumes USD100m ARR funds itself "
    "needs to be tested against this row.",
], ncols=18)

# ===================================================== VALUATION SIGNALS
vs = wb.create_sheet("Valuation Signals")
title(vs, "What the market actually pays for an FMS business at this scale",
      "Among the three LISTED companies the spread is about 15x; across every data point here it is "
      "about 67x. The growth and margin columns are present so the spread is explained rather than "
      "asserted. Market data as at 10-Aug-2026.")
header_row(vs, 4, ["Company", "Type of data point", "Date", "Scale metric",
                   "Scale metric (USD m)", "EV or valuation (USD m)", "Multiple of scale metric",
                   "EV / EBITDA", "Growth at that point", "Profitability at that point",
                   "Confidence", "What happened / why the multiple"],
           widths=[22, 24, 13, 17, 13, 14, 13, 11, 12, 20, 10, 84])

VS_FIRST = 5
r = VS_FIRST
val_rows = [
    dict(name="Karooooo", typ="Listed enterprise value", date="10-Aug-2026", metric="Revenue",
         scale_lit=None, ev=1960.0,
         evebitda=21.62, growth=0.1997, prof="42% adj. EBITDA margin", conf="H",
         note="THE upper anchor for a profitable operator: 5.5x sales and 21.6x EBITDA on USD344m of "
              "revenue growing 20% with a 42% EBITDA margin. This is what the market pays when growth "
              "AND margin are both present."),
    dict(name="Fleetio", typ="Private round (post-money)", date="Mar-2025", metric="Revenue",
         scale=None, scale_lit=58.0, ev=1500.0, evebitda=None, growth=None,
         prof="n/d; loss-making", conf="L",
         note="~26x sales, the highest multiple here, on the weakest revenue figure here. Priced on "
              "growth and the Auto Integrate network effect, not on profit. Revenue is a third-party "
              "estimate, so treat the multiple as indicative of appetite rather than a precise number."),
    dict(name="Netradyne", typ="Private round (post-money)", date="2025 Series D", metric="ARR",
         scale=None, scale_lit=210.0, ev=1300.0, evebitda=None, growth=0.62,
         prof="n/d", conf="L",
         note="~6.2x ARR on 62% ARR growth. The cleanest read on what fast growth is worth at this "
              "scale without disclosed profitability."),
    dict(name="ORBCOMM", typ="Take-private (LBO)", date="Apr-2021", metric="Revenue",
         scale=None, scale_lit=248.5, ev=1100.0, evebitda=None, growth=None,
         prof="n/d", conf="M",
         note="~4.4x sales including net debt, on FY2020 revenue of USD248.5m. A 2021 deal, struck in "
              "a materially cheaper rate environment than today's."),
    dict(name="Quartix", typ="Listed enterprise value", date="10-Aug-2026", metric="Revenue",
         scale=None, scale_lit=None, ev=None, evebitda=8.05, growth=0.123,
         prof="24% operating margin", conf="H",
         note="2.8x sales and 8.1x EBITDA. Profitable, high gross margin, but only 12% growth and "
              "sub-scale — so the market pays half of Karooooo's sales multiple and a third of its "
              "EBITDA multiple."),
    dict(name="Teletrac Navman", typ="Trade sale (strategic exit)", date="Jun-2026", metric="Revenue",
         scale=None, scale_lit=168.0, ev=220.0, evebitda=None, growth=None,
         prof="n/d", conf="M",
         note="~1.3x sales. Vontier, a strategic owner, chose to exit a USD168m-revenue FMS business "
              "for a USD220m total transaction value and only ~USD80m of cash. Scale bought no premium."),
    dict(name="Microlise", typ="Listed enterprise value", date="10-Aug-2026", metric="Revenue",
         scale=None, scale_lit=None, ev=None, evebitda=4.29, growth=0.057,
         prof="10% adj. EBITDA; statutory operating loss", conf="H",
         note="THE lower anchor and the most important row in the workbook. 0.39x sales. A GBP84m "
              "revenue business with a 64% gross margin carries an enterprise value of GBP32m — less "
              "than five months of revenue. Low growth plus a statutory loss equals near-zero "
              "enterprise value, whatever the ARR says."),
    dict(name="CalAmp", typ="Chapter 11 restructuring", date="Jun-2024", metric="Revenue",
         scale=None, scale_lit=197.0, ev=0.0, evebitda=None, growth=None,
         prof="6.4% EBITDA margin", conf="H",
         note="Equity value to zero. ~USD200m of revenue and 2.7m subscribers were not enough: a 6.4% "
              "EBITDA margin could not service the debt, USD230m was written off, lenders took the "
              "company and it delisted. The floor case for reaching this scale unprofitably."),
]
for d in val_rows:
    put(vs, r, 1, d["name"], size=9, bold=True)
    put(vs, r, 2, d["typ"], size=9, wrap=True)
    put(vs, r, 3, d["date"], size=9, halign="center")
    put(vs, r, 4, d["metric"], size=9, halign="center")
    # scale metric: link to the comparables sheets where possible, else a documented literal
    if d["name"] == "Karooooo":
        put(vs, r, 5, f"='Current Position'!$E${KARO_ROW}", fmt=NUM, color=GREEN)
    elif d["name"] == "Microlise":
        put(vs, r, 5, f"='At 100M (Today)'!$G${TD_ROW['Microlise']}", fmt=NUM, color=GREEN)
    elif d["name"] == "Quartix":
        put(vs, r, 5, f"='At 100M (Today)'!$G${TD_ROW['Quartix']}", fmt=NUM, color=GREEN)
    else:
        put(vs, r, 5, d["scale_lit"], fmt=NUM, color=BLUE)
    # EV: Microlise and Quartix are GBP market data, converted by formula
    if d["name"] == "Microlise":
        put(vs, r, 6, f"=32.35*{GBP}", fmt=NUM, color=GREEN,
            note="Enterprise value GBP32.35m (market cap GBP45.22m less net cash GBP12.87m), "
                 "10-Aug-2026.")
    elif d["name"] == "Quartix":
        put(vs, r, 6, f"=104.74*{GBP}", fmt=NUM, color=GREEN,
            note="Enterprise value GBP104.74m (market cap GBP109.02m less net cash GBP4.28m), "
                 "10-Aug-2026.")
    else:
        put(vs, r, 6, d["ev"], fmt=NUM, color=BLUE if d["ev"] is not None else GREY)
    put(vs, r, 7, f'=IF(AND(COUNT(E{r},F{r})=2,E{r}<>0),F{r}/E{r},"")', fmt="0.00x", bold=True,
        note="Computed here as EV divided by the scale metric on the comparables sheets. For the "
             "listed names this lands within ~0.15x of the ratio the data provider publishes "
             "(Karooooo 5.52x, Quartix 2.78x, Microlise 0.39x) because they use trailing-twelve-month "
             "revenue at their own FX, whereas this workbook uses the reported fiscal-year figure at "
             "the documented average rate. The gap does not change any conclusion."
        if d["name"] in ("Karooooo", "Quartix", "Microlise") else None)
    put(vs, r, 8, d["evebitda"], fmt="0.0x", color=BLUE if d["evebitda"] is not None else GREY)
    put(vs, r, 9, d["growth"], fmt=PCT, color=BLUE if d["growth"] is not None else GREY)
    put(vs, r, 10, d["prof"], size=8.5, color=GREY, wrap=True)
    put(vs, r, 11, d["conf"], size=9, bold=True, halign="center")
    put(vs, r, 12, d["note"], size=8.5, color=GREY, wrap=True)
    vs.row_dimensions[r].height = 52
    r += 1
VS_LAST = r - 1

r += 1
section(vs, r, "Dispersion", 12); r += 1
VS_MAX, VS_MIN, VS_RATIO = r, r + 1, r + 2
VS_LMAX, VS_LMIN, VS_LRATIO = r + 3, r + 4, r + 5
# the three listed rows: Karooooo, Quartix, Microlise
LG = f"G{VS_FIRST},G{VS_FIRST + 4},G{VS_FIRST + 6}"
disp = [
    (VS_MAX, "Highest multiple — all data points", f"=MAX(G{VS_FIRST}:G{VS_LAST})", "0.00x",
     "Fleetio, at about 26x. Priced on growth, and computed on the least reliable revenue figure in "
     "the table — treat it as evidence of appetite, not a precise multiple."),
    (VS_MIN, "Lowest multiple — all data points, excluding the zero",
     f'=_xlfn.MINIFS(G{VS_FIRST}:G{VS_LAST},G{VS_FIRST}:G{VS_LAST},">0")', "0.00x",
     "Microlise, at about 0.39x. CalAmp's zero is excluded so it does not swallow the range — but "
     "its equity really did go to nil, which is the true floor."),
    (VS_RATIO, "Ratio between them", f'=IFERROR(B{VS_MAX}/B{VS_MIN},"")', "0.0x",
     "Mixes a venture round with a listed enterprise value, so read it as an outer bound."),
    (VS_LMAX, "Highest multiple — listed companies only", f"=MAX({LG})", "0.00x",
     "Karooooo. The cleanest comparison: three public companies, same day, same methodology."),
    (VS_LMIN, "Lowest multiple — listed companies only", f"=MIN({LG})", "0.00x", "Microlise."),
    (VS_LRATIO, "Ratio between them (listed only)", f'=IFERROR(B{VS_LMAX}/B{VS_LMIN},"")', "0.0x",
     "THE number to remember. Same industry, same day, same disclosure regime — and roughly a 15x "
     "difference in what one dollar of revenue is worth, driven by growth and margin alone."),
]
for row, lab, formula, fmt, note in disp:
    put(vs, row, 1, lab, size=10, bold=(row == VS_LRATIO))
    put(vs, row, 2, formula, fmt=fmt, bold=True, halign="center")
    put(vs, row, 3, note, size=8.5, color=GREY, wrap=True)
    vs.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
    vs.row_dimensions[row].height = 26
r = VS_LRATIO + 1

r += 1
notes_block(vs, r, [
    "Multiples are not like-for-like and are not meant to be. A listed enterprise value, a "
    "venture round post-money and a completed trade sale are three different things — the 'Type of "
    "data point' column says which is which, and they should not be averaged.",
    "The pattern that survives the caveats: at USD100-250m of revenue, an FMS business is worth "
    "roughly 0.4x sales if it is growing slowly at a thin margin, and roughly 5-6x sales if it is "
    "growing 20%+ at a fat margin. The revenue number itself explains almost none of the variance.",
    "For McEasy this reframes the 2030 target. USD100m ARR at Microlise-like economics would be worth "
    "roughly USD40m of enterprise value. The same USD100m ARR at Cartrack-like economics would be "
    "worth several hundred million. The goal should be specified as ARR AND margin, never ARR alone.",
], ncols=12)

# ===================================================== McEASY BENCHMARK
mb = wb.create_sheet("McEasy Benchmark")
title(mb, "McEasy — what USD100m ARR by 2030 actually requires",
      "Yellow cells are for you to fill in. Everything else is calculated from the comparables.")

put(mb, 4, 1, "Your inputs", bold=True, size=11, color="1F3864")
inp = [("McEasy ARR today (USD m)", 15.0, NUM, "Replace with your actual figure. 15.0 is a "
        "placeholder so the formulas below show a worked example."),
       ("Base year", 2026, CNT, "Calendar year the ARR above refers to."),
       ("Target ARR (USD m)", f"={TGT_ARR}", NUM, "From the Assumptions sheet."),
       ("Target year", f"={TGT_YR}", CNT, "From the Assumptions sheet.")]
r = 5
for lab, val, fmt, note in inp:
    put(mb, r, 1, lab, size=10)
    is_formula = isinstance(val, str)
    put(mb, r, 2, val, fmt=fmt, color=GREEN if is_formula else BLUE, bold=True,
        halign="center", fill=None if is_formula else YEL)
    put(mb, r, 3, note, size=8.5, color=GREY, wrap=True)
    mb.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    mb.row_dimensions[r].height = 26
    r += 1
ARR_NOW, BASE_YR, TGT, TYR = 5, 6, 7, 8
mb.column_dimensions["A"].width = 52
mb.column_dimensions["B"].width = 16
for c in "CDEFG":
    mb.column_dimensions[c].width = 17

r += 1
section(mb, r, "What that implies", 7); r += 1
put(mb, r, 1, "Years to target", size=10)
put(mb, r, 2, f"=B{TYR}-B{BASE_YR}", fmt=CNT, bold=True, halign="center")
YRS = r; r += 1
put(mb, r, 1, "ARR multiple required", size=10)
put(mb, r, 2, f'=IFERROR(B{TGT}/B{ARR_NOW},"")', fmt="0.0x", bold=True, halign="center")
r += 1
put(mb, r, 1, "Required compound annual ARR growth", size=10)
put(mb, r, 2, f'=IFERROR((B{TGT}/B{ARR_NOW})^(1/B{YRS})-1,"")', fmt=PCT, bold=True,
    halign="center")
CAGR = r; r += 1
put(mb, r, 1, "Cartrack's actual ARR CAGR, FY2019 to FY2026 (USD)", size=10)
put(mb, r, 2, f'=({K}{LAST}{R["arr_u"]}/{K}{V19}{R["sub_u"]})^(1/7)-1', fmt=PCT,
    color=GREEN, bold=True, halign="center")
put(mb, r, 3, "Cartrack went from ~USD111m to USD325m over seven years. If your required CAGR "
              "above is far higher than this, you are not planning a Cartrack trajectory — you are "
              "planning a Motive one, which needs venture capital and tolerates losses.",
    size=8.5, color=GREY, wrap=True)
mb.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
mb.row_dimensions[r].height = 40
r += 2

section(mb, r, "Benchmark: what Cartrack looked like at ~USD100m ARR (FY2019)", 7); r += 1
for lab, ref, fmt, comment in [
    ("Subscription revenue (USD m)", f"={K}{V19}{R['sub_u']}", NUM, ""),
    ("Subscribers", f"={K}{V19}{R['subs']}", CNT, ""),
    ("ARPU (USD per subscriber per month)", f"={K}{V19}{R['arpu_m']}", USD2,
     "The number McEasy should stress-test hardest. Indonesian ARPU is structurally below this."),
    ("EBITDA margin", f"={K}{V19}{R['m_ebrev']}", PCT, ""),
    ("Operating margin", f"={K}{V19}{R['m_oprev']}", PCT, ""),
    ("Net margin", f"={K}{V19}{R['m_nprev']}", PCT, ""),
    ("Capex as % of revenue", f"={K}{V19}{R['m_cpxrev']}", PCT,
     "Devices, installation and contract assets. This is what eats the EBITDA."),
    ("Free-cash-flow margin", f"={K}{V19}{R['m_fcfrev']}", PCT,
     "Near zero. USD100m ARR did not make Cartrack self-funding."),
    ("Home market share of revenue", "='Revenue by Country'!F5", PCT,
     "Cartrack reached USD100m ARR on the back of one market, not many."),
]:
    put(mb, r, 1, lab, size=10)
    put(mb, r, 2, ref, fmt=fmt, color=GREEN, bold=True, halign="center")
    if comment:
        put(mb, r, 3, comment, size=8.5, color=GREY, wrap=True)
        mb.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        mb.row_dimensions[r].height = 30
    r += 1

r += 1
section(mb, r, "Reality check — the same metrics for companies actually at this scale today", 7)
r += 1
MB_RC_HDR = r
for i, lab in enumerate(["Metric", "Cartrack FY2019 (at USD111m ARR)",
                         "Microlise FY2025 (USD78m ARR)", "Quartix FY2025 (USD49m ARR)",
                         "Karooooo FY2026 (USD325m ARR)"], start=1):
    put(mb, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
mb.row_dimensions[r].height = 40
r += 1
MIC, QTX = TD_ROW["Microlise"], TD_ROW["Quartix"]
T = "'At 100M (Today)'!"
rc_rows = [
    ("ARR or subscription revenue (USD m)", f"={K}{V19}{R['sub_u']}", f"={T}E{MIC}",
     f"={T}E{QTX}", f"={K}{LAST}{R['arr_u']}", NUM),
    ("Total revenue (USD m)", f"={K}{V19}{R['rev_u']}", f"={T}G{MIC}", f"={T}G{QTX}",
     f"={K}{LAST}{R['rev_u']}", NUM),
    ("Revenue growth YoY", f"={K}{V19}{R['rev']}/{K}B{R['rev']}-1", f"={T}J{MIC}",
     f"={T}J{QTX}", f"='Current Position'!H{KARO_ROW}", PCT),
    ("Gross margin", None, f"={T}N{MIC}", f"={T}N{QTX}", f"={K}{LAST}{R['m_gprev']}", PCT),
    ("Operating margin (statutory)", f"={K}{V19}{R['m_oprev']}", f"={T}O{MIC}", f"={T}O{QTX}",
     f"={K}{LAST}{R['m_oprev']}", PCT),
    ("Net margin (statutory)", f"={K}{V19}{R['m_nprev']}", f"={T}P{MIC}", f"={T}P{QTX}",
     f"={K}{LAST}{R['m_nprev']}", PCT),
    ("EBITDA margin", f"={K}{V19}{R['m_ebrev']}", f"={T}Q{MIC}", None,
     f"={K}{LAST}{R['m_ebrev']}", PCT),
    ("Free-cash-flow margin", f"={K}{V19}{R['m_fcfrev']}", f"={T}S{MIC}", f"={T}S{QTX}",
     f"={K}{LAST}{R['m_fcfrev']}", PCT),
    ("ARPU (USD per subscriber per month)", f"={K}{V19}{R['arpu_m']}", None,
     f'=IFERROR({T}E{QTX}*1000000/{T}K{QTX}/12,"")', f"={K}{LAST}{R['arpu_m']}", USD2),
    ("Employees", None, f"={T}L{MIC}", f"={T}L{QTX}", f"='Current Position'!J{KARO_ROW}", CNT),
    ("Revenue per employee (USD k)", None, f"={T}M{MIC}", f"={T}M{QTX}",
     f"='Current Position'!K{KARO_ROW}", USD0),
    ("Enterprise value / sales", None, f"='Valuation Signals'!G{VS_FIRST + 6}",
     f"='Valuation Signals'!G{VS_FIRST + 4}", f"='Valuation Signals'!G{VS_FIRST}", "0.00x"),
]
for lab, a, b, c, d, fmt in rc_rows:
    put(mb, r, 1, lab, size=9)
    for col, v in ((2, a), (3, b), (4, c), (5, d)):
        put(mb, r, col, v, fmt=fmt, color=GREEN if v else GREY, halign="center")
    r += 1
r += 1
put(mb, r, 1, "Read across each row. USD100m ARR is not a single economic outcome — Cartrack and "
              "Microlise sit at comparable ARR with opposite margins and a ~14x gap in enterprise "
              "value per dollar of sales. Quartix is profitable at half the size. Specify the 2030 "
              "goal as ARR AND margin.", size=9, wrap=True, color="C00000")
mb.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
mb.row_dimensions[r].height = 44
r += 2

section(mb, r, "Subscribers required for USD100m ARR, by monthly ARPU", 7); r += 1
for i, lab in enumerate(["ARPU (USD per month)", "Implied annual ARPU (USD)",
                         "Subscribers required", "Multiple of Cartrack's FY2019 base",
                         "Comment"], start=1):
    put(mb, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
mb.row_dimensions[r].height = 32
r += 1
sens_first = r
for arpu, comment in [(5.0, "Low-end Indonesian SME price point."),
                      (8.0, "Typical Indonesian mid-market FMS subscription."),
                      (9.63, "Cartrack's actual FY2019 ARPU — the benchmark row above."),
                      (12.0, "Requires clear value-add well beyond basic tracking."),
                      (15.0, "Multi-product: telematics + video + fuel + driver app."),
                      (20.0, "Enterprise-only mix; a much smaller addressable base.")]:
    put(mb, r, 1, arpu, fmt=USD2, color=BLUE, halign="center", fill=YEL)
    put(mb, r, 2, f"=A{r}*12", fmt=USD0, halign="center")
    put(mb, r, 3, f'=IFERROR($B${TGT}*1000000/B{r},"")', fmt=CNT, bold=True, halign="center")
    put(mb, r, 4, f'=IFERROR(C{r}/{K}{V19}{R["subs"]},"")', fmt="0.00x", halign="center")
    put(mb, r, 5, comment, size=8.5, color=GREY, wrap=True)
    mb.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
    mb.row_dimensions[r].height = 24
    r += 1

r += 1
section(mb, r, "Headcount implied by peer revenue-per-employee", 7); r += 1
for i, lab in enumerate(["Benchmark", "Revenue per employee (USD k)",
                         "Headcount implied at USD100m ARR", "Comment"], start=1):
    put(mb, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
for lab, ref, comment in [
    ("Karooooo today (FY2026)", f"='Current Position'!K{KARO_ROW}",
     "Emerging-market cost base with in-house installation crews. The most relevant analogue."),
    ("Motive (LTM Sep-25)", f"='Current Position'!K{MOT_ROW}",
     "78% of Motive's staff are offshore, which is why this looks efficient despite the losses."),
    ("Gurtam (CY2025)", f"='Current Position'!K{CP_LAST}",
     "Pure software / channel model — no installation labour at all. The upper bound."),
]:
    put(mb, r, 1, lab, size=10)
    put(mb, r, 2, ref, fmt=USD0, color=GREEN, halign="center")
    put(mb, r, 3, f'=IFERROR($B${TGT}*1000/B{r},"")', fmt=CNT, bold=True, halign="center")
    put(mb, r, 4, comment, size=8.5, color=GREY, wrap=True)
    mb.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    mb.row_dimensions[r].height = 30
    r += 1

r += 1
notes_block(mb, r, [
    "This sheet is a benchmarking tool, not a financial model. It answers \"what would McEasy have "
    "to look like\", not \"how does McEasy get there\". Building the second thing is the next piece "
    "of work.",
    "The ARPU table is the crux. At USD8 per subscriber per month McEasy needs a little over one "
    "million subscribers to reach USD100m ARR — more than Cartrack had in its entire global base "
    "in FY2019, in a single country. Either the ARPU assumption rises materially, or the "
    "subscriber target becomes the binding constraint on the 2030 goal.",
], ncols=7)

# ========================================================== McEASY MODEL
# Read McEasy's own model directly so the figures cannot drift from the source.
MODEL_PATH = (r"C:\Users\Grady Kusmulyadi\OneDrive - PT. Otto Menara Globalindo"
              r"\Documents\H2 2026\Strategy 2027\McEasy model for claude reading.xlsx")
try:
    src = openpyxl.load_workbook(MODEL_PATH, data_only=True)
except Exception as exc:                                     # noqa: BLE001
    raise SystemExit(f"Cannot read McEasy model at {MODEL_PATH}: {exc}")
spl, sbs = src["PnL_Consol (USD)"], src["BS_Consol (USD)"]


def find_row(sheet, label, lo=1, hi=120):
    for rr in range(lo, hi + 1):
        v = sheet.cell(row=rr, column=1).value
        if isinstance(v, str) and v.strip() == label:
            return rr
    raise SystemExit(f"Label not found in {sheet.title!r}: {label!r}")


MYRS = list(range(2022, 2031))
PL_COL = {y: 2 + i for i, y in enumerate(MYRS)}          # P&L: B..J = 2022..2030
BS_COL = {y: 2 + i for i, y in enumerate(range(2023, 2031))}  # BS: B..I = 2023..2030


def pl(label):
    rr = find_row(spl, label)
    return {y: spl.cell(row=rr, column=PL_COL[y]).value for y in MYRS}


def bs(label):
    rr = find_row(sbs, label)
    return {y: sbs.cell(row=rr, column=BS_COL[y]).value for y in BS_COL}


M = {
    "arr": pl("ARR"), "rev_saas": pl("SaaS Solutions"), "rev_tot": pl("Total Revenue"),
    "cogs": pl("Total Cost of Revenue"), "gp": pl("Total Gross Profit"),
    "logi": pl("Logistic Cost"), "sm": pl("Sales & Marketing Costs"),
    "cm": pl("Contribution Margin"), "pers": pl("Personnel Expenses"),
    "ga": pl("General & Administrative Expenses"), "mkt": pl("Marketing Expenses"),
    "tech": pl("Technology Expenses"), "othx": pl("Other Expenses (Income)"),
    "opex": pl("Total Operational Expenses"), "ebitda": pl("EBITDA"),
    "tax": pl("Tax & Interest"), "da": pl("Depreciation"), "np": pl("Net Profit (Loss)"),
}
# "SaaS Solutions" appears under Revenue, COGS and Gross Profit; find_row takes the first
# (revenue). Sparepart revenue is the residual, which also proves the lines reconcile.
M["gfa"] = bs("Fixed Assets")
M["cash"] = bs("Cash and Cash Equivalent")
M["puc"] = bs("Paid Up Capital")

mm = wb.create_sheet("McEasy Model")
title(mm, "McEasy's own model, as supplied — read directly from the source file",
      "Blue = taken verbatim from 'McEasy model for claude reading.xlsx'. Yellow = the volume "
      "assumption Grady supplied in conversation, which is NOT in the model. Nothing on the "
      "benchmark sheets hardcodes a McEasy number; they all point here.")
header_row(mm, 4, ["Line item (USD)"] + [str(y) for y in MYRS], widths=[46] + [13] * 9)
MCOL = {y: 2 + i for i, y in enumerate(MYRS)}
M30 = get_column_letter(MCOL[2030])
M25 = get_column_letter(MCOL[2025])

put(mm, 5, 1, "Status per the model", bold=True, size=9)
for y in MYRS:
    put(mm, 5, MCOL[y], "Actual" if y <= 2025 else "Projection", size=9, color=GREY,
        halign="center")

MR = {}
r = 6
section(mm, r, "Profit and loss, as modelled", 10); r += 1
pl_rows = [
    ("arr", "ARR"), ("rev_saas", "Revenue — SaaS Solutions"),
    ("_spare", "Revenue — Sparepart Solutions"), ("rev_tot", "Total Revenue"),
    ("cogs", "Total Cost of Revenue"), ("gp", "Total Gross Profit"),
    ("logi", "Logistic Cost"), ("sm", "Sales & Marketing Costs"),
    ("cm", "Contribution Margin"), ("pers", "Personnel Expenses"),
    ("ga", "General & Administrative Expenses"), ("mkt", "Marketing Expenses"),
    ("tech", "Technology Expenses"), ("othx", "Other Expenses (Income)"),
    ("opex", "Total Operational Expenses"), ("ebitda", "EBITDA"),
    ("da", "Depreciation"), ("tax", "Tax & Interest"), ("np", "Net Profit (Loss)"),
]
for key, lab in pl_rows:
    put(mm, r, 1, lab, size=9, bold=key in ("arr", "rev_tot", "ebitda", "np"))
    for y in MYRS:
        if key == "_spare":     # residual, so the revenue lines are forced to reconcile
            cl = get_column_letter(MCOL[y])
            put(mm, r, MCOL[y], f"={cl}{MR['rev_tot'] if 'rev_tot' in MR else r + 1}"
                                f"-{cl}{MR['rev_saas']}", fmt=NUM0)
        else:
            put(mm, r, MCOL[y], M[key][y], fmt=NUM0, color=BLUE,
                bold=key in ("arr", "rev_tot", "ebitda", "np"))
    MR[key] = r
    r += 1
# the residual formula needed Total Revenue's row, which is written after it — repair in place
for y in MYRS:
    cl = get_column_letter(MCOL[y])
    put(mm, MR["_spare"], MCOL[y],
        f"={cl}{MR['rev_tot']}-{cl}{MR['rev_saas']}", fmt=NUM0)

r += 1
section(mm, r, "From the balance sheet", 10); r += 1
for key, lab in [("gfa", "Fixed assets, gross"), ("cash", "Cash and cash equivalents"),
                 ("puc", "Paid-up capital")]:
    put(mm, r, 1, lab, size=9)
    for y in MYRS:
        put(mm, r, MCOL[y], M[key].get(y), fmt=NUM0,
            color=BLUE if M[key].get(y) is not None else GREY)
    MR[key] = r
    r += 1
put(mm, r, 1, "Capex (year-on-year change in gross fixed assets)", size=9, bold=True)
for i, y in enumerate(MYRS):
    cl = get_column_letter(MCOL[y])
    if i == 0 or y == 2023:
        put(mm, r, MCOL[y], None, fmt=NUM0, color=GREY)
    else:
        pc = get_column_letter(MCOL[MYRS[i - 1]])
        put(mm, r, MCOL[y], f'=IF(COUNT({cl}{MR["gfa"]},{pc}{MR["gfa"]})=2,'
                            f'{cl}{MR["gfa"]}-{pc}{MR["gfa"]},"")', fmt=NUM0, bold=True)
MR["capex"] = r; r += 1

r += 1
section(mm, r, "Margins and ratios (calculated)", 10); r += 1
for lab, num, key in [("Gross margin", "gp", "m_gp"),
                      ("Contribution margin", "cm", "m_cm"),
                      ("Total operational expenses % of revenue", "opex", "m_opex"),
                      ("EBITDA margin", "ebitda", "m_eb"),
                      ("Depreciation % of revenue", "da", "m_da"),
                      ("Capex % of revenue", "capex", "m_cpx"),
                      ("Net margin", "np", "m_np")]:
    put(mm, r, 1, lab, size=9, bold=key in ("m_eb", "m_np"))
    for y in MYRS:
        cl = get_column_letter(MCOL[y])
        put(mm, r, MCOL[y], f'=IF(COUNT({cl}{MR[num]},{cl}{MR["rev_tot"]})=2,'
                            f'{cl}{MR[num]}/{cl}{MR["rev_tot"]},"")', fmt=PCT,
            bold=key in ("m_eb", "m_np"))
    MR[key] = r
    r += 1
put(mm, r, 1, "ARR growth year on year", size=9)
for i, y in enumerate(MYRS):
    cl = get_column_letter(MCOL[y])
    if i == 0:
        put(mm, r, MCOL[y], None, fmt=PCT, color=GREY)
    else:
        pc = get_column_letter(MCOL[MYRS[i - 1]])
        put(mm, r, MCOL[y], f'=IFERROR({cl}{MR["arr"]}/{pc}{MR["arr"]}-1,"")', fmt=PCT)
MR["arr_g"] = r; r += 1

r += 1
section(mm, r, "Volume assumption — SUPPLIED BY GRADY, NOT PRESENT IN THE MODEL", 10); r += 1
VOL_NOTE = ("Grady stated ~900k-1m vehicles at USD8-9 per vehicle per month for 2030. The two "
            "sheets in the model contain no subscriber, vehicle or ARPU driver at all. Every "
            "per-vehicle figure in this workbook flows from this single yellow cell — change it and "
            "the whole benchmark reflows. Turning this into a modelled year-by-year build is the "
            "highest-value fix to the model itself.")
put(mm, r, 1, "Vehicles at 2030 — central case", size=10, bold=True)
put(mm, r, 2, 950000, fmt=CNT, color=BLUE, bold=True, halign="center", fill=YEL, note=VOL_NOTE)
put(mm, r, 3, VOL_NOTE, size=8.5, color=GREY, wrap=True)
mm.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
mm.row_dimensions[r].height = 46
VEH = f"'McEasy Model'!$B${r}"
MR["veh"] = r; r += 1
for lab, val in [("Vehicles at 2030 — low case", 900000), ("Vehicles at 2030 — high case", 1000000)]:
    put(mm, r, 1, lab, size=9)
    put(mm, r, 2, val, fmt=CNT, color=BLUE, halign="center", fill=YEL)
    MR["veh_lo" if "low" in lab else "veh_hi"] = r
    r += 1
for lab, num, key in [("Implied total revenue per vehicle per month", "rev_tot", "arpu_tot"),
                      ("Implied SaaS revenue per vehicle per month", "rev_saas", "arpu_saas")]:
    put(mm, r, 1, lab, size=9, bold=True)
    put(mm, r, 2, f'=IFERROR({M30}{MR[num]}/{VEH}/12,"")', fmt=USD2, bold=True, halign="center")
    put(mm, r, 3, "Must land inside the USD8-9 band Grady stated for the total line.",
        size=8.5, color=GREY, wrap=True)
    mm.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    MR[key] = r
    r += 1
# Reconcile the two statements against each other — an investor will do this arithmetic.
put(mm, r, 1, "Vehicles implied by USD9.00/month (top of the stated band)", size=9)
put(mm, r, 2, f'=IFERROR({M30}{MR["rev_tot"]}/9/12,"")', fmt=CNT, halign="center")
put(mm, r, 3, "Reconciliation check: the USD8-9 band and the 900k-1m range are not quite consistent "
              "at the low end. 900,000 vehicles implies USD9.43/month, just above the quoted band. "
              "The internally consistent combination is roughly 943k-1,061k vehicles at USD8.00-9.00, "
              "i.e. the upper half of the vehicle range. Worth tightening before the meeting — it is "
              "exactly the arithmetic an investor will run.", size=8.5, color=GREY, wrap=True)
mm.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
mm.row_dimensions[r].height = 46
r += 1
put(mm, r, 1, "Vehicles implied by USD8.00/month (bottom of the stated band)", size=9)
put(mm, r, 2, f'=IFERROR({M30}{MR["rev_tot"]}/8/12,"")', fmt=CNT, halign="center")
r += 1

r += 1
section(mm, r, "The growth the plan requires", 10); r += 1
put(mm, r, 1, "Required ARR CAGR, 2025 to 2030", size=10, bold=True)
put(mm, r, 2, f'=IFERROR(({M30}{MR["arr"]}/{M25}{MR["arr"]})^(1/5)-1,"")', fmt=PCT, bold=True,
    halign="center")
MR["cagr"] = r; r += 1
put(mm, r, 1, "ARR multiple, 2025 to 2030", size=10)
put(mm, r, 2, f'=IFERROR({M30}{MR["arr"]}/{M25}{MR["arr"]},"")', fmt="0.0x", halign="center")
r += 2
notes_block(mm, r, [
    "Every blue figure on this sheet was read programmatically from Grady's model file at build "
    "time, not transcribed. Sparepart revenue is shown as Total Revenue less SaaS revenue, so the "
    "revenue lines are forced to reconcile.",
    "The model is internally consistent: the balance sheet balances in every year, and the change in "
    "total retained earnings equals P&L net profit in every year. The 'Previous / Current Earnings' "
    "split inside the equity block is a period-labelling quirk, not an articulation error.",
    "Capex is derived, not disclosed: it is the year-on-year change in GROSS fixed assets, which "
    "ignores disposals. If McEasy retires and replaces devices at scale, true gross capex is higher "
    "than shown and the device-cost argument on the benchmark sheet gets harder, not easier.",
], ncols=10)

# ==================================================== McEASY vs BENCHMARK
mvb = wb.create_sheet("McEasy vs Benchmark")
title(mvb, "McEasy 2030 against Cartrack FY2019 — the investor argument",
      "Cartrack crossed USD111m subscription ARR in FY2019 with 960,798 subscribers. McEasy targets "
      "USD100.9m ARR in 2030 with ~950k vehicles. Read the per-vehicle block, not the margin block.")
mvb.column_dimensions["A"].width = 46
for c, w in zip("BCDE", (15, 15, 12, 74)):
    mvb.column_dimensions[c].width = w

MM = "'McEasy Model'!"
CT_SUBS = f"{K}{V19}{R['subs']}"
CT_FX = f"{K}{V19}6"
r = 4
section(mvb, r, "Block 1 — margin comparison. This is the block that invites the attack.", 5)
r += 1
for i, lab in enumerate(["Metric", "McEasy 2030", "Cartrack FY2019", "Delta", "Comment"], start=1):
    put(mvb, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
m1 = [
    ("ARR (USD m)", f"={MM}{M30}{MR['arr']}/1000000", f"={K}{V19}{R['sub_u']}", NUM,
     "Within 9%. Cartrack's is subscription revenue, the closest disclosed analogue to ARR."),
    ("Total revenue (USD m)", f"={MM}{M30}{MR['rev_tot']}/1000000", f"={K}{V19}{R['rev_u']}", NUM,
     "McEasy earns less total revenue at the same subscriber count — see the ARPU row below."),
    ("Gross margin", f"={MM}{M30}{MR['m_gp']}", None, PCT,
     "Cartrack did not disclose gross profit in FY2019. Karooooo's Cartrack segment runs 72% today "
     "against McEasy's 75.6% SaaS-only line — near-identical software unit economics."),
    ("EBITDA margin", f"={MM}{M30}{MR['m_eb']}", f"={K}{V19}{R['m_ebrev']}", PCT,
     "THE headline gap: 7.7pp. Do not lead with this number. Block 2 explains why it is a "
     "denominator effect."),
    ("Net margin", f"={MM}{M30}{MR['m_np']}", f"={K}{V19}{R['m_nprev']}", PCT,
     "10.9pp gap, entirely attributable to lower depreciation per vehicle."),
    ("Depreciation % of revenue", f"={MM}{M30}{MR['m_da']}", f"={K}{V19}{R['m_ebrev']}"
     f"-{K}{V19}{R['m_oprev']}", PCT,
     "Cartrack's is EBITDA margin less operating margin. McEasy carries a lighter asset base."),
    ("Capex % of revenue", f"={MM}{M30}{MR['m_cpx']}", f"={K}{V19}{R['m_cpxrev']}", PCT,
     "11.4pp lighter. The single most important assumption in the whole plan."),
    ("Recurring % of revenue", f"={MM}{M30}{MR['rev_saas']}/{MM}{M30}{MR['rev_tot']}",
     f"={K}{V19}{R['m_subrev']}", PCT,
     "82.1% versus 89.8%. McEasy carries more low-margin spareparts revenue, which is exactly what "
     "drags the gross margin and flatters the EBITDA margin."),
]
MVB1 = r
for lab, a, b, fmt, note in m1:
    put(mvb, r, 1, lab, size=9)
    put(mvb, r, 2, a, fmt=fmt, color=GREEN, bold=True, halign="center")
    put(mvb, r, 3, b, fmt=fmt, color=GREEN if b else GREY, halign="center")
    put(mvb, r, 4, f'=IF(COUNT(B{r},C{r})=2,B{r}-C{r},"")',
        fmt=PCT if fmt is PCT else NUM, halign="center")
    put(mvb, r, 5, note, size=8.5, color=GREY, wrap=True)
    mvb.row_dimensions[r].height = 32
    r += 1

r += 1
section(mvb, r, "Block 2 — per vehicle per year. THIS IS THE ARGUMENT.", 5); r += 1
for i, lab in enumerate(["Metric", "McEasy 2030", "Cartrack FY2019", "Delta %", "Comment"], start=1):
    put(mvb, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
CT = {  # Cartrack FY2019 totals in USD, built from the Karooooo Path sheet
    "rev": f"{K}{V19}{R['rev_u']}*1000000",
    "sub": f"{K}{V19}{R['sub_u']}*1000000",
    "eb": f"{K}{V19}{R['eb_u']}*1000000",
    "da": f"({K}{V19}{R['eb']}-{K}{V19}{R['op']})/{CT_FX}*1000000",
    "cpx": f"{K}{V19}{R['cpx']}/{CT_FX}*1000000",
    "np": f"{K}{V19}{R['np']}/{CT_FX}*1000000",
}
m2 = [
    ("Vehicles / subscribers", f"={VEH}", f"={CT_SUBS}", CNT, True,
     "950,000 against 960,798 — 1.1% apart. Not approximately the same scale; the same scale."),
    ("Revenue", f"={MM}{M30}{MR['rev_tot']}/{VEH}", f"=({CT['rev']})/{CT_SUBS}", USD2, False,
     "McEasy earns ~17% less per vehicle, because it passes through less hardware revenue."),
    ("SaaS / subscription revenue", f"={MM}{M30}{MR['rev_saas']}/{VEH}",
     f"=({CT['sub']})/{CT_SUBS}", USD2, False,
     "~24% less. McEasy is assuming LESS pricing power than the benchmark achieved, not more."),
    ("EBITDA", f"={MM}{M30}{MR['ebitda']}/{VEH}", f"=({CT['eb']})/{CT_SUBS}", USD2, True,
     "2.3% apart. The two businesses generate the same cash per vehicle. This single row answers "
     "the margin objection."),
    ("Cash cost (revenue less EBITDA)",
     f"=({MM}{M30}{MR['rev_tot']}-{MM}{M30}{MR['ebitda']})/{VEH}",
     f"=(({CT['rev']})-({CT['eb']}))/{CT_SUBS}", USD2, False,
     "28% lower. Defensible on Indonesian versus South African fully-loaded cost per head — but it "
     "needs a headcount plan behind it."),
    ("Depreciation", f"={MM}{M30}{MR['da']}/{VEH}", f"=({CT['da']})/{CT_SUBS}", USD2, False,
     "39% lower, and the direct cause of the net-margin gap."),
    ("Capex", f"={MM}{M30}{MR['capex']}/{VEH}", f"=({CT['cpx']})/{CT_SUBS}", USD2, False,
     "48% lower. This is the one claim an investor should test — see Block 3."),
    ("Net profit", f"={MM}{M30}{MR['np']}/{VEH}", f"=({CT['np']})/{CT_SUBS}", USD2, False,
     "26% HIGHER, entirely explained by the lower depreciation above. One coherent story, not three "
     "unrelated stretches."),
    ("ARPU — total revenue per month", f"={MM}{M30}{MR['rev_tot']}/{VEH}/12",
     f"=({CT['rev']})/{CT_SUBS}/12", USD2, False, "USD8.94 against USD10.72."),
    ("ARPU — SaaS only per month", f"={MM}{M30}{MR['rev_saas']}/{VEH}/12",
     f"=({CT['sub']})/{CT_SUBS}/12", USD2, False, "USD7.34 against USD9.63."),
]
for lab, a, b, fmt, hi in [(x[0], x[1], x[2], x[3], x[4]) for x in m2]:
    note = [x[5] for x in m2 if x[0] == lab][0]
    put(mvb, r, 1, lab, size=9, bold=hi)
    put(mvb, r, 2, a, fmt=fmt, color=GREEN, bold=True, halign="center",
        fill=YEL if hi else None)
    put(mvb, r, 3, b, fmt=fmt, color=GREEN, bold=hi, halign="center")
    put(mvb, r, 4, f'=IF(AND(COUNT(B{r},C{r})=2,C{r}<>0),B{r}/C{r}-1,"")', fmt=PCT,
        bold=hi, halign="center")
    put(mvb, r, 5, note, size=8.5, color=GREY, wrap=True)
    mvb.row_dimensions[r].height = 32
    r += 1

r += 1
section(mvb, r, "Block 3 — the one remaining claim: cost per NEW INSTALL, not per base vehicle", 5)
r += 1
put(mvb, r, 1, "McEasy 2030 capex per NEW INSTALLATION", size=10, bold=True)
put(mvb, r, 2, f"={MM}{M30}{MR['capex']}/({VEH}-({VEH}*"
               f"{MM}{get_column_letter(MCOL[2029])}{MR['rev_saas']}/{MM}{M30}{MR['rev_saas']}))",
    fmt=USD2, bold=True, color=GREEN, halign="center", fill=YEL)
put(mvb, r, 5, "Capex is spent on new units, not on the whole base. This is the correct denominator "
               "and the number to quote. Full derivation on the 'Capex Decomposition' sheet.",
    size=8.5, color=GREY, wrap=True)
mvb.row_dimensions[r].height = 30
B3 = r; r += 1
put(mvb, r, 1, "Bottom-up bill of materials for one basic-tracker install", size=9)
put(mvb, r, 2, "$26 – $46", size=10, bold=True, halign="center")
put(mvb, r, 5, "4G Cat-1 tracker USD15-25 landed, installation labour USD8-15, SIM and logistics "
               "USD3-6. The model's budget sits ABOVE this — it carries headroom rather than being "
               "aggressive. That is the answer to the capex question.",
    size=8.5, color=GREY, wrap=True)
mvb.row_dimensions[r].height = 34
r += 1
put(mvb, r, 1, "Cartrack FY2019 capex per net addition, as reported", size=9)
put(mvb, r, 2, f"=({CT['cpx']})/({CT_SUBS}-{K}B{R['subs']})", fmt=USD2, color=GREEN,
    halign="center")
put(mvb, r, 5, "Not comparable as reported: the line is 'PP&E AND CONTRACT ASSETS' and bundles "
               "capitalised commissions with stolen-vehicle-recovery infrastructure — ground and air "
               "recovery teams, control rooms, recovery fleet. Strip 30-40% and it lands near "
               "USD106-124 per install.", size=8.5, color=GREY, wrap=True)
mvb.row_dimensions[r].height = 40
r += 1
put(mvb, r, 1, "Supporting evidence and the counter-risks", size=10, bold=True, color="1F3864")
r += 1
for ev in [
    "Queclink, the listed device vendor: FY2025 revenue FELL 7.8% while gross margin compressed from "
    "40.1% to 36.9% and operating margin collapsed from 13.4% to 5.4%. Streamax: revenue FELL 10.8%. "
    "A device maker losing pricing power is a device buyer gaining it. Both on 'At 100M (Today)'.",
    "Your own accounts carry the strongest argument: Sparepart Solutions is USD18.2m of 2030 revenue "
    "at a 21.2% gross margin — pass-through hardware. A device the customer BUYS is cost of revenue, "
    "not capex. Cartrack subsidised and retained its devices, so its spend sat in PP&E. Say the flip "
    "side too: this relocates cost rather than removing it, which is why the blended gross margin is "
    "65.9% against 75.6% SaaS-only — and why EBITDA per vehicle still matches.",
    "Counter-risk one: on the per-install denominator the gap versus Cartrack as reported is about "
    "-69%, WIDER than the -48% a per-base-vehicle comparison implies. Concede this before an investor "
    "derives it, then point at the bill of materials.",
    "Counter-risk two: at a blended dashcam install cost, only a single-digit percentage of new "
    "installs can be AI video units before this budget breaks. Check it against the product roadmap.",
    "Counter-risk three: capex here is derived as the change in GROSS fixed assets, so it excludes "
    "device retirement and replacement. If churned devices are not recovered and reused, true capex "
    "per install is higher and this argument weakens.",
]:
    put(mvb, r, 1, "•", size=9, bold=True, halign="center")
    put(mvb, r, 2, ev, size=8.5, color=GREY, wrap=True)
    mvb.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    mvb.row_dimensions[r].height = 34
    r += 1

r += 1
section(mvb, r, "Block 4 — bull / base / bear, each anchored on a real comparable", 5); r += 1
for i, lab in enumerate(["Scenario", "EBITDA margin", "EBITDA (USD m)", "Net (USD m)",
                         "Anchor and implied enterprise value"], start=1):
    put(mvb, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
REV30 = f"{MM}{M30}{MR['rev_tot']}/1000000"
# NOTE: these three refs are stored WITHOUT a leading "=" because they get embedded inside
# larger formulas below. A leading "=" here produces "=IFERROR(=...)", which Excel rejects
# outright — the file will not open at all. The guard before wb.save() catches it now.
scen = [
    ("Bear — Microlise economics", f"{T}Q{TD_ROW['Microlise']}", f"{T}P{TD_ROW['Microlise']}",
     f"'Valuation Signals'!G{VS_FIRST + 6}",
     "Microlise runs 10% adjusted EBITDA and a statutory operating loss at USD78m ARR, and trades at "
     "0.39x sales. USD100m of ARR at these economics is worth about USD40m of enterprise value. This "
     "is the real downside and it is not theoretical."),
    ("Base — Cartrack FY2019 economics", f"{K}{V19}{R['m_ebrev']}", f"{K}{V19}{R['m_nprev']}",
     f"'Valuation Signals'!G{VS_FIRST + 4}",
     "45.0% EBITDA and 21.3% net, valued on Quartix's 2.93x sales — a profitable, listed, "
     "modest-growth telematics business. The defensible ask: we match the best emerging-market "
     "comparable at the same scale."),
    ("Bull — McEasy's own model", f"{MM}{M30}{MR['m_eb']}", f"{MM}{M30}{MR['m_np']}",
     f"'Valuation Signals'!G{VS_FIRST}",
     "52.7% EBITDA and 32.2% net, valued on Karooooo's 5.70x sales. Requires the device-cost claim "
     "in Block 3 to hold AND the growth in the next sheet to be delivered."),
]
SC_FIRST = r
for lab, ebm, npm, mult, note in scen:
    put(mvb, r, 1, lab, size=9, bold=True)
    put(mvb, r, 2, f"={ebm}", fmt=PCT, color=GREEN, halign="center")
    put(mvb, r, 3, f'=IFERROR(B{r}*{REV30},"")', fmt=NUM, bold=True, halign="center")
    put(mvb, r, 4, f'=IFERROR({npm}*{REV30},"")', fmt=NUM, halign="center")
    put(mvb, r, 5, note, size=8.5, color=GREY, wrap=True)
    mvb.row_dimensions[r].height = 46
    r += 1
put(mvb, r, 1, "Implied enterprise value at the anchor's sales multiple (USD m)",
    size=9, bold=True)
for i, (lab, ebm, npm, mult, note) in enumerate(scen):
    put(mvb, r, 2 + i, f'=IFERROR({mult}*{REV30},"")', fmt=NUM0, bold=True, halign="center")
put(mvb, r, 5, "Bear / base / bull, left to right. The spread is what an investor is actually "
               "underwriting — the ARR figure alone tells them almost nothing.",
    size=8.5, color=GREY, wrap=True)
r += 2

notes_block(mvb, r, [
    "How to run the meeting: open on Block 2, not Block 1. 'We are targeting the same subscriber "
    "base as Cartrack at its USD100m milestone, at a LOWER price per vehicle, generating the same "
    "EBITDA per vehicle.' That is a matching claim, not a beating claim, and it is very hard to "
    "argue with.",
    "Then pre-empt the one real objection yourself by putting Block 3 up: 'the only place we differ "
    "is device cost per vehicle, and here is why an Indonesian device in 2030 costs half a South "
    "African device in 2019.' Volunteering your own weakest link buys enormous credibility.",
    "Everything in the McEasy column flows from one yellow cell on the 'McEasy Model' sheet — the "
    "2030 vehicle count. Flex it live in the room: at 900,000 vehicles EBITDA per vehicle is USD59.68 "
    "(+3.2% versus Cartrack), at 950,000 it is USD56.54 (-2.2%) and at 1,000,000 it is USD53.72 "
    "(-7.1%). The conclusion holds across the WHOLE range you quoted — it does not depend on picking "
    "a favourable vehicle count, which is what makes it safe to demonstrate live.",
    "One reconciliation to tighten first: 900,000 vehicles implies USD9.43 per month, just above the "
    "USD8-9 band. The self-consistent combination is ~943k-1,061k vehicles at USD8.00-9.00. See the "
    "check rows on the 'McEasy Model' sheet.",
    "For the downside, go to 'Conservative Cases'. It re-bases the whole 2026-2030 path to USD90m and "
    "USD80m of 2030 revenue and shows that this argument survives a VOLUME miss — EBITDA per vehicle "
    "stays within 7% of Cartrack even at USD80m — but not a PRICE miss, which lands 36% below the "
    "benchmark. If an investor pushes on the downside, that is the distinction to draw.",
], ncols=5)

# ================================================= GROWTH REALITY CHECK
gr = wb.create_sheet("Growth Reality Check")
title(gr, "The assumption that is genuinely more aggressive than the benchmark",
      "It is not the margin. It is delivering a 66% ARR CAGR while turning EBITDA-positive in 2026 "
      "and raising no further equity. No company in this peer set has done both at once.")
gr.column_dimensions["A"].width = 40
for c, w in zip("BCDE", (16, 16, 16, 76)):
    gr.column_dimensions[c].width = w

r = 4
section(gr, r, "What the plan requires versus what the benchmark delivered", 5); r += 1
for i, lab in enumerate(["Company / plan", "Period", "ARR or revenue CAGR", "FCF margin",
                         "Comment"], start=1):
    put(gr, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
grow = [
    ("McEasy — required", "2025-2030", f"={MM}$B${MR['cagr']}", None,
     "ARR from USD7.95m to USD100.9m: a 12.7x increase in five years.", True),
    ("Cartrack — realised", "FY2019-FY2026", f'=({K}{LAST}{R["arr_u"]}/{K}{V19}{R["sub_u"]})'
     f'^(1/7)-1', f"={K}{LAST}{R['m_fcfrev']}",
     "Cartrack's MATURE phase, from USD111m to USD325m. Not the right comparison for McEasy's stage, "
     "but it is what the benchmark actually did after the milestone.", False),
    ("Cartrack — at the milestone", "FY2019", f'={K}{V19}{R["rev"]}/{K}B{R["rev"]}-1',
     f"={K}{V19}{R['m_fcfrev']}",
     "Revenue growth in the year it crossed USD100m ARR: 27.8%, with a 2.2% FCF margin. Fast growth "
     "and free cash flow did not coexist even here.", False),
    ("Netradyne", "2023-2024", 0.62, None,
     "The only company in the set growing at McEasy's required rate. Private, USD1.3bn valuation, no "
     "disclosed profitability, venture-funded.", False),
    ("Motive", "LTM Sep-2025", 0.28, -0.23,
     "Grew to USD501m ARR and still burns 23% of revenue. Filed to IPO in Dec-2025 and remains "
     "unpriced. This is what funding hypergrowth looks like.", False),
    ("Karooooo", "FY2026", 0.1997, None,
     "20% growth at a 42% EBITDA margin and positive free cash flow — the profitable-growth case, at "
     "a third of McEasy's required rate.", False),
    ("Quartix", "FY2025", 0.123, 0.1252, "Profitable and slow. 2.93x sales.", False),
    ("Microlise", "FY2025", 0.057, 0.1127, "Slow AND unprofitable. 0.39x sales.", False),
]
for lab, per, cagr, fcf, note, hi in grow:
    put(gr, r, 1, lab, size=9, bold=hi)
    put(gr, r, 2, per, size=9, halign="center", color=GREY)
    put(gr, r, 3, cagr, fmt=PCT, color=GREEN if isinstance(cagr, str) else BLUE, bold=hi,
        halign="center", fill=YEL if hi else None)
    put(gr, r, 4, fcf, fmt=PCT, color=GREEN if isinstance(fcf, str) else
        (BLUE if fcf is not None else GREY), halign="center")
    put(gr, r, 5, note, size=8.5, color=GREY, wrap=True)
    gr.row_dimensions[r].height = 34
    r += 1
put(gr, r, 1, "McEasy's required rate as a multiple of Cartrack's realised rate", size=9, bold=True)
put(gr, r, 3, f'=IFERROR(C{r - 8}/C{r - 7},"")', fmt="0.0x", bold=True, halign="center")
put(gr, r, 5, "In fairness: 66% off a USD7.95m base is a very different proposition from 66% off "
               "USD100m, and Cartrack's own growth at USD8m of ARR is not disclosed. The honest "
               "framing is that McEasy needs Netradyne-class growth with Cartrack-class margins.",
    size=8.5, color=GREY, wrap=True)
gr.row_dimensions[r].height = 40
r += 2

section(gr, r, "Funding: the plan self-funds from 2026 onward", 5); r += 1
for i, lab in enumerate(["Year", "Cash at year end (USD)", "EBITDA (USD)",
                         "Paid-up capital (USD)", "Comment"], start=1):
    put(gr, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
CASH_NOTE = {
    2025: "Thinnest point in the plan: cash of USD1.9m against a USD1.6m EBITDA loss — under "
          "twelve months of runway before the 2026 raise.",
    2026: "Paid-up capital rises USD4.43m and EBITDA turns positive in the same year. Both have to "
          "happen; the plan has no slack if either slips.",
    2027: "Last year of low cash. From here the model is self-funding.",
    2030: "USD42.4m of cash and no equity raised since 2026.",
}
for y in (2025, 2026, 2027, 2028, 2029, 2030):
    cl = get_column_letter(MCOL[y])
    put(gr, r, 1, str(y), size=9, halign="center", bold=y in CASH_NOTE)
    put(gr, r, 2, f"={MM}{cl}{MR['cash']}", fmt=NUM0, color=GREEN, halign="center")
    put(gr, r, 3, f"={MM}{cl}{MR['ebitda']}", fmt=NUM0, color=GREEN, halign="center")
    put(gr, r, 4, f"={MM}{cl}{MR['puc']}", fmt=NUM0, color=GREEN, halign="center")
    put(gr, r, 5, CASH_NOTE.get(y, ""), size=8.5, color=GREY, wrap=True)
    gr.row_dimensions[r].height = 30
    r += 1
r += 1
notes_block(gr, r, [
    "The investor question this sheet exists to answer: 'you are asking me to believe you will grow "
    "four times faster than Cartrack did and reach better unit economics than Cartrack ever had, "
    "without raising again.' Blocks 2 and 3 of the benchmark sheet answer the unit-economics half. "
    "This sheet is where you should concede that the growth half is the genuine risk, and show the "
    "bear case is survivable.",
    "The single most useful thing you can add to the model before the meeting is a year-by-year "
    "vehicle and ARPU build. Without it, the 66% CAGR is an assertion; with it, it becomes a sales "
    "capacity and installation-throughput question that you can actually defend with an operating plan.",
], ncols=5)

# ==================================================== CONSERVATIVE CASES
import conservative_engine as CE                                        # noqa: E402

CM = CE.load()
CE_LIFE, CE_CUM, CE_TGT = CE.calibrate_opening_block(CM)
CE_REF = CE.base_cohort(CM, CE_LIFE)
CKEYS = CM["keys"]
PLAN30 = sum(CM["P"]["rev"][k] for k in CKEYS if k.startswith("2030"))

SCEN = [("base", "Base — the plan", PLAN30, "volume"),
        ("v90", "$90m — fewer vehicles", 90_000_000.0, "volume"),
        ("a90", "$90m — lower ARPU", 90_000_000.0, "arpu"),
        ("v80", "$80m — fewer vehicles", 80_000_000.0, "volume"),
        ("a80", "$80m — lower ARPU", 80_000_000.0, "arpu")]
YRS5 = (2026, 2027, 2028, 2029, 2030)
CS, CRESID = {}, None
for key, lab, tgt, drv in SCEN:
    S = CE.build(CM, tgt, drv, CE_LIFE, CE_REF)
    if key == "base":
        _, CRESID = CE.cash_path(CM, S, None)
    path, _ = CE.cash_path(CM, S, CRESID)
    A = {n: CE.annual(S[n], CKEYS, YRS5)
         for n in ("rev", "saas", "spare", "cogs_saas", "cogs_spare", "logi", "sm",
                   "opex", "ebitda", "dep", "tax", "np", "capex")}
    mn = min(path.items(), key=lambda kv: kv[1])
    CS[key] = dict(S=S, A=A, path=path, veh=S["veh"]["2030-12"], h=S["h"],
                   min_cash=mn[1], min_month=mn[0],
                   cash={y: path.get(f"{y}-12", 0.0) for y in YRS5})

cc = wb.create_sheet("Conservative Cases")
title(cc, "Conservative cases — 2030 revenue $10m and $20m below plan",
      "Re-based month by month from Jul-2026 off observed H1 actuals. Two drivers per revenue level, "
      "because a volume miss and a price miss produce opposite conclusions at the same revenue.")
cc.column_dimensions["A"].width = 50
for c, w in zip("BCDEF", (15,) * 5):
    cc.column_dimensions[c].width = w
cc.column_dimensions["G"].width = 68
SC_COL = {k: 2 + i for i, (k, *_rest) in enumerate(SCEN)}
BASEC = get_column_letter(SC_COL["base"])

r = 4
section(cc, r, "Block 1 — why the path diverges from Jul-2026, not from plan", 7); r += 1
h1 = [k for k in CKEYS if k.startswith("2026-0") and int(k[-2:]) <= 6]
g_obs = (CM["P"]["arr"][h1[-1]] / CM["P"]["arr"][h1[0]]) ** (1 / (len(h1) - 1)) - 1
jul_req = CM["P"]["arr"]["2026-07"] / CM["P"]["arr"]["2026-06"] - 1
exit_obs = CM["P"]["arr"]["2026-06"] * (1 + g_obs) ** 6
arr_mult = 1.0
for y in (2027, 2028, 2029, 2030):
    arr_mult *= CM["P"]["arr"][f"{y}-12"] / CM["P"]["arr"][f"{y - 1}-12"]
for lab, val, fmtx, note in [
    ("Monthly ARR growth, H1-2026 actual", "3.54% → 2.98% → 2.93% → 2.83% → 2.65%", None,
     "Decelerating through the half. These are reported actuals, not estimates."),
    ("Average observed monthly ARR growth", g_obs, PCT,
     "Geometric mean, Jan to Jun 2026."),
    ("Jun-2026 actual ARR", CM["P"]["arr"]["2026-06"], NUM0, "Last actual month in the model."),
    ("Plan Dec-2026 exit ARR", CM["P"]["arr"]["2026-12"], NUM0, "The 2026 target."),
    ("Growth the plan needs in Jul-2026 alone", jul_req, PCT,
     "A step-change from the observed ~2.7% to nearly 13% in a single month, then about 6% a month "
     "for the rest of the year. This is the assumption the conservative cases test."),
    ("Dec-2026 exit ARR if observed momentum continues", exit_obs, NUM0, ""),
    ("Gap to plan on that basis", None, PCT, "Formula below."),
    ("Plan ARR multiple, 2026 exit to 2030 exit", arr_mult, "0.000x",
     "The plan's own compounding, applied to a lower starting point."),
    ("Implied 2030 ARR at observed momentum", exit_obs * arr_mult, NUM0,
     "So the $80m case is approximately the CURRENT TRAJECTORY, not an arbitrary haircut, and $90m "
     "is a partial-recovery case. That is a far stronger way to present these."),
]:
    put(cc, r, 1, lab, size=9, bold="Implied 2030" in lab or "Growth the plan" in lab)
    if lab.startswith("Gap to plan"):
        put(cc, r, 2, f"=B{r - 3}/B{r - 5}-1", fmt=PCT, bold=True, halign="center")
    elif fmtx is None:
        put(cc, r, 2, val, size=9, color=BLUE, halign="center")
        cc.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    else:
        put(cc, r, 2, val, fmt=fmtx, color=BLUE,
            bold="Implied 2030" in lab or "Growth the plan" in lab, halign="center")
    if note:
        put(cc, r, 7, note, size=8.5, color=GREY, wrap=True)
        cc.row_dimensions[r].height = 34 if len(note) > 90 else 24
    r += 1

r += 1
section(cc, r, "Block 2 — 2030 profit and loss by scenario", 7); r += 1
for k, lab, tgt, drv in SCEN:
    c = SC_COL[k]
    put(cc, r, c, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
put(cc, r, 1, "USD, year to Dec-2030", bold=True, size=9, fill=SUB_FILL)
put(cc, r, 7, "Notes", bold=True, size=9, fill=SUB_FILL)
cc.row_dimensions[r].height = 32
r += 1
CR = {}
pl_block = [
    ("h", "Growth haircut applied from Jul-2026", "0.0000", "h"),
    ("rev", "Total revenue", NUM0, "A"), ("saas", "  SaaS Solutions", NUM0, "A"),
    ("spare", "  Spare Part Solutions", NUM0, "A"),
    ("cogs_saas", "Cost of revenue — SaaS", NUM0, "A"),
    ("cogs_spare", "Cost of revenue — Spare Part", NUM0, "A"),
    ("logi", "Logistic cost", NUM0, "A"), ("sm", "Sales & marketing", NUM0, "A"),
    ("opex", "Operational expenses", NUM0, "A"),
    ("ebitda", "EBITDA", NUM0, "A"), ("dep", "Depreciation", NUM0, "A"),
    ("tax", "Tax & interest", NUM0, "A"), ("np", "Net profit", NUM0, "A"),
    ("capex", "Capex", NUM0, "A"),
]
for key, lab, fmtx, src in pl_block:
    put(cc, r, 1, lab, size=9, bold=key in ("rev", "ebitda", "np"))
    for k, *_ in SCEN:
        v = CS[k]["h"] if src == "h" else CS[k]["A"][key][2030]
        put(cc, r, SC_COL[k], v, fmt=fmtx, color=BLUE,
            bold=key in ("rev", "ebitda", "np"), halign="center")
    CR[key] = r
    if key == "h":
        put(cc, r, 7, "One lever: the plan's own monthly ARR growth rates multiplied by h from "
                      "Jul-2026, solved so FY2030 revenue hits the target.", size=8.5,
            color=GREY, wrap=True)
        cc.row_dimensions[r].height = 30
    r += 1
for lab, num, den, key in [("Gross margin", None, "rev", "gm"),
                           ("EBITDA margin", "ebitda", "rev", "ebm"),
                           ("Net margin", "np", "rev", "npm"),
                           ("Capex % of revenue", "capex", "rev", "cxm")]:
    put(cc, r, 1, lab, size=9, bold=key == "ebm")
    for k, *_ in SCEN:
        cl = get_column_letter(SC_COL[k])
        if key == "gm":
            f = (f'=IFERROR((({cl}{CR["rev"]}-{cl}{CR["cogs_saas"]}-{cl}{CR["cogs_spare"]})'
                 f'/{cl}{CR["rev"]}),"")')
        else:
            f = f'=IFERROR({cl}{CR[num]}/{cl}{CR[den]},"")'
        put(cc, r, SC_COL[k], f, fmt=PCT, bold=key == "ebm", halign="center")
    CR[key] = r
    r += 1
put(cc, r, 1, "Sparepart share of revenue (must hold constant)", size=9)
for k, *_ in SCEN:
    cl = get_column_letter(SC_COL[k])
    put(cc, r, SC_COL[k], f'=IFERROR({cl}{CR["spare"]}/{cl}{CR["rev"]},"")', fmt=PCT,
        halign="center")
put(cc, r, 7, "Grady's constraint. Holds at 17.85% in every scenario because the plan's own monthly "
              "revenue mix is preserved.", size=8.5, color=GREY, wrap=True)
cc.row_dimensions[r].height = 30
r += 2

section(cc, r, "Block 3 — the benchmark check. This is where the two drivers part company.", 7)
r += 1
put(cc, r, 1, "Vehicles at Dec-2030", size=9)
for k, *_ in SCEN:
    put(cc, r, SC_COL[k], CS[k]["veh"], fmt=CNT, color=BLUE, halign="center")
CR["veh"] = r; r += 1
put(cc, r, 1, "Revenue per vehicle per month", size=9)
for k, *_ in SCEN:
    cl = get_column_letter(SC_COL[k])
    put(cc, r, SC_COL[k], f'=IFERROR({cl}{CR["rev"]}/{cl}{CR["veh"]}/12,"")', fmt=USD2,
        halign="center")
put(cc, r, 7, "Vehicles are derived from exit ARR at constant ARR-per-vehicle, so slower growth "
              "mechanically lifts this ratio slightly in the volume cases — full-year revenue is "
              "measured against a smaller exit base. A definitional artefact, not a price rise.",
    size=8.5, color=GREY, wrap=True)
cc.row_dimensions[r].height = 40
r += 1
put(cc, r, 1, "EBITDA per vehicle per year", size=10, bold=True)
for k, *_ in SCEN:
    cl = get_column_letter(SC_COL[k])
    put(cc, r, SC_COL[k], f'=IFERROR({cl}{CR["ebitda"]}/{cl}{CR["veh"]},"")', fmt=USD2,
        bold=True, halign="center", fill=YEL)
CR["ebpv"] = r; r += 1
put(cc, r, 1, "Cartrack FY2019 EBITDA per subscriber", size=9)
put(cc, r, 2, f"={K}{V19}{R['eb_u']}*1000000/{K}{V19}{R['subs']}", fmt=USD2, color=GREEN,
    halign="center")
cc.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
CR["ct"] = r; r += 1
put(cc, r, 1, "Gap to the benchmark", size=10, bold=True)
for k, *_ in SCEN:
    cl = get_column_letter(SC_COL[k])
    put(cc, r, SC_COL[k], f'=IFERROR({cl}{CR["ebpv"]}/$B${CR["ct"]}-1,"")', fmt=PCT, bold=True,
        halign="center")
put(cc, r, 7, "THE FINDING: a volume miss leaves the Cartrack argument intact — EBITDA per vehicle "
              "stays within 7% even at $80m. A price miss destroys it, landing 36% below the "
              "benchmark you are presenting against. Same revenue, opposite conclusion.",
    size=8.5, color=GREY, wrap=True)
cc.row_dimensions[r].height = 44
r += 2

section(cc, r, "Block 4 — re-based path, 2026 to 2030", 7); r += 1
for k, lab, tgt, drv in SCEN:
    put(cc, r, 1, lab, bold=True, size=9.5, color="1F3864")
    for i, y in enumerate(YRS5):
        put(cc, r, 2 + i, str(y), bold=True, size=9, halign="center", fill=SUB_FILL)
    put(cc, r, 7, f"Haircut h = {CS[k]['h']:.4f}", size=8.5, color=GREY)
    r += 1
    for key, lab2 in [("rev", "Revenue"), ("ebitda", "EBITDA"), ("dep", "Depreciation"),
                      ("tax", "Tax & interest"), ("np", "Net profit"), ("capex", "Capex")]:
        put(cc, r, 1, "   " + lab2, size=9)
        for i, y in enumerate(YRS5):
            put(cc, r, 2 + i, CS[k]["A"][key][y], fmt=NUM0, color=BLUE, halign="center")
        r += 1
    put(cc, r, 1, "   Free cash flow (EBITDA less tax less capex)", size=9)
    for i, y in enumerate(YRS5):
        cl = get_column_letter(2 + i)
        put(cc, r, 2 + i, f"={cl}{r - 5}-{cl}{r - 3}-{cl}{r - 1}", fmt=NUM0, halign="center")
    r += 1
    put(cc, r, 1, "   Closing cash", size=9, bold=True)
    for i, y in enumerate(YRS5):
        put(cc, r, 2 + i, CS[k]["cash"][y], fmt=NUM0, color=BLUE, bold=True, halign="center")
    r += 2

section(cc, r, "Block 5 — the cash trough and the funding gap", 7); r += 1
for k, lab, tgt, drv in SCEN:
    put(cc, r, SC_COL[k], lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
put(cc, r, 1, "Cash and funding", bold=True, size=9, fill=SUB_FILL)
cc.row_dimensions[r].height = 32
r += 1
put(cc, r, 1, "Minimum cash on the whole path", size=10, bold=True)
for k, *_ in SCEN:
    put(cc, r, SC_COL[k], CS[k]["min_cash"], fmt=NUM0, color=BLUE, bold=True, halign="center")
CR["minc"] = r; r += 1
put(cc, r, 1, "Month it occurs", size=9)
for k, *_ in SCEN:
    put(cc, r, SC_COL[k], CS[k]["min_month"], size=9, color=BLUE, halign="center")
put(cc, r, 7, "Every scenario troughs in 2027 — including the plan itself. The pinch point is not "
              "2030, it is the year after next.", size=8.5, color=GREY, wrap=True)
cc.row_dimensions[r].height = 30
r += 1
put(cc, r, 1, "Buffer: 6 months of 2027 operating expenses", size=9)
buf27 = CS["base"]["A"]["opex"][2027] / 12 * 6
put(cc, r, 2, buf27, fmt=NUM0, color=BLUE, halign="center", fill=YEL)
cc.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
CR["buf"] = r; r += 1
put(cc, r, 1, "Funding gap against that buffer", size=10, bold=True)
for k, *_ in SCEN:
    cl = get_column_letter(SC_COL[k])
    put(cc, r, SC_COL[k], f"=MAX(0,$B${CR['buf']}-{cl}{CR['minc']})", fmt=NUM0, bold=True,
        halign="center")
put(cc, r, 7, "Shown, not plugged. Note the plan itself is already about $0.5m short of a six-month "
              "buffer at its May-2027 trough, and the worst case is only about $1.3m short. The 2027 "
              "path is thin but not fragile — a modest facility covers every case here.",
    size=8.5, color=GREY, wrap=True)
cc.row_dimensions[r].height = 44
r += 2

notes_block(cc, r, [
    "METHOD. Jan–Jun 2026 are reported actuals and are untouched. From Jul-2026 the plan's own "
    "monthly ARR growth rates are multiplied by a single constant h per scenario, solved so FY2030 "
    "revenue equals the target. Monthly revenue follows ARR using the plan's own revenue-to-ARR "
    "ratio for that month, which preserves seasonality and the revenue mix.",
    "WHAT FLEXES. Volume driver: vehicles fall, price holds, so cost of revenue and capex scale with "
    "revenue. ARPU driver: the vehicle count holds and price falls, so SaaS cost of revenue and capex "
    "are held in absolute terms — the devices are already bought — while spare-part cost of revenue "
    "scales with spare-part revenue at plan margin.",
    "WHAT IS HELD. Everything below gross profit — logistics, sales & marketing and operating "
    "expenses — is held at plan in every scenario. That is full operating deleverage with no assumed "
    "cost cuts, which is the conservative reading and the one an investor will credit.",
    "DEPRECIATION. Device life is 48 months, which the actuals support: the implied blended life "
    "over H1-2026 averages 47.7 months. Depreciation is carried on a delta basis — the plan's own "
    "figure plus the incremental 48-month cohort effect of the capex change — so base and ARPU cases "
    "tie to the plan exactly while the volume case picks up the capex reduction with the correct lag. "
    f"The cohort reproduces reported accumulated depreciation at Jun-2026 to within "
    f"{abs(CE_CUM - CE_TGT) / CE_TGT:.2%}.",
    "CASH. Rolled monthly from the Jun-2026 actual balance as EBITDA less tax less capex less the "
    "change in working capital, with working capital scaled to trailing revenue. Non-modelled items "
    "are calibrated on the plan and carried into every scenario, so the base path reproduces the "
    "plan's cash at every year end to the cent. The 2026 equity raise is already in the Jun-2026 "
    "balance — paid-up capital reads $20,086,600 — so no further equity is assumed anywhere.",
    "THE HONEST CAVEAT. Margins fall only gently in the volume cases (52.7% to 49.1%) because "
    "operating expenses are just 9.5% of 2030 revenue. That shallow decline is a property of this "
    "model's very thin fixed-cost base, not a law of the business. A company carrying a normal "
    "fixed-cost load would be hurt considerably more by the same revenue miss.",
], ncols=7)

# ==================================================== CAPEX DECOMPOSITION
cx = wb.create_sheet("Capex Decomposition")
title(cx, "Capex per vehicle — decomposed, because the headline gap is the wrong number",
      "Capex is spent on NEW installations, not on the whole base. On the correct denominator the gap "
      "is wider, not narrower — and that is fine, because the absolute cost per install is what "
      "matters and it is conservative.")
cx.column_dimensions["A"].width = 48
for c, w in zip("BCDE", (16, 16, 13, 72)):
    cx.column_dimensions[c].width = w

CTC = f"{K}{V19}{R['cpx']}/{K}{V19}6*1000000"          # Cartrack FY2019 capex, USD
CT_ADDS = f"({K}{V19}{R['subs']}-{K}B{R['subs']})"      # FY2019 net adds
MC_CAPEX = f"{MM}{M30}{MR['capex']}"
V29 = f"({VEH}*{MM}{get_column_letter(MCOL[2029])}{MR['rev_saas']}/{MM}{M30}{MR['rev_saas']})"
MC_ADDS = f"({VEH}-{V29})"

r = 4
section(cx, r, "Block 1 — the denominator problem", 5); r += 1
for i, lab in enumerate(["Basis", "McEasy 2030", "Cartrack FY2019", "Delta", "Why it matters"],
                        start=1):
    put(cx, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
put(cx, r, 1, "Capex per vehicle in the BASE", size=9)
put(cx, r, 2, f"={MC_CAPEX}/{VEH}", fmt=USD2, color=GREEN, halign="center")
put(cx, r, 3, f"=({CTC})/{K}{V19}{R['subs']}", fmt=USD2, color=GREEN, halign="center")
put(cx, r, 4, f'=IFERROR(B{r}/C{r}-1,"")', fmt=PCT, halign="center")
put(cx, r, 5, "The figure originally shown. It divides a flow (a year's spend) by a stock (the whole "
              "installed base), so it is not a unit cost of anything. Do not lead with it.",
    size=8.5, color=GREY, wrap=True)
cx.row_dimensions[r].height = 34
r += 1
put(cx, r, 1, "Net additions in the year", size=9)
put(cx, r, 2, f"={MC_ADDS}", fmt=CNT, halign="center")
put(cx, r, 3, f"={CT_ADDS}", fmt=CNT, color=GREEN, halign="center")
put(cx, r, 4, None, color=GREY)
put(cx, r, 5, "McEasy's 2029 vehicle count is NOT in the model. It is proxied here by scaling the "
              "2030 count by the ratio of 2029 to 2030 SaaS revenue, i.e. assuming flat ARPU. "
              "Replace with the real ramp when you build one.", size=8.5, color=GREY, wrap=True)
cx.row_dimensions[r].height = 34
r += 1
put(cx, r, 1, "Capex per NET ADDITION", size=10, bold=True)
put(cx, r, 2, f"={MC_CAPEX}/{MC_ADDS}", fmt=USD2, bold=True, halign="center", fill=YEL)
put(cx, r, 3, f"=({CTC})/{CT_ADDS}", fmt=USD2, bold=True, color=GREEN, halign="center")
put(cx, r, 4, f'=IFERROR(B{r}/C{r}-1,"")', fmt=PCT, bold=True, halign="center")
put(cx, r, 5, "The economically correct denominator, and the gap is WIDER here. Concede this before "
              "an investor derives it. Then point at Block 2 — the absolute number is conservative.",
    size=8.5, color=GREY, wrap=True)
CX_PERADD = r
cx.row_dimensions[r].height = 34
r += 2

section(cx, r, "Block 2 — what one installation actually costs at 2026 prices", 5); r += 1
for i, lab in enumerate(["Component", "Low (USD)", "High (USD)", "", "Source / note"], start=1):
    put(cx, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
BOM_FIRST = r
for comp, lo, hi, note in [
    ("4G Cat-1 basic tracker, volume, landed Indonesia", 15, 25,
     "Concox / Jimi IoT wholesale runs USD12-49 with high-volume pricing as low as USD5-12; "
     "Teltonika FMB920 is about EUR35 at retail. Replace with your own quotes."),
    ("Installation labour, roughly one hour", 8, 15,
     "Indonesian technician, fully loaded. Materially below the South African equivalent."),
    ("SIM activation, logistics, spares, wastage", 3, 6, "Estimate — refine from your own actuals."),
]:
    put(cx, r, 1, comp, size=9, wrap=True)
    put(cx, r, 2, lo, fmt=USD0, color=BLUE, halign="center")
    put(cx, r, 3, hi, fmt=USD0, color=BLUE, halign="center")
    put(cx, r, 5, note, size=8.5, color=GREY, wrap=True)
    cx.row_dimensions[r].height = 30
    r += 1
BOM_LAST = r - 1
put(cx, r, 1, "Bill of materials — one basic tracker install", size=10, bold=True)
put(cx, r, 2, f"=SUM(B{BOM_FIRST}:B{BOM_LAST})", fmt=USD0, bold=True, halign="center")
put(cx, r, 3, f"=SUM(C{BOM_FIRST}:C{BOM_LAST})", fmt=USD0, bold=True, halign="center")
put(cx, r, 5, "Compare against the per-net-add budget above.", size=8.5, color=GREY, wrap=True)
BOM_LO, BOM_HI = f"B{r}", f"C{r}"
r += 1
put(cx, r, 1, "Headroom in the model's budget versus the high case", size=9, bold=True)
put(cx, r, 2, f"=B{CX_PERADD}-{BOM_HI}", fmt=USD2, bold=True, halign="center")
put(cx, r, 5, "Positive means the model is CONSERVATIVE per install, not aggressive. This is the "
              "single most useful sentence you can say on this topic.", size=8.5, color=GREY,
    wrap=True)
cx.row_dimensions[r].height = 30
r += 1
put(cx, r, 1, "AI dashcam / video unit install, all-in", size=9)
put(cx, r, 2, 80, fmt=USD0, color=BLUE, halign="center")
put(cx, r, 3, 200, fmt=USD0, color=BLUE, halign="center")
put(cx, r, 5, "Materially dearer than a basic tracker.", size=8.5, color=GREY, wrap=True)
DASH_LO, DASH_HI = f"B{r}", f"C{r}"
r += 1
put(cx, r, 1, "Maximum dashcam share the budget supports", size=10, bold=True)
put(cx, r, 2, f'=IFERROR(MAX(0,(B{CX_PERADD}-{BOM_HI})/(AVERAGE({DASH_LO},{DASH_HI})-{BOM_HI})),"")',
    fmt=PCT, bold=True, halign="center", fill=YEL)
put(cx, r, 5, "At a blended dashcam install cost, this is the share of new installs that can be "
              "video units before the capex budget breaks. CHECK THIS AGAINST THE PRODUCT ROADMAP — "
              "if video is a bigger part of the 2030 mix, capex is understated.",
    size=8.5, color=GREY, wrap=True)
cx.row_dimensions[r].height = 40
r += 2

section(cx, r, "Block 3 — RESOLVED. Cartrack's actual FY2019 device capex, from Note 5 of the "
               "Karooooo F-1", 5)
r += 1
put(cx, r, 1, "Source: Karooooo Ltd Form F-1 (Mar-2021), Note 5 'Property, plant and equipment' "
              "FY2019 movement table, and Note 6 'Capitalized commission assets'. Figures in ZAR "
              "thousands as reported; the reconciliation below ties exactly.",
    size=8.5, color=GREY, wrap=True)
cx.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
cx.row_dimensions[r].height = 30
r += 1
for i, lab in enumerate(["FY2019 line item (ZAR 000s)", "Amount", "% of cash-flow line", "",
                         "Note"], start=1):
    put(cx, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
CFL = 507151.0
recon = [
    ("Capitalised telematics devices — additions", 353655.0, True,
     "Note 5. The device AND its directly related installation costs, capitalised together and "
     "depreciated straight-line over the expected average contract life of 60 months. THIS is the "
     "figure a device-cost comparison should use."),
    ("Other owned-asset additions", 6836.0, False,
     "Buildings, furniture and fixtures, leasehold improvements, plant and machinery — ZAR6.8m in "
     "total. Immaterial."),
    ("Owned-asset additions = cash PP&E capex", 360491.0, False,
     "Note 5 total owned additions of ZAR360,491k. Right-of-use additions of ZAR74,364k (IT "
     "equipment, motor vehicles, property) are IFRS 16 lease recognitions, not cash capex, and are "
     "excluded here."),
    ("Cash-flow line: PP&E AND CONTRACT ASSETS", CFL, False,
     "As reported in the FY2019 SENS. This is the figure that produced the USD176.77 per add."),
    ("Residual = contract assets / capitalised commissions", 146660.0, False,
     "28.9% of the line. Note 6 confirms a separate 'Capitalized sales commissions' asset — cost "
     "ZAR178,330k at Feb-2019. McEasy expenses commissions through S&M, so this has no McEasy "
     "equivalent at all."),
]
for lab, amt, hi, note in recon:
    put(cx, r, 1, lab, size=9, bold=hi, wrap=True)
    put(cx, r, 2, amt, fmt=NUM0, color=BLUE, bold=hi, halign="center",
        fill=YEL if hi else None)
    put(cx, r, 3, f"=B{r}/{CFL}", fmt=PCT, halign="center")
    put(cx, r, 5, note, size=8.5, color=GREY, wrap=True)
    cx.row_dimensions[r].height = 42
    if hi:
        CX_DEV_ZAR = r
    r += 1

r += 1
for i, lab in enumerate(["Restated on device capex only", "McEasy 2030", "Cartrack FY2019",
                         "Gap", "Note"], start=1):
    put(cx, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
FX19 = f"{K}{V19}6"
DEVU = f"B{CX_DEV_ZAR}/{FX19}*1000"        # ZAR 000s -> USD
put(cx, r, 1, "Device + installation capex per NET ADDITION", size=10, bold=True)
put(cx, r, 2, f"=B{CX_PERADD}", fmt=USD2, bold=True, halign="center")
put(cx, r, 3, f"={DEVU}/{CT_ADDS}", fmt=USD2, bold=True, color=GREEN, halign="center")
put(cx, r, 4, f'=IFERROR(B{r}/C{r}-1,"")', fmt=PCT, bold=True, halign="center", fill=YEL)
put(cx, r, 5, "The gap to defend, correctly stated: device-and-install against device-and-install, "
              "per new unit. Down from -69% on the as-reported line.", size=8.5, color=GREY,
    wrap=True)
cx.row_dimensions[r].height = 34
r += 1
put(cx, r, 1, "Device + installation capex per BASE vehicle", size=10, bold=True)
put(cx, r, 2, f"={MC_CAPEX}/{VEH}", fmt=USD2, bold=True, halign="center")
put(cx, r, 3, f"={DEVU}/{K}{V19}{R['subs']}", fmt=USD2, bold=True, color=GREEN, halign="center")
put(cx, r, 4, f'=IFERROR(B{r}/C{r}-1,"")', fmt=PCT, bold=True, halign="center", fill=YEL)
put(cx, r, 5, "On the base-vehicle view the restated gap is modest. This is the most favourable "
              "true framing available — and unlike the original -48% it is like-for-like.",
    size=8.5, color=GREY, wrap=True)
cx.row_dimensions[r].height = 34
r += 1
put(cx, r, 1, "Cartrack device depreciation life", size=9)
put(cx, r, 2, "60 months", size=10, bold=True, halign="center")
put(cx, r, 5, "Disclosed accounting policy: capitalised telematics devices are depreciated "
              "straight-line over the expected average contract life of 60 months, reduced where "
              "contracts are expected to be materially shorter. CHECK YOUR OWN ASSUMED DEVICE LIFE "
              "against this — if yours is longer, that explains part of the depreciation-per-vehicle "
              "gap and is a separate assumption to defend.", size=8.5, color=GREY, wrap=True)
cx.row_dimensions[r].height = 46
r += 2

section(cx, r, "Block 4 — an internal tension to reconcile before the meeting", 5); r += 1
for i, lab in enumerate(["Metric", "McEasy 2030", "Cartrack FY2019", "", "Note"], start=1):
    put(cx, r, i, lab, bold=True, size=9, halign="center", fill=SUB_FILL, wrap=True)
r += 1
put(cx, r, 1, "Base growth in the year", size=9)
put(cx, r, 2, f'=IFERROR({VEH}/{V29}-1,"")', fmt=PCT, halign="center")
put(cx, r, 3, f'=IFERROR({K}{V19}{R["subs"]}/{K}B{R["subs"]}-1,"")', fmt=PCT, color=GREEN,
    halign="center")
r += 1
put(cx, r, 1, "Capex as a multiple of depreciation", size=9, bold=True)
put(cx, r, 2, f"=IFERROR({MC_CAPEX}/{MM}{M30}{MR['da']},\"\")", fmt="0.00x", bold=True,
    halign="center")
put(cx, r, 3, f'=IFERROR(({CTC})/(({K}{V19}{R["eb"]}-{K}{V19}{R["op"]})/{K}{V19}6*1000000),"")',
    fmt="0.00x", bold=True, color=GREEN, halign="center")
put(cx, r, 5, "A faster-growing base should carry a HIGHER capex-to-depreciation multiple, because "
              "more of the spend is on units not yet depreciated. McEasy grows the base roughly twice "
              "as fast as Cartrack did, on a LOWER multiple. Either device life is assumed longer, "
              "capex is light, or the vehicle ramp is back-loaded. An investor building a device "
              "cohort model will find this.", size=8.5, color=GREY, wrap=True)
cx.row_dimensions[r].height = 52
r += 2

section(cx, r, "Block 5 — the justifications, ranked by how well they survive scrutiny", 5); r += 1
just = [
    ("Tier 1", "You sell hardware; Cartrack subsidised it",
     "Sparepart Solutions is USD18.2m of 2030 revenue at a 21.2% gross margin — the signature of "
     "pass-through hardware at low markup. A device the customer BUYS is cost of revenue, not capex. "
     "Cartrack retained device ownership so its device spend sat in PP&E. Say the flip side yourself: "
     "this relocates cost rather than removing it, which is exactly why the blended gross margin is "
     "65.9% against 75.6% SaaS-only, and exactly why EBITDA per vehicle still matches."),
    ("Tier 1", "Cartrack's line bundles things you do not capitalise — now proven, not asserted",
     "Note 5 of the Karooooo F-1 shows FY2019 capitalised telematics device additions of "
     "ZAR353,655k against a cash-flow line of ZAR507,151k. The ZAR146,660k residual (28.9%) is "
     "capitalised sales commissions, which Note 6 carries as a separate asset and which McEasy "
     "expenses through S&M. A further ZAR74,364k of right-of-use additions in the note are IFRS 16 "
     "lease recognitions, not cash capex. Restating device-for-device closes most of the gap."),
    ("Tier 2", "Hardware price deflation, 2019 to 2026",
     "Listed device vendors are visibly losing pricing power. Queclink FY2025: revenue -7.8%, gross "
     "margin 40.1% to 36.9%, operating margin 13.4% to 5.4%. Streamax FY2025: revenue -10.8%. Both "
     "on the 'At 100M (Today)' sheet with sources."),
    ("Tier 2", "Installation labour and asset mix",
     "Indonesian install labour sits well below South African. Two-wheelers and light assets cost a "
     "fraction of a truck unit with CAN bus and video — Queclink discloses a dedicated two-wheeler "
     "terminal line precisely because Southeast Asia has that volume."),
    ("Do not use", "\"We are more capital-efficient operators\"",
     "Unquantifiable, and it invites the auditor question instead of answering it. Every argument "
     "above is either in your own accounts or in a listed company's disclosure. Stay there."),
]
for tier, head, body in just:
    put(cx, r, 1, head, size=9, bold=True,
        color="C00000" if tier == "Do not use" else BLACK)
    put(cx, r, 2, tier, size=8.5, bold=True, halign="center",
        color="C00000" if tier == "Do not use" else (GREEN if tier == "Tier 1" else GREY))
    put(cx, r, 5, body, size=8.5, color=GREY, wrap=True)
    put(cx, r, 3, None); put(cx, r, 4, None)
    cx.row_dimensions[r].height = 62
    r += 1

r += 1
notes_block(cx, r, [
    "What to say: 'our capex is USD54 per new installation against a USD26-46 bill of materials, so we "
    "are carrying headroom. Cartrack's USD177 is not comparable — it bundles capitalised commissions "
    "and stolen-vehicle-recovery infrastructure we do not operate.' That answers the question instead "
    "of defending a percentage.",
    "Four things still to collect: (1) your own year-by-year vehicle ramp, so the per-add figure stops "
    "being a constant-ARPU proxy; (2) 2026 volume quotes from Teltonika, Queclink and Howen landed in "
    "Jakarta, by device type; (3) your device-mix plan against the dashcam ceiling in Block 2; (4) your "
    "assumed device life and churned-device recovery and refurbishment rate, against Cartrack's "
    "disclosed 60 months. Item (1) on the old list — Cartrack's PP&E note split — is now obtained and "
    "sits in Block 3.",
    "Where to find it yourself: Karooooo Ltd Form F-1, SEC accession 0001104659-21-029334, document "
    "tm2034233-5_f1.htm. Note 5 is 'Property, plant and equipment' — the second and third tables are "
    "the FY2020 and FY2019 movement reconciliations, with additions by asset class. Note 6 is "
    "'Capitalized commission assets'. The F-1 is the only English SEC filing carrying audited FY2019 "
    "notes, because it was filed for the April 2021 NASDAQ listing with two comparative years.",
], ncols=5)

# ============================================================== SOURCES
sc = wb.create_sheet("Sources")
title(sc, "Sources",
      "Every figure in this workbook traces to one of these. Retrieved 10-Aug-2026.")
header_row(sc, 4, ["Company", "What it supports", "Document", "URL"],
           widths=[22, 52, 46, 96])
src = [
    ("Karooooo", "FY2026 revenue, subscription revenue, ARR, subscribers, segment split, margins, cash flow, FY2027 guidance", "Form 6-K, Q4 & FY2026 results release (13-May-2026)", "https://www.sec.gov/Archives/edgar/data/1828102/000121390026055767/ea029054101ex99-1.htm"),
    ("Karooooo", "Same, mirrored with full segment and geographic tables", "StockTitan filing summary of the FY2026 6-K", "https://www.stocktitan.net/sec-filings/KARO/6-k-karooooo-ltd-current-report-foreign-issuer-cc4ab432d528.html"),
    ("Karooooo", "FY2026 press release, dividend, adjusted FCF, regional commentary", "Karooooo, \"Delivers Accelerating Subscription Revenue Growth and Strong Cash Flow in FY 2026\"", "https://www.businesswire.com/news/home/20260513986164/en/"),
    ("Karooooo", "FY2022-FY2026 revenue, gross profit, operating income, net income, FCF history", "StockAnalysis.com financial statements, KARO", "https://stockanalysis.com/stocks/karo/financials/"),
    ("Karooooo", "FY2025 revenue by geography (South Africa, APAC/ME/USA, Europe, Africa-other)", "TradingKey revenue breakdown, NASDAQ: KARO", "https://www.tradingkey.com/markets/stocks/nasdaq-karo/revenue"),
    ("Karooooo", "FY2026 20-F (filed Jun-2026) — referenced for headcount and geography", "Form 20-F, FY2026", "https://www.sec.gov/Archives/edgar/data/0001828102/000121390026066795/ea0293645-20f_karooooo.htm"),
    ("Cartrack", "FY2019 full income statement, cash flow, subscribers and the complete geographic segment table — the USD100m-ARR vintage", "JSE SENS: Summarised Preliminary Consolidated Audited Results, year ended 28-Feb-2019", "https://www.sharenet.co.za/v3/sens_display.php?tdate=20190528070700&seq=3"),
    ("Cartrack", "FY2020 subscription revenue ZAR1,888m (97% of total) and 1,126,515 subscribers", "Cartrack 2020 Integrated Annual Report", "https://www.sharedata.co.za/data/015736/pdfs/CARTRACK_ar_feb20.pdf"),
    ("Karooooo", "Headcount indication (>5,000 group staff)", "Q4 & FY2025 earnings call transcript", "https://karooooo.com/wp-content/uploads/2025/05/Transcript-Q4-and-FY-2025.pdf"),
    ("Karooooo", "Third-party headcount estimate, 4,550 as at Mar-2026", "Revelio Labs company profile, Cartrack", "https://www.reveliolabs.com/companies/cartrack/employees"),
    ("Ituran", "FY2025 revenue USD359.0m, net income USD58.0m, EBITDA USD96.2m, operating cash flow USD88.6m, 2.63m subscribers, 74/26 services-products split, Q4 geography 55/23/22", "Q4 & FY2025 results press release (05-Mar-2026)", "https://www.prnewswire.com/news-releases/ituran-presents-fourth-quarter--full-year-2025-results-302705204.html"),
    ("Ituran", "CY2006-2007 revenue run-rate, 407k subscribers, ~800 employees — the USD100m vintage", "Form 6-K, FY2007 quarterly results", "https://www.sec.gov/Archives/edgar/data/0001337117/000117891307001064/exhibit_99-1.htm"),
    ("Powerfleet", "FY2026 revenue USD443.8m, services USD359.8m, products USD84.0m, 55.5% gross margin, USD19.6m operating income, USD(20.6)m net loss, USD97.0m adjusted EBITDA, cash flow and net debt", "Q4 & full-year FY2026 results release (15-Jun-2026)", "https://www.prnewswire.com/news-releases/powerfleet-reports-results-for-fourth-quarter-and-full-year-fiscal-2026-302799808.html"),
    ("MiX Telematics", "FY2014 revenue and subscription guidance, adjusted EBITDA, 25.3% subscriber growth — the USD100m vintage for the Powerfleet lineage", "Form 6-K, FY2014", "https://www.sec.gov/Archives/edgar/data/0001576914/000115752314002468/a50878981-ex991.htm"),
    ("Motive", "ARR USD501m, LTM revenue USD429m, 70% gross margin, -17% non-GAAP operating margin, -23% FCF margin, 4,508 employees, NDR, customer counts", "Form S-1 (filed 23-Dec-2025) and the Mostly Metrics S-1 breakdown", "https://www.sec.gov/Archives/edgar/data/1646681/000162828025058773/motive-sx1.htm"),
    ("Motive", "9M-2025 revenue USD327.3m, net loss USD138.5m, Q3 detail, IPO status", "CNBC, \"Alphabet-backed Motive files for IPO\" and Seeking Alpha S-1 analysis", "https://www.cnbc.com/2025/12/23/alphabet-backed-motive-files-for-ipo.html"),
    ("Motive", "KeepTruckin ARR history: USD150m in 2021, 55k+ customers Apr-2019, USD149m Series D", "Contrary Research company report", "https://research.contrary.com/company/motive"),
    ("Geotab", "5m+ subscriptions, 55,000+ customers, 2,900+ employees", "Geotab press release and ABI Research 2025 vendor ranking", "https://www.geotab.com/press-release/geotab-5-million-subscriptions-milestone/"),
    ("Geotab", "Revenue estimates USD412.3m (2021) to USD681m (2024) — third-party, unaudited", "Tracxn company profile", "https://tracxn.com/d/companies/geotab/"),
    ("Lytx", "1,082 employees, Permira ownership, 5.5m+ drivers, revenue range estimate only", "Lytx corporate profile and Growjo estimate", "https://growjo.com/company/Lytx"),
    ("Gurtam", "CY2025 revenue USD37.3m, 339 employees", "Latka company profile", "https://getlatka.com/companies/gurtam.com"),
    # ---- players currently at USD80-250m revenue -------------------------------
    ("Microlise", "FY2025 adjusted revenue GBP84.0m, ARR GBP59.2m, recurring GBP58.8m, adjusted EBITDA GBP8.3m at 10%, 615 employees, 1.4% churn, 417 new customers, OEM 27% of revenue", "Results for the year ended 31 December 2025 (RNS)", "https://uk.advfn.com/stock-market/london/microlise-SAAS/share-news/Microlise-Group-PLC-Results-for-the-year-ended-31/98522274"),
    ("Microlise", "FY2025 statutory income statement: gross profit GBP54.14m, operating loss GBP(2.42)m, net loss GBP(2.17)m, FCF GBP9.47m", "StockAnalysis.com financial statements, AIM: SAAS", "https://stockanalysis.com/quote/aim/SAAS/financials/"),
    ("Microlise", "Enterprise value GBP32.35m, market cap GBP45.22m, EV/Sales 0.39x, EV/EBITDA 4.29x, net cash GBP12.87m (10-Aug-2026)", "StockAnalysis.com statistics, AIM: SAAS", "https://stockanalysis.com/quote/aim/SAAS/statistics/"),
    ("Quartix", "FY2025 ARR GBP37.0m (+14%), 310,701 subscriptions, 31,040 customers, gross margin 73.2%", "FY2025 results coverage and Quartix 2025 annual report", "https://www.quartix.com/content/uploads/2026/03/Quartix-Technologies-Plc-Annual-Report-2025.pdf"),
    ("Quartix", "FY2025 revenue GBP35.71m, gross profit GBP26.13m, operating income GBP8.68m, net income GBP6.38m, FCF GBP4.47m", "StockAnalysis.com financial statements, AIM: QTX", "https://stockanalysis.com/quote/aim/QTX/financials/"),
    ("Quartix", "Enterprise value GBP104.74m, EV/Sales 2.78x, EV/EBITDA 8.05x (10-Aug-2026)", "StockAnalysis.com statistics, AIM: QTX", "https://stockanalysis.com/quote/aim/QTX/statistics/"),
    ("CalAmp", "CY2024 revenue USD197m, EBITDA USD12.7m, 2.7m subscribers", "CalAmp, \"Delivers Strong Financial Performance in 2024\"", "https://www.globenewswire.com/news-release/2025/04/24/3067737/0/en/CalAmp-Delivers-Strong-Financial-Performance-in-2024.html"),
    ("CalAmp", "Chapter 11 terms: USD230m debt eliminated, Lynrock Lake sole owner, delisted Oct-2024; Q3-FY2024 S&SS USD34.5m vs products USD19.2m; 644 employees at Feb-2023", "Go-private press release and last Form 10-K", "https://www.calamp.com/press-releases/lynrock-lake-takes-calamp-private/"),
    ("Teletrac Navman", "~USD168m revenue; Vontier divestiture to Respida Capital completed 30-Jun-2026, USD220m total transaction value, ~USD80m cash to Vontier", "Vontier press release, \"Completes Divestiture of Teletrac Navman\"", "https://investors.vontier.com/press-releases/press-releases-details/2026/Vontier-Completes-Divestiture-of-Teletrac-Navman/default.aspx"),
    ("Netradyne", "ARR USD210m (2024) from USD129.8m (2023); USD1.3bn valuation; ~1,000 employees", "Latka and Tracxn company profiles", "https://getlatka.com/companies/netradyne.com"),
    ("ORBCOMM", "TTM revenue ~USD250m (Sep-2025), 878 employees (Feb-2026); GI Partners take-private Apr-2021 at ~USD1.1bn including net debt", "GI Partners acquisition announcement and Growjo profile", "https://www.gipartners.com/news/orbcomm-announces-completion-of-acquisition-by-gi-partners"),
    ("Queclink", "FY2025 revenue CNY891.1m, gross profit CNY329.17m, operating income CNY48.51m, net income CNY74.56m, FCF CNY45.79m", "StockAnalysis.com financial statements, SHE: 300590", "https://stockanalysis.com/quote/she/300590/financials/"),
    ("Queclink", "FY2025 revenue by product line and by region (94.34% outside China)", "Futubull revenue main-composition breakdown, 300590.SZ", "https://www.futunn.com/en/stock/300590-SZ/financial/main-composition"),
    ("Streamax", "FY2025 revenue CNY2,477m, gross margin 44.84%, operating income CNY375m, net income CNY383m, FCF CNY294m", "StockAnalysis.com financial statements, SZ: 002970", "https://stockanalysis.com/quote/she/002970/financials/"),
    ("Streamax", "65% overseas revenue, 5m+ commercial vehicles, ~2,231 employees", "Streamax company profile and Frost & Sullivan 2025 recognition release", "https://www.prnewswire.com/news-releases/streamax-technology-receives-frost--sullivans-2025-global-aiot-video-hardware-customer-value-leadership-recognition-for-excellence-in-ai-driven-fleet-safety-and-operational-efficiency-302780528.html"),
    ("Fleetio", "USD450m+ Series D Mar-2025, Auto Integrate acquisition, combined valuation above USD1.5bn, 8,500+ fleets, 8m+ vehicles", "Fleetio Series D press release", "https://www.globenewswire.com/news-release/2025/03/25/3048952/0/en/Fleetio-Raises-over-450-Million-Series-D-and-Acquires-Auto-Integrate-to-Create-Customer-Centric-One-Stop-Shop-for-Fleet-Maintenance.html"),
    ("Fleetio", "~USD58m revenue estimate, 463 employees — third-party, unaudited", "Latka and Tracxn company profiles", "https://getlatka.com/companies/fleetio"),
    ("Karooooo", "Enterprise value USD1.96bn, EV/Sales 5.52x, EV/EBITDA 21.62x (10-Aug-2026)", "StockAnalysis.com statistics, KARO", "https://stockanalysis.com/stocks/karo/statistics/"),
    ("FX", "GBP/USD 2025 full-year average 1.3190; USD/CNY 2025 full-year average 7.1873", "exchange-rates.org and x-rates 2025 annual averages", "https://www.exchange-rates.org/exchange-rate-history/gbp-usd-2025"),
    ("McEasy", "Every figure on the 'McEasy Model' sheet — ARR, revenue by line, COGS, gross profit, S&M, opex by line, EBITDA, depreciation, tax & interest, net profit 2022-2030; gross fixed assets, cash and paid-up capital 2023-2030. Read programmatically from the file at build time, not transcribed.", "McEasy Pte. Ltd. consolidated model, sheets 'PnL_Consol (USD)' and 'BS_Consol (USD)'", MODEL_PATH),
    ("Cartrack", "FY2019 capitalised telematics device additions ZAR353,655k; other owned-asset additions ZAR6,836k; right-of-use additions ZAR74,364k; capitalised sales commissions cost ZAR178,330k at Feb-2019 with FY2020 additions of ZAR64,437k; device depreciation over 60 months. Note 5 'Property, plant and equipment' and Note 6 'Capitalized commission assets'.", "Karooooo Ltd Form F-1 (Mar-2021) — the only English SEC filing with audited FY2019 notes", "https://www.sec.gov/Archives/edgar/data/1828102/000110465921029334/tm2034233-5_f1.htm"),
    ("Device prices", "2026 wholesale tracker pricing used in the bill of materials on 'Capex Decomposition': Concox / Jimi IoT USD12-49 with high-volume pricing as low as USD5-12; Teltonika FMB920 about EUR35 retail. Indicative only — replace with your own landed-Jakarta quotes.", "Alibaba / Jimi IoT supplier listings and Teltonika product pages", "https://www.teltonika-gps.com/products/trackers/basic/fmb920"),
    ("Cartrack", "Stolen-vehicle-recovery infrastructure that sits inside Cartrack's PP&E and not in McEasy's model: ground AND air recovery teams across South Africa and neighbouring countries, 1,402 recoveries in January 2025, over 100,000 in a decade.", "Cartrack South Africa, stolen vehicle recovery product pages", "https://www.cartrack.co.za/platform/features/risk-management-compliance/stolen-vehicle-recovery"),
    ("McEasy", "Monthly P&L and balance sheet columns to Dec-2030, with ACTUALS through 30-Jun-2026. Used to re-base the conservative cases from Jul-2026, to derive the 48-month device depreciation cohort (implied blended life over H1-2026: 47.7 months), and to calibrate the monthly cash roll so the base case reproduces the plan exactly.", "McEasy Pte. Ltd. consolidated model, monthly columns on both sheets", MODEL_PATH),
    ("McEasy", "2030 volume assumption: ~900k-1m vehicles at USD8-9 per vehicle per month, Southeast-Asia-weighted. NOT PRESENT IN THE MODEL — supplied by Grady Kusmulyadi in conversation on 11-Aug-2026. Every per-vehicle figure in this workbook depends on it.", "Verbal, unmodelled. Flagged yellow throughout.", "internal"),
]
r = 5
for co, what, doc, url in src:
    put(sc, r, 1, co, size=9, bold=True)
    put(sc, r, 2, what, size=8.5, wrap=True, color=GREY)
    put(sc, r, 3, doc, size=8.5, wrap=True, color=GREY)
    is_link = url.startswith("http")
    c = put(sc, r, 4, url, size=8.5, wrap=True, color="0563C1" if is_link else GREY)
    if is_link:
        c.hyperlink = url
    sc.row_dimensions[r].height = 34
    r += 1
r += 1
notes_block(sc, r, [
    "Third-party estimates (Tracxn, Growjo, Latka, Revelio) are flagged L on every sheet. They are "
    "the only public data for Geotab, Lytx and Gurtam. Do not present them to a board without a "
    "second independent source.",
    "SEC EDGAR pages may need a browser user-agent header; every SEC figure above was also "
    "cross-checked against a press-release mirror.",
], ncols=4)

# ============================================================== README
rd = wb.create_sheet("README")
title(rd, "McEasy — FMS comparables benchmark", None)
rd.column_dimensions["A"].width = 4
rd.column_dimensions["B"].width = 34
rd.column_dimensions["C"].width = 118

r = 3
put(rd, r, 2, "Built for", bold=True, size=10, color="1F3864")
put(rd, r, 3, "Grady Kusmulyadi, McEasy — benchmarking a 2030 target of USD100m ARR.",
    size=10, wrap=True)
r += 1
put(rd, r, 2, "Built on", bold=True, size=10, color="1F3864")
put(rd, r, 3, "10-Aug-2026. Latest reported full year for each company.", size=10)
r += 1
put(rd, r, 2, "Peer set", bold=True, size=10, color="1F3864")
put(rd, r, 3, "Emerging-market SMB fleet operators (Karooooo/Cartrack, Ituran, Powerfleet incl. "
              "MiX Telematics history), private VC-backed FMS (Motive, Geotab, Lytx, Gurtam), and "
              "the players currently at USD80-250m of revenue (Microlise, CalAmp, Teletrac Navman, "
              "Netradyne, ORBCOMM), plus Chinese-listed hardware vendors (Queclink, Streamax) in a "
              "separate block and two below-band closest-model comps (Quartix, Fleetio).",
    size=10, wrap=True)
rd.row_dimensions[r].height = 46
r += 1
put(rd, r, 2, "Three cuts", bold=True, size=10, color="1F3864")
put(rd, r, 3, "The workbook answers the USD100m question three ways, and they disagree with each "
              "other in the most useful way possible: what USD100m ARR looked like historically "
              "(Cartrack FY2019), who is at that size right now (Microlise, Quartix and peers), and "
              "where the big players sit today (Karooooo, Motive, Powerfleet, Ituran).",
    size=10, wrap=True)
rd.row_dimensions[r].height = 46
r += 2

section_titles = [
    ("McEasy vs Benchmark", "THE ANSWER. McEasy's 2030 model against Cartrack FY2019, with the "
                            "per-vehicle unit economics that carry the argument, the one claim that "
                            "needs defending, and bull/base/bear. Start here."),
    ("Conservative Cases", "2030 revenue $10m and $20m below plan, re-based month by month from "
                           "Jul-2026 off actuals. Two drivers each, because a volume miss and a price "
                           "miss end up in different places. Includes the 2027 cash trough."),
    ("Capex Decomposition", "The one claim that needs defending, taken apart. Capex per NEW INSTALL "
                            "(USD54) against a bottom-up bill of materials (USD26-46), why Cartrack's "
                            "USD177 is not a device cost, and the justifications ranked."),
    ("Growth Reality Check", "The assumption that IS more aggressive than the benchmark — a 66% ARR "
                             "CAGR delivered while turning profitable and not raising again."),
    ("McEasy Model", "Grady's own model, read programmatically from the source file. Yellow cell = "
                     "the 2030 vehicle count, which drives every per-vehicle figure."),
    ("At 100M (Today)", "WHO IS AT THIS SIZE RIGHT NOW — revenue USD80-250m. Three labelled "
                        "blocks: subscription-FMS operators, hardware vendors, and below-band "
                        "closest models. Start here."),
    ("At 100M (Historic)", "The same metrics at the year each company crossed ~USD100m ARR. "
                           "Cartrack FY2019 is the only fully-populated row."),
    ("Current Position", "The big players at their latest reported year — the destination view. "
                         "All in USD millions."),
    ("Valuation Signals", "What the market pays for an FMS business at USD100-250m of revenue, and "
                          "what it depends on. The dispersion is roughly 20x wide."),
    ("Revenue by Country", "Geographic mix across every company that discloses one. Karooooo FY2019 "
                           "and FY2025 disclosed, FY2026 rolled forward; plus Queclink, Streamax, "
                           "Ituran, Motive and Cartrack subscribers by region."),
    ("Revenue by Product", "Product-line mix. Recurring versus hardware, Karooooo's "
                           "Delivery-as-a-Service drag, and Queclink's full device line-up."),
    ("Karooooo Path", "Cartrack FY2018 to FY2026 in full — the actual trajectory from USD90m to "
                      "USD325m ARR."),
    ("McEasy Benchmark", "Yellow cells for your ARR. Required CAGR, subscribers needed at each "
                         "ARPU, implied headcount, and a side-by-side reality check against "
                         "Microlise and Quartix."),
    ("Assumptions", "FX rates and the confidence legend. Change a rate here and every USD figure "
                    "updates."),
    ("Sources", "Every URL, with what it supports."),
]
put(rd, r, 2, "Sheets", bold=True, size=11, color="1F3864")
r += 1
for name, desc in section_titles:
    put(rd, r, 2, name, size=9.5, bold=True)
    put(rd, r, 3, desc, size=9, color=GREY, wrap=True)
    rd.row_dimensions[r].height = 26
    r += 1
r += 1

put(rd, r, 2, "The answer, in one paragraph", bold=True, size=11, color="C00000")
r += 1
put(rd, r, 3, "Cartrack at FY2019 is the benchmark, and the match is far closer than a margin "
              "comparison suggests. McEasy's 2030 plan and Cartrack's FY2019 actuals sit 1.1% apart "
              "on vehicles (950,000 vs 960,798) and 2.3% apart on EBITDA per vehicle ($56.54 vs "
              "$57.85) — at a LOWER price per vehicle ($8.94/month vs $10.72). The 7.7pp EBITDA "
              "margin gap is a denominator effect from carrying less low-margin hardware revenue, "
              "not a claim to superior operating performance. That collapses the whole investor "
              "challenge to one testable number: capex per vehicle of $20.00 against Cartrack's "
              "$38.53. Lead with EBITDA per vehicle, never with 52.7%.",
    size=9.5, wrap=True, color="C00000")
rd.row_dimensions[r].height = 92
r += 2

put(rd, r, 2, "Seven things the data says", bold=True, size=11, color="1F3864")
r += 1
takeaways = [
    "USD100m ARR is not one destination — it is at least three different businesses. At that "
    "milestone, EBITDA margin across this peer set ranged from roughly 21% (MiX Telematics) to 45% "
    "(Cartrack) to deeply negative (Motive), and time from founding ranged from 6 years to 18. "
    "Decide which one you are building before you plan the number.",
    "The company closest to USD100m ARR today is a warning, not a target. Microlise runs ARR of "
    "USD78m on USD111m of revenue with a 64% gross margin — and posts a STATUTORY OPERATING LOSS. "
    "Its 10% figure is adjusted EBITDA, down from 14%. Cartrack at a comparable ARR in FY2019 ran a "
    "45% EBITDA margin and a 21% net margin. Same scale, same industry, opposite economics.",
    "The market prices that difference brutally. Microlise carries an enterprise value of 0.39x "
    "sales — GBP32m against GBP84m of revenue. Karooooo carries 5.52x. Teletrac Navman, at USD168m "
    "of revenue, was sold by a strategic owner for about 1.3x. CalAmp reached USD200m of revenue and "
    "2.7m subscribers, then wiped out its equity in Chapter 11 on a 6.4% EBITDA margin. Reaching "
    "USD100m ARR at thin margins creates almost no enterprise value.",
    "Scale is not what creates margin. Quartix is profitable at HALF the target size — 73% gross, "
    "24% operating, 18% net, 12.5% FCF margin on USD47m of revenue. Pricing power and product mix "
    "create margin; revenue does not.",
    "Cartrack's free-cash-flow margin at USD100m ARR was about 2%, not the 15% it runs today. A 45% "
    "EBITDA margin was almost entirely consumed by capex running at 30% of revenue — devices, "
    "installation and contract assets. Hardware-subsidised FMS does not self-fund at USD100m.",
    "Geographic diversification did not drive Cartrack from USD100m to USD325m. South Africa was "
    "73.6% of revenue at the milestone and is still about 74% today; Southeast Asia is 12.6% of the "
    "subscriber base after roughly fifteen years. Depth in the home market, not breadth, was the "
    "engine. Note the mirror image: Queclink, a device vendor, earns 94% of revenue outside its home "
    "market. Exporting hardware and operating a subscription base are different businesses.",
    "Karooooo management states plainly that Southeast Asian ARPU — Indonesia, Philippines, "
    "Thailand — sits materially below South African ARPU and will dilute group ARPU as the mix "
    "shifts. For an Indonesia-based company this is the binding constraint, and it makes the "
    "subscriber count, not the ARR figure, the real target.",
]
for i, t in enumerate(takeaways, start=1):
    put(rd, r, 2, f"{i}", bold=True, size=10, halign="center", color="1F3864")
    put(rd, r, 3, t, size=9, wrap=True)
    rd.row_dimensions[r].height = 46
    r += 1
r += 1

put(rd, r, 2, "One caution on the framing", bold=True, size=11, color="C00000")
r += 1
put(rd, r, 3, "Karooooo's current position — USD325m ARR, 18% net margin, 15% FCF margin, "
              "dividend-paying — is the destination, not the path, and benchmarking against it will "
              "flatter any 2030 model. Those margins reflect twenty-two years of compounding and a "
              "dominant position in one market. But the opposite error is just as costly: "
              "benchmarking against today's USD100m players alone would suggest the milestone is "
              "barely worth reaching. The honest reading needs both. And the real conclusion is that "
              "\"USD100m ARR by 2030\" is an underspecified goal — the same ARR is worth roughly "
              "USD40m of enterprise value at Microlise economics and several hundred million at "
              "Cartrack economics. Specify the margin alongside the ARR.",
    size=9.5, wrap=True, color="C00000")
rd.row_dimensions[r].height = 92
r += 2
put(rd, r, 2, "Colour convention", bold=True, size=10, color="1F3864")
r += 1
for col, lab in [(BLUE, "Blue — a figure taken from a source, or an input you can change."),
                 (BLACK, "Black — calculated on this sheet."),
                 (GREEN, "Green — links to another sheet in this workbook."),
                 ("FFC000", "Yellow fill — fill this in with your own numbers."),
                 (GREY, "Grey / blank — not disclosed. Deliberately left empty.")]:
    put(rd, r, 3, lab, size=9, color=col if col != "FFC000" else BLACK,
        fill=YEL if col == "FFC000" else None)
    r += 1
rd.sheet_view.showGridLines = False

# order the sheets
order = ["README", "McEasy vs Benchmark", "Conservative Cases", "Capex Decomposition",
         "Growth Reality Check", "McEasy Model",
         "At 100M (Today)", "At 100M (Historic)", "Current Position", "Valuation Signals",
         "Revenue by Country", "Revenue by Product", "Karooooo Path", "McEasy Benchmark",
         "Assumptions", "Sources"]
assert sorted(order) == sorted(wb.sheetnames), (
    set(order) ^ set(wb.sheetnames))

# ---------------------------------------------------------------- guards
# A nested "=" (from embedding a ref that already carried one) makes Excel refuse to open the
# file at all — openpyxl saves it happily and every value-level check still passes. Catch it here.
problems = []
for _ws in wb.worksheets:
    for _row in _ws.iter_rows():
        for _c in _row:
            f = _c.value
            if not (isinstance(f, str) and f.startswith("=")):
                continue
            body = f[1:]
            for tok in ("(=", "*=", "/=", ",=", "+=", "-="):
                if tok in body:
                    problems.append(f"{_ws.title}!{_c.coordinate}: nested '=' ({tok}) in {f[:90]}")
            if body.count("(") != body.count(")"):
                problems.append(f"{_ws.title}!{_c.coordinate}: unbalanced parens in {f[:90]}")
if problems:
    raise SystemExit("Malformed formulas — Excel would refuse to open this file:\n  "
                     + "\n  ".join(problems))
wb._sheets = [wb[n] for n in order]
wb.active = 0

wb.save(OUT)
print("wrote", OUT)
