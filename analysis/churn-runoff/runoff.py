# -*- coding: utf-8 -*-
"""Project the run-off of the CURRENT active book from the Kaplan-Meier curve.

The v2 workbook has the survival curve (tab 03_Hazard_Survival) and the target
list has each active account's ARR and tenure, but nothing joins them. The
workbook can therefore say what happens to a hypothetical new cohort, and the
target list can say what one account is worth over 12 months, but neither
answers "of the IDR 94.99B we hold today, how much is still here in a year".
This adds that as tab 13_ARR_Runoff.

Method, per account at tenure t holding ARR a:

    projected surviving ARR at horizon h  =  a x S(t+h) / S(t)

summed over accounts, where S() is the ARR-weighted KM survival curve. That
conditional form matters: an account already 25 months in is no longer exposed
to months 0-24, so its forward risk is read from where it actually stands.

The honest limit, and the reason this tab carries an evidence-quality block:
the curve is observed only to tenure month 48 and is statistically reliable
only to month 29 (at-risk pool above 15% of month 0). 61% of active ARR is
already past month 24, so most of a 24- or 36-month projection is the tail
assumption talking, not the data.

Reads   outputs/churn-cohort-v2-2026-08-11.xlsx   (03_Hazard_Survival)
        outputs/churn-am-target-list-2026-08-11.xlsx
Writes  outputs/churn-cohort-v2-2026-08-13.xlsx   (all 13 original tabs + the new one)

The output is a COPY written through Excel COM rather than openpyxl, because
openpyxl silently drops the workbook's 5 charts on save. The 08-11 workbook is
never modified.

Run:  python analysis/churn-runoff/runoff.py
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
SRC_WB = REPO / "outputs" / "churn-cohort-v2-2026-08-11.xlsx"
TARGET_LIST = REPO / "outputs" / "churn-am-target-list-2026-08-11.xlsx"
OUT_WB = REPO / "outputs" / "churn-cohort-v2-2026-08-13.xlsx"
TAB = "13_ARR_Runoff"

RELIABLE_TO = 29          # last month with >15% of the month-0 at-risk pool
HORIZONS = (6, 12, 18, 24, 36)
BANDS = ((0, 12, "0-11 months"), (12, 18, "12-17 months"),
         (18, 24, "18-23 months"), (24, 36, "24-35 months"),
         (36, 10 ** 6, "36+ months"))

# Workbook styling, lifted from the existing tabs.
NAVY = 0x64381F        # BGR for Excel COM (1F3864)
GREY = 0x808080
WHITE = 0xFFFFFF
IDR_FMT = "#,##0"
PCT_FMT = "0.0%"


def read_curve() -> tuple[dict, dict]:
    """Survival and monthly hazard by tenure month, off tab 03_Hazard_Survival."""
    tmp = Path(REPO / "outputs" / "_runoff_curve_read.xlsx")
    shutil.copy2(SRC_WB, tmp)
    try:
        ws = openpyxl.load_workbook(tmp, data_only=True)["03_Hazard_Survival"]
        surv, haz = {}, {}
        for row in ws.iter_rows(min_row=6, values_only=True):
            month, _atrisk, _share, _churned, h, s = row[:6]
            if isinstance(month, (int, float)) and isinstance(s, (int, float)):
                surv[int(month)] = float(s)
                haz[int(month)] = float(h)
        return surv, haz
    finally:
        tmp.unlink(missing_ok=True)


def read_accounts() -> list[tuple[str, float, int]]:
    """Active accounts ex-Corin: (name, ARR, tenure months).

    Corin is dropped to match the KM basis -- core.py excludes it from both
    numerator and denominator of every rate, so projecting it off this curve
    would be applying a curve it was not part of building.
    """
    tmp = Path(REPO / "outputs" / "_runoff_targets_read.xlsx")
    shutil.copy2(TARGET_LIST, tmp)
    try:
        ws = openpyxl.load_workbook(tmp, data_only=True)["At_Risk_Active_Accounts"]
        out = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            name, arr, tenure = row[0], row[4], row[6]
            if (name and isinstance(arr, (int, float))
                    and isinstance(tenure, (int, float)) and "Corin" not in name):
                out.append((name, float(arr), int(tenure)))
        return out
    finally:
        tmp.unlink(missing_ok=True)


class Curve:
    """KM survival, extended past the observed window at a constant hazard."""

    def __init__(self, surv: dict, tail_hazard: float, assume_from: int | None = None):
        self.s = surv
        self.last = max(surv)
        self.tail = tail_hazard
        # Where the observed curve stops being used. Default: its true end.
        self.anchor = min(assume_from, self.last) if assume_from else self.last

    def __call__(self, t: int) -> float:
        if t <= self.anchor:
            return self.s[t]
        return self.s[self.anchor] * ((1.0 - self.tail) ** (t - self.anchor))


def project(accounts, curve: Curve, h: int) -> float:
    keep = 0.0
    for _name, arr, t in accounts:
        s_now = curve(t)
        keep += arr * (curve(t + h) / s_now) if s_now > 0 else 0.0
    return keep


def build(accounts, surv, haz) -> dict:
    opening = sum(a[1] for a in accounts)
    last = max(surv)

    def mean_haz(lo, hi):
        months = [m for m in range(lo, hi + 1) if m in haz]
        return sum(haz[m] for m in months) / len(months)

    h_early = mean_haz(12, 19)
    h_exit = mean_haz(20, 29)
    peak_month = max((m for m in haz if m <= RELIABLE_TO), key=lambda m: haz[m])
    h_peak = haz[peak_month]

    base = Curve(surv, h_exit)
    rows = []
    for h in HORIZONS:
        keep = project(accounts, base, h)
        lost = opening - keep
        annualised = 1.0 - (keep / opening) ** (12.0 / h)
        rows.append((h, keep, lost, lost / opening, annualised))

    scenarios = [
        (f"{h_early * 100:.2f}% — mean of months 12–19 (calmer tail)", Curve(surv, h_early)),
        (f"{h_exit * 100:.2f}% — mean of months 20–29 (BASE CASE)", base),
        (f"{h_peak * 100:.2f}% — peak observed, month {peak_month}", Curve(surv, h_peak)),
        (f"{h_exit * 100:.2f}% but applied from month 30, distrusting the thin region",
         Curve(surv, h_exit, assume_from=30)),
    ]
    sens = [(label, [(opening - project(accounts, c, h)) / opening for h in (12, 24, 36)])
            for label, c in scenarios]

    quality = []
    for h in (12, 24, 36):
        rel = sum(a[1] for a in accounts if a[2] + h <= RELIABLE_TO)
        thin = sum(a[1] for a in accounts if RELIABLE_TO < a[2] + h <= last)
        extr = sum(a[1] for a in accounts if a[2] + h > last)
        quality.append((h, rel / opening, thin / opening, extr / opening))

    bands = []
    for lo, hi, label in BANDS:
        sub = [a for a in accounts if lo <= a[2] < hi]
        if not sub:
            continue
        o = sum(a[1] for a in sub)
        lost = o - project(sub, base, 12)
        bands.append((label, len(sub), o, lost, lost / o))

    series = []
    for h in range(0, 37):
        keep = project(accounts, base, h)
        series.append((h, keep, opening - keep))

    return {
        "opening": opening, "rows": rows, "sens": sens, "quality": quality,
        "bands": bands, "series": series, "last": last,
        "h_exit": h_exit, "peak_month": peak_month,
        "km12": rows[1][2], "km12_pct": rows[1][3],
    }


def write_tab(model: dict) -> None:
    import win32com.client as com

    shutil.copy2(SRC_WB, OUT_WB)
    app = com.Dispatch("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    wb = app.Workbooks.Open(str(OUT_WB))
    try:
        for sh in wb.Sheets:
            if sh.Name == TAB:
                sh.Delete()
        ws = wb.Sheets.Add()
        ws.Name = TAB
        r = [1]

        def title(text):
            c = ws.Cells(r[0], 1)
            c.Value = text
            c.Font.Size = 13
            c.Font.Bold = True
            c.Font.Color = NAVY
            r[0] += 1

        def note(text):
            c = ws.Cells(r[0], 1)
            c.Value = text
            c.Font.Size = 9
            c.Font.Color = GREY
            r[0] += 1

        def head(labels):
            for i, lab in enumerate(labels, start=1):
                c = ws.Cells(r[0], i)
                c.Value = lab
                c.Font.Size = 10
                c.Font.Bold = True
                c.Font.Color = WHITE
                c.Interior.Color = NAVY
            r[0] += 1

        def line(values, fmts=None, bold=False):
            for i, v in enumerate(values, start=1):
                c = ws.Cells(r[0], i)
                c.Value = v
                if fmts and fmts[i - 1]:
                    c.NumberFormat = fmts[i - 1]
                if bold:
                    c.Font.Bold = True
            r[0] += 1

        def gap():
            r[0] += 1

        title("Projected run-off of the CURRENT active book — Kaplan-Meier conditional survival")
        note("Answers: of the active ARR we hold today, how much is still here at each horizon "
             "if nothing changes. This is a gross projection with no intervention, expansion or new sales.")
        note("Per account at tenure t holding ARR a:  projected surviving ARR at horizon h = "
             "a × S(t+h) / S(t), summed. S() is the ARR-weighted survival curve on tab 03_Hazard_Survival.")
        note("The conditional form is the point: an account already 25 months in is no longer exposed "
             "to months 0-24, so its forward risk is read from where it actually stands on the curve.")
        note("Corin Mulia Gemilang is excluded, matching the KM basis (see 01_Read_Me). "
             "Account tenure is account-level while the curve is line-level, so accounts whose lines "
             "started at different times carry some error.")
        note(f"The curve is OBSERVED only to tenure month {model['last']} and is statistically "
             f"reliable only to month {RELIABLE_TO}. Beyond the observed window a constant monthly "
             "hazard is assumed — see the sensitivity and evidence-quality blocks below.")
        gap()

        title("Base case — tail hazard %.2f%%/month (mean of months 20-29)" % (model["h_exit"] * 100))
        head(["Horizon", "Opening ARR", "Projected surviving ARR",
              "Projected churn (IDR)", "Projected churn", "Annualised equivalent"])
        for h, keep, lost, pct, ann in model["rows"]:
            line([f"{h} months", model["opening"], keep, lost, pct, ann],
                 [None, IDR_FMT, IDR_FMT, IDR_FMT, PCT_FMT, PCT_FMT],
                 bold=(h == 12))
        note("The 12-month row is the one to quote. Everything below it is dominated by the tail "
             "assumption rather than by observed data — see the evidence-quality block.")
        gap()

        title("Sensitivity to the tail assumption — cumulative churn as % of opening ARR")
        head(["Assumed monthly hazard beyond the observed curve", "12 months", "24 months", "36 months"])
        for label, vals in model["sens"]:
            line([label] + vals, [None, PCT_FMT, PCT_FMT, PCT_FMT],
                 bold=("BASE CASE" in label))
        note("At 36 months the spread across plausible tail assumptions is roughly 20 percentage "
             "points of the opening book. That is the honest width of the estimate, not noise.")
        note("The last row is the reassuring one: throwing away the thin months 30-48 entirely and "
             "assuming from month 30 moves the 12-month figure by about 0.2 points, because the "
             "observed thin region happens to run close to the assumed rate. The 12-month number is "
             "therefore robust to distrusting the thin data — it is the 36-month number that is not.")
        gap()

        title("How much of each horizon rests on data vs on the assumption")
        head(["Horizon", "Fully reliable (t+h ≤ %d)" % RELIABLE_TO,
              "Thin (months %d-%d)" % (RELIABLE_TO + 1, model["last"]),
              "Extrapolated (beyond month %d)" % model["last"]])
        for h, rel, thin, extr in model["quality"]:
            line([f"{h} months", rel, thin, extr], [None, PCT_FMT, PCT_FMT, PCT_FMT])
        note("Share of opening ARR by the quality of the curve region its projection lands in. "
             "This is why the 24- and 36-month rows must not be planned against: at 24 months only "
             "a twentieth of the book has a reliable answer, and at 36 months none of it does. "
             "It is consistent with 12_Data_Quality: nothing about months 30+ should be planned against.")
        gap()

        title("Where the next 12 months of loss comes from, by tenure today")
        head(["Tenure today", "Accounts", "Opening ARR", "Projected 12-month churn", "Rate"])
        for label, n, o, lost, pct in model["bands"]:
            line([label, n, o, lost, pct], [None, "#,##0", IDR_FMT, IDR_FMT, PCT_FMT])
        gap()

        title("Reconciliation — why this differs from the IDR 13.23B in the target list")
        head(["Estimator", "12-month expected loss", "As % of opening ARR"])
        line(["KM conditional survival, portfolio curve (this tab)",
              model["km12"], model["km12_pct"]], [None, IDR_FMT, PCT_FMT], bold=True)
        line(["Per-account peer-band hazard (churn-am-target-list, and the deck)",
              13.225e9, 13.225e9 / model["opening"]], [None, IDR_FMT, PCT_FMT])
        line(["Difference", model["km12"] - 13.225e9,
              (model["km12"] - 13.225e9) / model["opening"]], [None, IDR_FMT, PCT_FMT])
        note("Both are legitimate and they answer slightly different questions. This tab applies ONE "
             "portfolio curve to each account's own tenure, so it captures the tenure mix of the book "
             "but ignores segment differences. The target list applies each account's SIZE-BAND hazard, "
             "so it captures segment differences but rests on thinner per-band samples. The portfolio "
             "curve reads higher mainly because 61% of active ARR sits past month 24, where observed "
             "monthly hazard runs 1.6-1.8%.")
        note("Quote the target list's 13.23B for account prioritisation, since that is what the ranked "
             "list is built from. Quote this tab for a book-level run-off. Do not mix them in one sentence.")
        gap()

        title("Monthly run-off series, base case — chart this")
        head(["Months from today", "Projected surviving ARR", "Cumulative projected churn"])
        for h, keep, lost in model["series"]:
            line([h, keep, lost], [None, IDR_FMT, IDR_FMT])

        for col, width in (("A", 58), ("B", 20), ("C", 24), ("D", 24), ("E", 20), ("F", 20)):
            ws.Columns(col).ColumnWidth = width
        ws.Rows(1).RowHeight = 18

        # Sheets.Add ignores both Before= and After= under late binding and drops
        # the sheet before whichever sheet was active when the file was saved, so
        # the position is set afterwards with Move. Move must be called
        # POSITIONALLY -- Move(Before, After) -- because the keyword form is
        # dropped the same way. The "13_" prefix means this tab belongs last.
        ws.Move(None, wb.Sheets("12_Data_Quality"))
        wb.Save()
    finally:
        wb.Close(SaveChanges=True)
        app.Quit()


def main() -> None:
    surv, haz = read_curve()
    accounts = read_accounts()
    model = build(accounts, surv, haz)

    opening = model["opening"]
    assert abs(opening / 1e9 - 94.99) < 0.02, f"opening ARR drifted: {opening/1e9:.3f}B"
    assert len(accounts) == 2002, f"expected 2,002 active accounts, got {len(accounts)}"
    assert model["series"][0][1] == opening, "month-0 projection must equal opening ARR"
    for h, keep, lost, pct, _ann in model["rows"]:
        assert 0 < keep < opening and abs(keep + lost - opening) < 1.0, f"row {h} does not tie"

    print("opening active ARR ex-Corin  %.2fB across %d accounts" % (opening / 1e9, len(accounts)))
    print("curve observed to month %d, reliable to month %d" % (model["last"], RELIABLE_TO))
    print("base tail hazard %.2f%%/mo\n" % (model["h_exit"] * 100))
    print("  %-10s %12s %12s %8s %11s" % ("horizon", "surviving", "churned", "%", "annualised"))
    for h, keep, lost, pct, ann in model["rows"]:
        print("  %-10s %11.2fB %11.2fB %7.1f%% %10.1f%%"
              % ("%d mo" % h, keep / 1e9, lost / 1e9, 100 * pct, 100 * ann))
    print("\n  evidence quality (share of opening ARR)")
    for h, rel, thin, extr in model["quality"]:
        print("    %2d mo: reliable %5.1f%%  thin %5.1f%%  extrapolated %5.1f%%"
              % (h, 100 * rel, 100 * thin, 100 * extr))
    print("\n  reconciliation: KM run-off %.2fB vs target list 13.23B  (gap %+.2fB)"
          % (model["km12"] / 1e9, (model["km12"] - 13.225e9) / 1e9))

    write_tab(model)

    # Verify the saved file rather than trusting the COM session: the tab must be
    # last, all 13 original tabs must survive, and the workbook's 5 charts must
    # still be there (openpyxl would have silently dropped them, which is why
    # this is written through Excel).
    import zipfile
    check = openpyxl.load_workbook(OUT_WB, data_only=True)
    assert check.sheetnames[-1] == TAB, f"tab order wrong: {check.sheetnames}"
    assert len(check.sheetnames) == 14, f"expected 14 tabs, got {len(check.sheetnames)}"
    names = zipfile.ZipFile(OUT_WB).namelist()
    charts = [n for n in names if "charts/chart" in n and n.endswith(".xml")]
    assert len(charts) == 5, f"expected 5 charts to survive, found {len(charts)}"
    band_sum = sum(b[3] for b in model["bands"])
    assert abs(band_sum - model["km12"]) < 1.0, "tenure bands do not sum to the 12-month total"

    print(f"\nwrote {OUT_WB.relative_to(REPO)}")
    print(f"  {len(check.sheetnames)} tabs ({TAB} last), {len(charts)} charts preserved")


if __name__ == "__main__":
    main()
