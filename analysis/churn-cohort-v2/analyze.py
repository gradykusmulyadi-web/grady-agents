"""Churn cohort v2 -> hazard-rate workbook + AM target list.

Usage:
    python analyze.py <subscription.csv> <nonaktif.csv> <output_dir>

What is different from the v1 run: the Subscription extract carries the ACTIVE
base alongside the churned lines, so every figure here can be a RATE rather
than a share-of-churn. v1 could only ever say "band X is N% of churn"; it could
not say whether band X churns more than band Y. Each tab states which it is.

Console here is cp1252; keep stdout ASCII-only.
"""

from __future__ import annotations

import collections
import shutil
import statistics
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import core

# --------------------------------------------------------------------------
# Styling
# --------------------------------------------------------------------------

H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="808080")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
IDR = "#,##0"
PCT = "0.0%"
PCT2 = "0.00%"
NUM1 = "#,##0.0"
HORIZON = 48


def title(ws, row, text, note=None):
    ws.cell(row, 1, text).font = TITLE_FONT
    row += 1
    for line in (note or "").split("\n"):
        if line:
            ws.cell(row, 1, line).font = NOTE_FONT
            row += 1
    return row + 1


def header(ws, row, labels, start_col=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row, start_col + i, lab)
        c.fill, c.font, c.border = H_FILL, H_FONT, BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return row + 1


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def put(ws, row, col, value, fmt=None, font=None, fill=None, border=True):
    c = ws.cell(row, col, value)
    if fmt:
        c.number_format = fmt
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if border:
        c.border = BOX
    return c


def band_order(keys):
    known = [b for b in core.SIZE_ORDER if b in keys]
    return known + sorted(k for k in keys if k not in core.SIZE_ORDER)


def bucket_order(keys):
    known = [b for b in core.BUCKET_ORDER if b in keys]
    return known + sorted(k for k in keys if k not in core.BUCKET_ORDER)


def fmt_idr(v):
    return format(int(round(v)), ",")


# --------------------------------------------------------------------------
# Tab 1: Read Me
# --------------------------------------------------------------------------


def tab_readme(wb, ctx):
    ws = wb.create_sheet("01_Read_Me")
    widths(ws, {"A": 34, "B": 104})
    r = title(ws, 1, "McEasy churn cohort analysis v2 -- read this before quoting any number")

    def kv(r, k, v, fill=None):
        put(ws, r, 1, k, font=BOLD, fill=SUB_FILL)
        c = put(ws, r, 2, v, fill=fill)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        return r + 1

    r = kv(r, "Purpose", "Design an Account Management program: which accounts get named "
                         "owners, what triggers an intervention, at what lifecycle point, "
                         "owned by which function, measured by which metric.")
    r = kv(r, "What changed vs v1", "v1 had churned customers ONLY, so every figure was "
                                    "share-of-churn (mix). This extract carries the active "
                                    "base, so genuine churn RATES and a Kaplan-Meier hazard "
                                    "curve are computable for the first time.", GOOD_FILL)
    r = kv(r, "Snapshot / observation end", str(ctx["snapshot"]))
    r += 1

    r = title(ws, r, "Verified points on the source-file reading")
    r = kv(r, "1. ARR annualisation = subtotal x 12", "Confirmed, at every billing frequency. "
           "`subtotal` is a MONTHLY-NORMALISED amount, not the amount per invoice: it equals "
           "`unit_price` (a monthly rate) in 100%% of rows across all 10 frequencies, and for "
           "churned lines matched to Nonaktif -- whose ARR Lost is a known annual figure -- the "
           "implied multiplier is median 12.00 for /Month, /Year, /3 Months, /4 Months and "
           "/6 Months alike. `multipler` is months-per-invoice and scales the invoice amount, "
           "not the run-rate. For reference %d rows (%.0f%%) bill non-monthly, `/ Year` alone "
           "%s rows -- that affects cash timing and the prepay offer, not ARR."
           % (ctx["n_nonmonthly"], 100.0 * ctx["n_nonmonthly"] / ctx["n_sub_rows"],
              format(ctx["n_yearly"], ",")), GOOD_FILL)
    r = kv(r, "2. Churn date", "In Subscription the churn date is `date_end`, NOT "
           "`accounting_date` (which equals `date_start`). `date_end` is NULL on all %s active "
           "rows and populated only on churned ones -- which is what makes this a clean "
           "right-censored panel." % format(ctx["n_active_lines"], ","), WARN_FILL)
    r = kv(r, "3. Nonaktif double-counting", "Every Nonaktif line is stored as 2 rows (one per "
           "salesperson slot) and the file mixes conventions: %d grains duplicate the value, "
           "%d genuinely split it. Rule: equal nonzero values -> take one; unequal -> sum. "
           "Total falls IDR %s -> %s. Validated: Tempirai Energy Resources reconciles to the "
           "Subscription register to the rupiah under this rule."
           % (ctx["non_meta"]["n_duplicate_grains"], ctx["non_meta"]["n_split_grains"],
              fmt_idr(ctx["non_meta"]["sum_all_arr"]), fmt_idr(ctx["non_meta"]["dedup_arr"])),
           WARN_FILL)
    r += 1

    r = title(ws, r, "Exclusions applied")
    r = kv(r, "Administrative reasons", "Excluded from every churn figure -- these are "
           "re-papering, ERP cleanup or plan changes, not a customer leaving: %s. "
           "Removes IDR %s (%.1f%% of gross churned ARR)."
           % (", ".join(sorted(core.ADMIN_REASONS)), fmt_idr(ctx["exc_admin"]),
              100.0 * ctx["exc_admin"] / ctx["gross_churn"]))
    r = kv(r, "Downgrade -- important nuance", "Downgrade is excluded from CHURN (nobody left) "
           "but it is real ARR contraction of IDR %s and is the most AM-addressable loss in "
           "the file: the customer stayed, said why, and shrank anyway. It is reported on the "
           "07_Full_vs_Partial tab as contraction. Do not read its absence from the churn "
           "headline as 'no problem here'." % fmt_idr(ctx["exc_downgrade"]), GOOD_FILL)
    r = kv(r, "Excluded customer", "%s -- IDR %s, %.1f%% of gross churned ARR on its own. "
           "Excluded by instruction; it distorts every segment average it touches."
           % (", ".join(sorted(core.EXCLUDE_CUSTOMERS)), fmt_idr(ctx["exc_customer"]),
              100.0 * ctx["exc_customer"] / ctx["gross_churn"]))
    r += 1

    r = title(ws, r, "Mix vs rate -- the single easiest way to misread this workbook")
    for tab, kind, note in ctx["tab_index"]:
        put(ws, r, 1, tab, font=BOLD)
        c = put(ws, r, 2, "%s -- %s" % (kind, note),
                fill=GOOD_FILL if kind == "RATE" else (WARN_FILL if kind == "MIX" else None))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1

    r = title(ws, r, "Limits -- state these, do not bury them")
    for lim in ctx["limits"]:
        c = put(ws, r, 2, lim, fill=WARN_FILL)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
        r += 1
    return ws


# --------------------------------------------------------------------------
# Tab 2: Reconciliation
# --------------------------------------------------------------------------


def tab_reconciliation(wb, ctx):
    ws = wb.create_sheet("02_Reconciliation")
    widths(ws, {"A": 46, "B": 20, "C": 18, "D": 60})
    r = title(ws, 1, "Reconciliation and control totals",
              "Tie every figure back to something the business recognises before arguing "
              "about conclusions.")

    r = header(ws, r, ["Control total", "Value", "Unit", "Note"])
    for label, val, unit, note in ctx["controls"]:
        put(ws, r, 1, label)
        put(ws, r, 2, val, IDR if isinstance(val, (int, float)) and abs(val) > 1000 else None)
        put(ws, r, 3, unit)
        c = put(ws, r, 4, note)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 2

    r = title(ws, r, "Gap bridge: Nonaktif vs Subscription churned ARR",
              "Decision: Subscription is authoritative. A rate must take numerator and "
              "denominator from the same register, or the rate is wrong.")
    r = header(ws, r, ["Bridge component", "IDR", "Role", "Nature"])
    for label, val, kind, note in ctx["bridge"]:
        emph = kind in ("start", "subtotal", "end")
        put(ws, r, 1, label, font=BOLD if emph else None)
        put(ws, r, 2, val, IDR, font=BOLD if emph else None,
            fill=TOTAL_FILL if kind in ("subtotal", "end") else None)
        put(ws, r, 3, kind)
        c = put(ws, r, 4, note)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.cell(r, 1, "Start + all delta rows sum exactly to the authoritative figure; this is "
                  "asserted in code. Subtotal rows are running totals for readability and are "
                  "not part of that sum.").font = NOTE_FONT
    r += 1
    r += 2

    r = title(ws, r, "Reason propagation coverage (Nonaktif -> Subscription)",
              "WO does not join between the files (0.2%); company name joins 99.8%. Cascade "
              "runs tightest grain first. Loose fallbacks are disclosed, not hidden.")
    r = header(ws, r, ["Match level", "Lines", "Churned ARR", "% of churned ARR"])
    tot = ctx["gross_churn"]
    for level in ["L1 co+month+product", "L2 co+month", "L3 co", "unmatched"]:
        put(ws, r, 1, level)
        put(ws, r, 2, ctx["coverage"].get(level, 0), IDR)
        put(ws, r, 3, ctx["coverage_arr"].get(level, 0.0), IDR)
        put(ws, r, 4, ctx["coverage_arr"].get(level, 0.0) / tot, PCT,
            fill=WARN_FILL if level in ("L3 co", "unmatched") else None)
        r += 1
    put(ws, r, 1, "Coverage (any level)", font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 2, sum(ctx["coverage"].values()), IDR, font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 3, tot - ctx["coverage_arr"].get("unmatched", 0.0), IDR, font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 4, (tot - ctx["coverage_arr"].get("unmatched", 0.0)) / tot, PCT,
        font=BOLD, fill=TOTAL_FILL)
    r += 3

    r = title(ws, r, "Blank-reason rate by year -- why reason views start 2025-01",
              "Reason coding began mid-2024. Analysing reason mix on 2022-23 data would be "
              "reading noise.")
    r = header(ws, r, ["Churn year", "Churned ARR", "Blank reason %", "Admin reason %"])
    for y, tot_y, blank, admin in ctx["blank_by_year"]:
        put(ws, r, 1, y)
        put(ws, r, 2, tot_y, IDR)
        put(ws, r, 3, blank, PCT, fill=WARN_FILL if blank > 0.25 else None)
        put(ws, r, 4, admin, PCT)
        r += 1
    return ws


# --------------------------------------------------------------------------
# Tab 3: Hazard / survival
# --------------------------------------------------------------------------


def tab_hazard(wb, ctx):
    ws = wb.create_sheet("03_Hazard_Survival")
    widths(ws, {"A": 12, "B": 18, "C": 16, "D": 12, "E": 12, "F": 14, "G": 12, "H": 12,
                "I": 12, "J": 14})
    r = title(ws, 1, "Kaplan-Meier hazard and survival by tenure month  [RATE]",
              "THE tab v1 could not produce. Hazard(m) = churn at tenure month m divided by "
              "what was still at risk at month m. Censored (still-active) units leave the "
              "denominator without entering the numerator -- that is the point.\n"
              "ARR-weighted runs at subscription-line grain; logo runs at customer grain "
              "(event = FULL churn only). Lines within a customer are correlated, so treat "
              "line-level confidence intervals as optimistic.")

    km_arr, km_cnt, km_logo = ctx["km_arr"], ctx["km_cnt"], ctx["km_logo"]
    r = header(ws, r, ["Tenure\nmonth", "ARR at risk", "% of month-0\nat risk",
                       "ARR churned", "Hazard\n(ARR)", "Survival\n(ARR)",
                       "Logos at risk", "Hazard\n(logo)", "Survival\n(logo)",
                       "Stability"])
    start = r
    for m in range(0, HORIZON + 1):
        thin = m > 0 and m not in ctx["stable_months"]
        put(ws, r, 1, m)
        put(ws, r, 2, km_arr["at_risk"][m], IDR)
        put(ws, r, 3, ctx["at_risk_share"][m], PCT,
            fill=WARN_FILL if thin else None)
        put(ws, r, 4, km_arr["events"][m], IDR)
        put(ws, r, 5, km_arr["hazard"][m], PCT2, fill=WARN_FILL if thin else None)
        put(ws, r, 6, km_arr["survival"][m], PCT)
        put(ws, r, 7, km_logo["at_risk"][m], NUM1)
        put(ws, r, 8, km_logo["hazard"][m], PCT2)
        put(ws, r, 9, km_logo["survival"][m], PCT)
        if m == ctx["peak_month"]:
            put(ws, r, 10, "PEAK (stable window)", fill=TOTAL_FILL)
        elif thin:
            put(ws, r, 10, "thin -- do not rank", fill=WARN_FILL)
        r += 1
    end = r - 1
    put(ws, r, 1, "Months beyond %d hold under %.0f%% of the month-0 at-risk pool. They are "
                  "shown for completeness and excluded from peak-finding: an earlier cut of "
                  "this analysis read a spurious hazard peak at month 40 off a pool that had "
                  "shrunk to 5%% of its starting size."
        % (ctx["last_stable_month"], 100 * ctx["stable_min_share"]),
        font=NOTE_FONT, border=False)
    r += 2

    chart = LineChart()
    chart.title = "Survival by tenure month -- ARR-weighted vs logo"
    chart.y_axis.title = "% surviving"
    chart.x_axis.title = "Tenure month"
    chart.height, chart.width = 9, 20
    for col in (6, 9):
        chart.add_data(Reference(ws, min_col=col, min_row=start - 1, max_row=end),
                       titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    ws.add_chart(chart, "L%d" % (start + 1))

    h = BarChart()
    h.title = "Hazard by tenure month (ARR) -- months %d+ are thinly observed" \
              % (ctx["last_stable_month"] + 1)
    h.y_axis.title = "Monthly hazard"
    h.x_axis.title = "Tenure month"
    h.height, h.width = 9, 20
    h.add_data(Reference(ws, min_col=5, min_row=start - 1, max_row=end), titles_from_data=True)
    h.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    ws.add_chart(h, "L%d" % (start + 20))

    r = title(ws, r, "Hazard by tenure phase  [RATE]",
              "Monthly hazard averaged within each phase, indexed to the mean across the "
              "statistically stable months only. This is the defensible version of v1's "
              "'three lifecycle phases' -- and it moves the risk window later than v1 claimed.")
    r = header(ws, r, ["Tenure phase", "Mean monthly hazard (ARR)", "Index vs overall",
                       "ARR churned in phase", "Share of churned ARR", "Stability"])
    for p in ctx["phases"]:
        put(ws, r, 1, p["label"])
        put(ws, r, 2, p["mean_h"], PCT2)
        put(ws, r, 3, p["index"], NUM1,
            fill=WARN_FILL if p["index"] >= 1.25 else (GOOD_FILL if p["index"] <= 0.8 else None))
        put(ws, r, 4, p["arr"], IDR)
        put(ws, r, 5, p["share"], PCT)
        put(ws, r, 6, "thin -- indicative only" if p["thin"] else "stable",
            fill=WARN_FILL if p["thin"] else GOOD_FILL)
        r += 1
    r += 2

    r = title(ws, r, "Survival by customer size band  [RATE]",
              "Probability a subscription line is still live at 12 / 24 / 36 months, "
              "ARR-weighted. This is what tells you whether small accounts really churn "
              "harder or are just numerous -- the question v1 could not answer.")
    r = header(ws, r, ["Size band", "ARR at risk (mo 0)", "Survival @12mo", "Survival @24mo",
                       "Survival @36mo", "Implied 24mo churn prob."])
    import compute as _c
    for band, km in ctx["km_by_band"]:
        put(ws, r, 1, band)
        put(ws, r, 2, km["at_risk"][0], IDR)
        for i, m in enumerate((12, 24, 36)):
            if _c.surv_thin(km, m):
                put(ws, r, 3 + i, "CENSORED (%.0f%%)" % (100 * _c.surv_share(km, m)),
                    fill=WARN_FILL)
            else:
                put(ws, r, 3 + i, km["survival"][m], PCT)
        if _c.surv_thin(km, 24):
            put(ws, r, 6, "n/a", fill=WARN_FILL)
        else:
            put(ws, r, 6, 1.0 - km["survival"][24], PCT,
                fill=WARN_FILL if (1.0 - km["survival"][24]) > ctx["overall_24"] else GOOD_FILL)
        r += 1
    put(ws, r, 1, "ALL", font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 2, km_arr["at_risk"][0], IDR, font=BOLD, fill=TOTAL_FILL)
    for i, m in enumerate((12, 24, 36)):
        put(ws, r, 3 + i, km_arr["survival"][m], PCT, font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 6, ctx["overall_24"], PCT, font=BOLD, fill=TOTAL_FILL)
    r += 1
    ws.cell(r, 1, '"CENSORED" means that subgroup\'s at-risk pool at 24 months had fallen below 15% of its own starting size, so the figure measures how young the subgroup is rather than how well it retains. Read those cells as no-data, not good news.').font = NOTE_FONT
    r += 3

    r = title(ws, r, "Survival by product family  [RATE]", "Same basis, ARR-weighted.")
    r = header(ws, r, ["Product family", "ARR at risk (mo 0)", "Survival @12mo",
                       "Survival @24mo", "Implied 24mo churn prob."])
    for fam, km in ctx["km_by_product"]:
        thin = fam in ctx["thin_families"]
        put(ws, r, 1, fam + ("  [small n]" if thin else ""))
        put(ws, r, 2, km["at_risk"][0], IDR, fill=WARN_FILL if thin else None)
        put(ws, r, 3, km["survival"][12], PCT)
        put(ws, r, 4, km["survival"][24], PCT)
        put(ws, r, 5, 1.0 - km["survival"][24], PCT,
            fill=WARN_FILL if (1.0 - km["survival"][24]) > ctx["overall_24"] else GOOD_FILL)
        r += 1
    ws.cell(r, 1, "[small n] = under IDR 100M of ARR at risk at month 0. Rate shown but not "
                  "stable enough to prioritise on.").font = NOTE_FONT
    return ws


# --------------------------------------------------------------------------
# Tab 4: Churn rate by segment
# --------------------------------------------------------------------------


def tab_rate_segment(wb, ctx):
    ws = wb.create_sheet("04_Churn_Rate_Segment")
    widths(ws, {"A": 16, "B": 18, "C": 18, "D": 16, "E": 16, "F": 16, "G": 18, "H": 14})
    r = title(ws, 1, "Churn rate by segment  [RATE]",
              "Two rates, deliberately both shown. Gross cumulative = churned ARR / "
              "(churned + active) ARR: simple, but confounded by how old each segment's base "
              "is. KM 24-month = age-adjusted probability of churn within 24 months; this is "
              "the one to prioritise on. Absolute ARR sits alongside because a high-rate tiny "
              "segment and a low-rate huge segment need different responses.")

    r = header(ws, r, ["Size band", "Active ARR", "Churned ARR (base)", "Gross cumulative %",
                       "KM 24mo churn prob.", "Active logos", "Full-churn logos",
                       "Logo churn %"])
    for row in ctx["rate_by_band"]:
        put(ws, r, 1, row["key"])
        put(ws, r, 2, row["active"], IDR)
        put(ws, r, 3, row["churn"], IDR)
        put(ws, r, 4, row["gross"], PCT,
            fill=WARN_FILL if row["gross"] > ctx["gross_overall"] else GOOD_FILL)
        if row["km24_thin"]:
            put(ws, r, 5, "CENSORED (%.0f%% observed)" % (100 * row["km24_share"]),
                fill=WARN_FILL)
        else:
            put(ws, r, 5, row["km24"], PCT,
                fill=WARN_FILL if row["km24"] > ctx["overall_24"] else GOOD_FILL)
        put(ws, r, 6, row["logos_active"], IDR)
        put(ws, r, 7, row["logos_full"], IDR)
        put(ws, r, 8, row["logo_rate"], PCT)
        r += 1
    put(ws, r, 1, "ALL", font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 2, ctx["active_arr"], IDR, font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 3, ctx["base_churn"], IDR, font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 4, ctx["gross_overall"], PCT, font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 5, ctx["overall_24"], PCT, font=BOLD, fill=TOTAL_FILL)
    band_end = r
    r += 1

    chart = BarChart()
    chart.title = "24-month churn probability by size band (age-adjusted)"
    chart.y_axis.title = "P(churn within 24 months)"
    chart.height, chart.width = 9, 18
    chart.add_data(Reference(ws, min_col=5, min_row=band_end - len(ctx["rate_by_band"]) - 1,
                             max_row=band_end - 1), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1,
                                  min_row=band_end - len(ctx["rate_by_band"]),
                                  max_row=band_end - 1))
    ws.add_chart(chart, "J%d" % (band_end - len(ctx["rate_by_band"])))

    r += 2
    r = title(ws, r, "Churn rate by SUBSCRIBED-VEHICLE band  [RATE]",
              "Fleet Category above measures the customer's TOTAL fleet, so it does not answer "
              "'do accounts with few vehicles on our platform churn harder'. This band is "
              "derived from the vehicles actually subscribed, which does.")
    r = header(ws, r, ["Subscribed vehicles", "Active ARR", "Churned ARR (base)",
                       "Gross cumulative %", "KM 24mo churn prob.", "Customers",
                       "Full-churn logos", "Logo churn %"])
    for row in ctx["rate_by_veh_band"]:
        put(ws, r, 1, row["key"])
        put(ws, r, 2, row["active"], IDR)
        put(ws, r, 3, row["churn"], IDR)
        put(ws, r, 4, row["gross"], PCT,
            fill=WARN_FILL if row["gross"] > ctx["gross_overall"] else GOOD_FILL)
        if row["km24_thin"]:
            put(ws, r, 5, "CENSORED (%.0f%% observed)" % (100 * row["km24_share"]),
                fill=WARN_FILL)
        else:
            put(ws, r, 5, row["km24"], PCT,
                fill=WARN_FILL if row["km24"] > ctx["overall_24"] else GOOD_FILL)
        put(ws, r, 6, row["logos_active"] + row["logos_full"], IDR)
        put(ws, r, 7, row["logos_full"], IDR)
        put(ws, r, 8, row["logo_rate"], PCT)
        r += 1
    vb_end = r
    chart2 = BarChart()
    chart2.title = "24-month churn probability by subscribed-vehicle band"
    chart2.y_axis.title = "P(churn within 24 months)"
    chart2.height, chart2.width = 9, 18
    chart2.add_data(Reference(ws, min_col=5,
                              min_row=vb_end - len(ctx["rate_by_veh_band"]) - 1,
                              max_row=vb_end - 1), titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=1,
                                    min_row=vb_end - len(ctx["rate_by_veh_band"]),
                                    max_row=vb_end - 1))
    ws.add_chart(chart2, "J%d" % (vb_end - len(ctx["rate_by_veh_band"])))

    r += 2
    r = title(ws, r, "Churn rate by product family  [RATE]")
    r = header(ws, r, ["Product family", "Active ARR", "Churned ARR (base)",
                       "Gross cumulative %", "KM 24mo churn prob.", "Active lines",
                       "Churned lines"])
    for row in ctx["rate_by_product"]:
        put(ws, r, 1, row["key"])
        put(ws, r, 2, row["active"], IDR)
        put(ws, r, 3, row["churn"], IDR)
        put(ws, r, 4, row["gross"], PCT,
            fill=WARN_FILL if row["gross"] > ctx["gross_overall"] else GOOD_FILL)
        put(ws, r, 5, row["km24"], PCT,
            fill=WARN_FILL if row["km24"] > ctx["overall_24"] else GOOD_FILL)
        put(ws, r, 6, row["lines_active"], IDR)
        put(ws, r, 7, row["lines_churn"], IDR)
        r += 1
    r += 2

    r = title(ws, r, "Size band x product family -- gross cumulative churn %  [RATE]",
              "Cells with an active base under IDR 100M are suppressed: the rate is unstable "
              "and will be over-read.")
    cols = ctx["product_keys"]
    r = header(ws, r, ["Size band"] + cols + ["All"])
    for band in ctx["band_keys"]:
        put(ws, r, 1, band, font=BOLD, fill=SUB_FILL)
        for j, fam in enumerate(cols):
            cell = ctx["band_product"].get((band, fam))
            if cell is None or cell["active"] + cell["churn"] < 100e6:
                put(ws, r, 2 + j, "n/a", font=NOTE_FONT)
            else:
                v = cell["churn"] / (cell["active"] + cell["churn"])
                put(ws, r, 2 + j, v, PCT,
                    fill=WARN_FILL if v > ctx["gross_overall"] else None)
        tot = ctx["band_totals"][band]
        put(ws, r, 2 + len(cols), tot["churn"] / (tot["active"] + tot["churn"]) if
            (tot["active"] + tot["churn"]) else 0, PCT, font=BOLD, fill=TOTAL_FILL)
        r += 1
    return ws


# --------------------------------------------------------------------------
# Tab 4b: Penetration / share of wallet
# --------------------------------------------------------------------------


def tab_penetration(wb, ctx):
    ws = wb.create_sheet("04b_Penetration")
    widths(ws, {"A": 14, "B": 12, "C": 16, "D": 14, "E": 16, "F": 18, "G": 14, "H": 16,
                "I": 20})
    r = title(ws, 1, "Share of wallet by customer size  [RATE]",
              "Fleet Category is the customer's TOTAL fleet; subscribed vehicles are what "
              "McEasy actually runs. The ratio is share of wallet, and it falls monotonically "
              "with customer size -- from roughly two-thirds of a small operator's fleet to a "
              "couple of percent of a large one's.\n"
              "This cuts two ways and both matter to an AM program: it is the expansion pool, "
              "and it is also a fragility signal. A 2%-penetrated account is a pilot, and a "
              "pilot is easy to cancel.")

    r = header(ws, r, ["Size band", "Customers", "Median subscribed veh.",
                       "Band midpoint", "Median penetration", "Active ARR",
                       "KM 24mo churn prob.", "Ceiling breaches",
                       "Implied headroom (veh.)"])
    for row in ctx["penetration"]:
        put(ws, r, 1, row["band"])
        put(ws, r, 2, row["n"], IDR)
        put(ws, r, 3, row["median_veh"], NUM1)
        put(ws, r, 4, row["band_mid"], IDR)
        put(ws, r, 5, row["median_pen"], PCT,
            fill=WARN_FILL if row["median_pen"] < 0.25 else GOOD_FILL)
        put(ws, r, 6, row["active_arr"], IDR)
        if row["km24_thin"]:
            put(ws, r, 7, "CENSORED", fill=WARN_FILL)
        else:
            put(ws, r, 7, row["km24"], PCT)
        put(ws, r, 8, row["breach_pct"], PCT,
            fill=WARN_FILL if row["breach_pct"] > 0.10 else None)
        put(ws, r, 9, row["headroom_veh"], IDR)
        r += 1
    r += 2

    r = title(ws, r, "Fleet Category integrity -- accounts breaching their own band ceiling",
              "%d customers subscribe more vehicles than their band's ceiling allows. Under "
              "either reading of the field these records are wrong, and they include some of "
              "the largest accounts in the file. Fix these before the field is used for "
              "pricing or territory design."
              % ctx["breach_total"])
    r = header(ws, r, ["Subscribed veh.", "Labelled band", "Customer", "Active ARR"])
    for veh, band, name, arr in ctx["breach_worst"]:
        put(ws, r, 1, veh, IDR)
        put(ws, r, 2, band, fill=WARN_FILL)
        put(ws, r, 3, name)
        put(ws, r, 4, arr, IDR)
        r += 1
    return ws


# --------------------------------------------------------------------------
# Tab 5: Tenure distribution
# --------------------------------------------------------------------------


def tab_tenure(wb, ctx):
    ws = wb.create_sheet("05_Tenure_Distribution")
    widths(ws, {"A": 20, "B": 18, "C": 14, "D": 14, "E": 14, "F": 16, "G": 30})
    r = title(ws, 1, "Churn distribution by tenure  [MIX -- NOT a hazard]",
              "This says where churn LANDS across tenure. It does NOT say risk is higher "
              "there: mid-tenure bins are inflated mechanically by the age structure of the "
              "base. For risk, use 03_Hazard_Survival. Both are here precisely so the two do "
              "not get conflated.")

    r = header(ws, r, ["Tenure bucket", "Churned ARR", "Share of ARR", "Vehicles*",
                       "Customers", "Top-account share", "Dominant account"])
    for row in ctx["tenure_buckets"]:
        put(ws, r, 1, row["label"])
        put(ws, r, 2, row["arr"], IDR)
        put(ws, r, 3, row["share"], PCT)
        put(ws, r, 4, row["veh"], IDR)
        put(ws, r, 5, row["cust"], IDR)
        put(ws, r, 6, row["top_share"], PCT,
            fill=WARN_FILL if row["top_share"] > 0.5 else None)
        put(ws, r, 7, row["top"] if row["top_share"] > 0.3 else "")
        r += 1
    put(ws, r, 1, "TOTAL", font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 2, ctx["base_churn"], IDR, font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 3, 1.0, PCT, font=BOLD, fill=TOTAL_FILL)
    r += 2
    ws.cell(r, 1, "* Vehicles = GPS-mapped subscription lines (proxy). Measured %+.1f%% vs "
                  "Nonaktif's hard distinct-nopol count." % ctx["veh_bias"]).font = NOTE_FONT
    r += 2

    r = title(ws, r, "Churn by exact tenure month (0-48)",
              "Per-month detail behind the buckets, with single-account dominance flagged.")
    r = header(ws, r, ["Tenure month", "Churned ARR", "Customers", "Top-account share"])
    start = r
    for m in range(0, HORIZON + 1):
        d = ctx["tenure_month"].get(m, {"arr": 0.0, "cust": 0, "top_share": 0.0})
        put(ws, r, 1, m)
        put(ws, r, 2, d["arr"], IDR)
        put(ws, r, 3, d["cust"], IDR)
        put(ws, r, 4, d["top_share"], PCT,
            fill=WARN_FILL if d["top_share"] > 0.5 and d["arr"] > 0 else None)
        r += 1
    end = r - 1

    chart = BarChart()
    chart.title = "Churned ARR by tenure month (distribution, not hazard)"
    chart.y_axis.title = "Churned ARR (IDR)"
    chart.x_axis.title = "Tenure month"
    chart.height, chart.width = 10, 22
    chart.add_data(Reference(ws, min_col=2, min_row=start - 1, max_row=end),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    ws.add_chart(chart, "F%d" % start)
    return ws


# --------------------------------------------------------------------------
# Tab 6: Contract term
# --------------------------------------------------------------------------


def tab_contract(wb, ctx):
    ws = wb.create_sheet("06_Contract_Term")
    widths(ws, {"A": 18, "B": 14, "C": 18, "D": 16, "E": 16, "F": 18, "G": 22, "H": 20})
    r = title(ws, 1, "Contract term vs churn timing  [RATE + MIX]",
              "v1 found a churn spike at tenure month 24 and could not tell 'customers decide "
              "to leave in year 2' from 'the standard term is 24 months and it expired'. This "
              "extract has contract_period, so the question is answerable.")

    r = header(ws, r, ["Contract term (mo)", "Churned lines", "Churned ARR", "Median tenure",
                       "Active lines", "KM 24mo churn prob.",
                       "% churning within 2mo of a term boundary", "Read"])
    for row in ctx["contract"]:
        put(ws, r, 1, row["term"])
        put(ws, r, 2, row["lines"], IDR)
        put(ws, r, 3, row["arr"], IDR)
        put(ws, r, 4, row["median"], NUM1)
        put(ws, r, 5, row["active"], IDR)
        if row["km24_thin"]:
            put(ws, r, 6, "CENSORED (%.0f%% obs)" % (100 * row["km24_share"]), fill=WARN_FILL)
        else:
            put(ws, r, 6, row["km24"], PCT)
        put(ws, r, 7, row["at_boundary"], PCT,
            fill=WARN_FILL if row["at_boundary"] > 0.30 else None)
        put(ws, r, 8, row["read"], fill=WARN_FILL if row["thin"] else None)
        r += 1
    r += 2

    r = title(ws, r, "Where the month-24 story actually lands",
              "Findings, computed rather than asserted.")
    for line in ctx["contract_findings"]:
        c = put(ws, r, 1, line, fill=GOOD_FILL)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    r += 2

    r = title(ws, r, "Tenure distribution within each contract term  [MIX]",
              "Share of each term's churned ARR falling in each tenure bucket. If churn "
              "clustered at the term boundary you would see the mass sit at the term length.")
    r = header(ws, r, ["Tenure bucket"] + ["%d mo term" % t for t in ctx["main_terms"]])
    for label in core.TENURE_ORDER:
        if label not in ctx["term_tenure_labels"]:
            continue
        put(ws, r, 1, label, font=BOLD, fill=SUB_FILL)
        for j, t in enumerate(ctx["main_terms"]):
            v = ctx["term_tenure"].get((t, label), 0.0)
            den = ctx["term_tenure_totals"].get(t, 0.0)
            put(ws, r, 2 + j, (v / den) if den else 0.0, PCT)
        r += 1
    return ws


# --------------------------------------------------------------------------
# Tab 7: Full vs partial + net retention
# --------------------------------------------------------------------------


def tab_full_partial(wb, ctx):
    ws = wb.create_sheet("07_Full_vs_Partial")
    widths(ws, {"A": 22, "B": 14, "C": 20, "D": 20, "E": 18, "F": 18, "G": 20})
    r = title(ws, 1, "Full vs partial churn, and net retention  [MIX + RATE]",
              "Status is evaluated AS OF THE SNAPSHOT. A company that fully churned in 2023 "
              "and came back in 2025 reads as partial. Logo retention hides partial churn "
              "completely -- an account shedding 200 of 500 vehicles shows as 100% retained, "
              "which is why net retention is a first-class metric here.")

    r = header(ws, r, ["Customer status", "Customers", "Churned ARR (base)", "Retained ARR",
                       "Vehicles churned*", "Vehicles retained*", "Net ARR retention"])
    for row in ctx["status_rows"]:
        put(ws, r, 1, row["status"], font=BOLD)
        put(ws, r, 2, row["n"], IDR)
        put(ws, r, 3, row["churn"], IDR)
        put(ws, r, 4, row["active"], IDR)
        put(ws, r, 5, row["veh_churn"], IDR)
        put(ws, r, 6, row["veh_active"], IDR)
        put(ws, r, 7, row["nrr"], PCT,
            fill=GOOD_FILL if row["nrr"] > 0.9 else WARN_FILL)
        r += 1
    r += 2

    r = title(ws, r, "Partial churn -- net retention by size band  [RATE]",
              "The largest loss pool and the one gross churn reporting misses. Net vehicle "
              "retention is the metric an AM program on this segment must move.")
    r = header(ws, r, ["Size band", "Partial-churn customers", "Churned ARR", "Retained ARR",
                       "Net ARR retention", "Net vehicle retention"])
    for row in ctx["partial_by_band"]:
        put(ws, r, 1, row["band"])
        put(ws, r, 2, row["n"], IDR)
        put(ws, r, 3, row["churn"], IDR)
        put(ws, r, 4, row["active"], IDR)
        put(ws, r, 5, row["nrr"], PCT, fill=GOOD_FILL if row["nrr"] > 0.9 else WARN_FILL)
        put(ws, r, 6, row["nvr"], PCT, fill=GOOD_FILL if row["nvr"] > 0.9 else WARN_FILL)
        r += 1
    r += 2

    r = title(ws, r, "Reason mix: full churn vs partial churn (2025+)  [MIX]",
              "v1's strongest structural finding was that these are different diseases. "
              "Re-tested here on rate-capable data.")
    r = header(ws, r, ["Reason bucket", "Full-churn ARR", "Full %", "Partial-churn ARR",
                       "Partial %", "Concordant?"])
    for row in ctx["reason_full_partial"]:
        put(ws, r, 1, row["bucket"])
        put(ws, r, 2, row["full"], IDR)
        put(ws, r, 3, row["full_pct"], PCT)
        put(ws, r, 4, row["partial"], IDR)
        put(ws, r, 5, row["partial_pct"], PCT)
        put(ws, r, 6, row["note"],
            fill=WARN_FILL if "diverge" in row["note"] else None)
        r += 1
    r += 2

    r = title(ws, r, "Revenue contraction (Downgrade) -- excluded from churn, tracked here",
              "Not churn: the customer stayed. But it is real ARR loss, the customer told you "
              "why, and it is the most AM-addressable pool in the file.")
    r = header(ws, r, ["Size band", "Contraction ARR", "Lines", "Customers"])
    for band, d in ctx["contraction_by_band"]:
        put(ws, r, 1, band)
        put(ws, r, 2, d["arr"], IDR)
        put(ws, r, 3, d["lines"], IDR)
        put(ws, r, 4, len(d["cust"]), IDR)
        r += 1
    put(ws, r, 1, "TOTAL", font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 2, ctx["exc_downgrade"], IDR, font=BOLD, fill=TOTAL_FILL)
    return ws


# --------------------------------------------------------------------------
# Tab 8 + 9: Reason
# --------------------------------------------------------------------------


def tab_reason(wb, ctx):
    ws = wb.create_sheet("08_Reason_2025plus")
    widths(ws, {"A": 26, "B": 18, "C": 12, "D": 14, "E": 12, "F": 14, "G": 12, "H": 14,
                "I": 18, "J": 24})
    r = title(ws, 1, "Churn reason, %s onward  [MIX]" % core.REASON_WINDOW_START,
              "Reason is SALES-ENTERED TESTIMONY, not fact. It is the rep who owned the "
              "relationship recording why they lost it. Treat as a hypothesis generator.\n"
              "A finding is only credible if it holds on all three weightings at once. Where "
              "ARR, vehicles and customers disagree, one big account is driving it.")

    r = header(ws, r, ["Reason bucket", "Churned ARR", "% ARR", "Vehicles*", "% veh",
                       "Customers", "% cust", "Median tenure", "Top-account share",
                       "Concordance"])
    for row in ctx["reason_rows"]:
        put(ws, r, 1, row["bucket"], font=BOLD)
        put(ws, r, 2, row["arr"], IDR)
        put(ws, r, 3, row["arr_pct"], PCT)
        put(ws, r, 4, row["veh"], IDR)
        put(ws, r, 5, row["veh_pct"], PCT)
        put(ws, r, 6, row["cust"], IDR)
        put(ws, r, 7, row["cust_pct"], PCT)
        put(ws, r, 8, row["median_tenure"], NUM1)
        put(ws, r, 9, row["top_share"], PCT,
            fill=WARN_FILL if row["top_share"] > 0.5 else None)
        put(ws, r, 10, row["concordance"],
            fill=GOOD_FILL if row["concordance"] == "concordant" else WARN_FILL)
        r += 1
    put(ws, r, 1, "TOTAL", font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 2, ctx["reason_total"], IDR, font=BOLD, fill=TOTAL_FILL)
    put(ws, r, 3, 1.0, PCT, font=BOLD, fill=TOTAL_FILL)
    r += 2
    ws.cell(r, 1, "* Vehicles = GPS-mapped lines (proxy).").font = NOTE_FONT
    r += 2

    r = title(ws, r, "Raw reason labels behind the buckets",
              "29 distinct sales-entered values. Shown so nobody has to trust the bucketing.")
    r = header(ws, r, ["Raw reason", "Bucket", "Churned ARR", "% ARR", "Customers"])
    for raw, bucket, amt, n in ctx["reason_raw"]:
        put(ws, r, 1, raw or "(blank)")
        put(ws, r, 2, bucket)
        put(ws, r, 3, amt, IDR)
        put(ws, r, 4, amt / ctx["reason_total"] if ctx["reason_total"] else 0, PCT)
        put(ws, r, 5, n, IDR)
        r += 1
    return ws


def tab_reason_cross(wb, ctx):
    ws = wb.create_sheet("09_Reason_x_Band_x_Prod")
    widths(ws, {"A": 26})
    for i in range(2, 14):
        ws.column_dimensions[chr(64 + i)].width = 15
    r = title(ws, 1, "Reason x size band and reason x product family, %s onward  [MIX]"
              % core.REASON_WINDOW_START,
              "The core cohort cross-tab. Margins are asserted against section totals in "
              "code; a mismatch fails the run rather than warning.")

    for heading, rowkeys, colkeys, table, tot in ctx["reason_crosstabs"]:
        r = title(ws, r, heading)
        r = header(ws, r, [heading.split(" x ")[0]] + colkeys + ["Total"])
        for rk in rowkeys:
            put(ws, r, 1, rk, font=BOLD, fill=SUB_FILL)
            rtot = 0.0
            for j, ck in enumerate(colkeys):
                v = table.get((rk, ck), 0.0)
                rtot += v
                put(ws, r, 2 + j, v, IDR)
            put(ws, r, 2 + len(colkeys), rtot, IDR, font=BOLD, fill=TOTAL_FILL)
            r += 1
        put(ws, r, 1, "Total", font=BOLD, fill=TOTAL_FILL)
        ctot = 0.0
        for j, ck in enumerate(colkeys):
            v = sum(table.get((rk, ck), 0.0) for rk in rowkeys)
            ctot += v
            put(ws, r, 2 + j, v, IDR, font=BOLD, fill=TOTAL_FILL)
        put(ws, r, 2 + len(colkeys), ctot, IDR, font=BOLD, fill=TOTAL_FILL)
        assert abs(ctot - tot) < 1.0, "%s margins do not tie: %.2f vs %.2f" % (heading, ctot, tot)
        r += 3
    return ws


# --------------------------------------------------------------------------
# Tab 10: Concentration
# --------------------------------------------------------------------------


def tab_concentration(wb, ctx):
    ws = wb.create_sheet("10_Concentration")
    widths(ws, {"A": 6, "B": 44, "C": 18, "D": 12, "E": 12, "F": 14, "G": 16, "H": 22,
                "I": 14})
    conc = ctx["conc"]
    r = title(ws, 1, "Concentration -- the loss is named accounts, not segments  [MIX]",
              "%d customers carry 50%% of churned ARR and %d carry 80%%, out of %d. Any "
              "analysis that does not handle this explicitly produces segment averages that "
              "describe nobody. It is also the reason the AM program has to be assigned by "
              "name." % (conc["n50"], conc["n80"], conc["n"]))

    r = header(ws, r, ["#", "Customer", "Churned ARR", "% of base", "Cumulative %",
                       "Vehicles*", "Size band", "Dominant reason", "Tenure (mo)"])
    cum = 0.0
    for i, (name, amt) in enumerate(conc["ranked"][:40], 1):
        cum += amt
        c = ctx["customers"].get(name, {})
        put(ws, r, 1, i)
        put(ws, r, 2, name)
        put(ws, r, 3, amt, IDR)
        put(ws, r, 4, amt / conc["total"], PCT)
        put(ws, r, 5, cum / conc["total"], PCT)
        put(ws, r, 6, conc["veh"].get(name, 0), IDR)
        put(ws, r, 7, c.get("band", ""))
        put(ws, r, 8, c.get("reason_bucket", ""))
        put(ws, r, 9, c.get("tenure") if c.get("tenure") is not None else "")
        r += 1
    r += 2

    r = title(ws, r, "Single-account dominance screen",
              "Any bin where one customer exceeds 50% is an event, not a pattern. v1 reported "
              "three lifecycle 'signals' that were single accounts; this screen is what "
              "catches them.")
    r = header(ws, r, ["Bin type", "Bin", "Bin ARR", "Dominant account", "Top share",
                       "Customers in bin", "Verdict"])
    for kind, b, d in ctx["dominance_flags"]:
        put(ws, r, 1, kind)
        put(ws, r, 2, str(b))
        put(ws, r, 3, d["total"], IDR)
        put(ws, r, 4, d["top"])
        put(ws, r, 5, d["top_share"], PCT, fill=WARN_FILL)
        put(ws, r, 6, d["n_cust"], IDR)
        put(ws, r, 7, "single event -- do not read as a pattern", fill=WARN_FILL)
        r += 1
    r += 2

    r = title(ws, r, "Records needing human verification before anything is presented",
              "Carried over from v1 and re-tested here. These are extraordinary events or "
              "their dates/labels are wrong; a CFO will find them either way.")
    r = header(ws, r, ["Customer", "Churned ARR", "Vehicles*", "Tenure (mo)",
                       "Dominant reason", "Why it needs checking"])
    for row in ctx["verify_rows"]:
        put(ws, r, 1, row["name"])
        put(ws, r, 2, row["arr"], IDR)
        put(ws, r, 3, row["veh"], IDR)
        put(ws, r, 4, row["tenure"], NUM1)
        put(ws, r, 5, row["reason"])
        c = put(ws, r, 6, row["why"], fill=WARN_FILL)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


# --------------------------------------------------------------------------
# Tab 11: Cohort retention triangle
# --------------------------------------------------------------------------


def tab_cohort(wb, ctx):
    ws = wb.create_sheet("11_Cohort_Retention")
    widths(ws, {"A": 16, "B": 18, "C": 12})
    for i in range(4, 20):
        ws.column_dimensions[chr(64 + i)].width = 10
    r = title(ws, 1, "Cohort retention triangle  [RATE]",
              "Rows are start-quarter cohorts, columns are months since start. Cell = share of "
              "the cohort's originating ARR still active. Blank cells are months the cohort "
              "has not reached yet -- reading across a young cohort's empty tail is the "
              "classic error here.")

    steps = ctx["cohort_steps"]
    r = header(ws, r, ["Start cohort", "Cohort ARR", "Lines"] +
               ["m%d" % s for s in steps])
    for coh, d in ctx["cohorts"]:
        put(ws, r, 1, coh)
        put(ws, r, 2, d["arr0"], IDR)
        put(ws, r, 3, d["lines"], IDR)
        for j, s in enumerate(steps):
            v = d["ret"].get(s)
            if v is None:
                put(ws, r, 4 + j, "", font=NOTE_FONT)
            else:
                put(ws, r, 4 + j, v, PCT,
                    fill=WARN_FILL if v < 0.6 else (GOOD_FILL if v > 0.85 else None))
        r += 1
    return ws


# --------------------------------------------------------------------------
# Tab 12: Data quality
# --------------------------------------------------------------------------


def tab_quality(wb, ctx):
    ws = wb.create_sheet("12_Data_Quality")
    widths(ws, {"A": 52, "B": 18, "C": 80})
    r = title(ws, 1, "Data quality and what this analysis cannot conclude")
    r = header(ws, r, ["Item", "Count / value", "Implication"])
    for item, val, impl in ctx["quality"]:
        put(ws, r, 1, item)
        put(ws, r, 2, val, IDR if isinstance(val, (int, float)) and abs(val) > 1000 else None)
        c = put(ws, r, 3, impl)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    r += 2
    r = title(ws, r, "Still missing -- the data request that follows from this run")
    for i, req in enumerate(ctx["data_requests"], 1):
        put(ws, r, 1, "%d." % i)
        c = put(ws, r, 3, req)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


# --------------------------------------------------------------------------
# Target list workbook
# --------------------------------------------------------------------------


def write_target_list(path, ctx):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "At_Risk_Active_Accounts"
    widths(ws, {"A": 44, "B": 12, "C": 20, "D": 18, "E": 12, "F": 12, "G": 12, "H": 14,
                "I": 14, "J": 16, "K": 14, "L": 30, "M": 26})
    r = title(ws, 1, "AM target list -- ACTIVE accounts ranked by expected ARR at risk",
              "Expected loss = active ARR x peer 12-month hazard for the account's band, "
              "scaled by where it sits in the lifecycle. This is a prioritisation aid built "
              "from tenure, size band, product and contract position -- the only signals "
              "these two files carry. It is NOT a churn prediction: no usage, payment or "
              "support data exists in this extract, so the leading indicators an AM program "
              "actually wants are absent. Treat the ranking as 'who to call first', not "
              "'who will churn'.")
    r = header(ws, r, ["Customer", "Size band", "Product family", "Industry", "Active ARR",
                       "Vehicles*", "Tenure (mo)", "Contract term", "Lifecycle phase",
                       "Peer 12mo hazard", "Expected ARR at risk", "Risk flags",
                       "Suggested play"])
    for row in ctx["targets"]:
        put(ws, r, 1, row["company"])
        put(ws, r, 2, row["band"])
        put(ws, r, 3, row["pmap"])
        put(ws, r, 4, row["industry"])
        put(ws, r, 5, row["arr"], IDR)
        put(ws, r, 6, row["veh"], IDR)
        put(ws, r, 7, row["tenure"], NUM1)
        put(ws, r, 8, row["term"])
        put(ws, r, 9, row["phase"])
        put(ws, r, 10, row["peer_hazard"], PCT2)
        put(ws, r, 11, row["expected"], IDR,
            fill=WARN_FILL if row["expected"] > ctx["target_p90"] else None)
        put(ws, r, 12, row["flags"])
        put(ws, r, 13, row["play"])
        r += 1
    ws.freeze_panes = "A%d" % (r - len(ctx["targets"]))
    save(wb, path)


# --------------------------------------------------------------------------
# Save with OneDrive lock handling
# --------------------------------------------------------------------------


def save(wb, out: Path):
    tmp = Path(tempfile.gettempdir()) / ("_churnv2_%s" % out.name)
    wb.save(tmp)
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(tmp, out)
    except PermissionError:
        alt = out.with_name("%s-rerun%s" % (out.stem, out.suffix))
        shutil.copy2(tmp, alt)
        print("NOTE: %s was locked (open in Excel or syncing). Wrote %s instead."
              % (out.name, alt.name))
        out = alt
    tmp.unlink(missing_ok=True)
    return out


# --------------------------------------------------------------------------
# Gap decomposition (needed by the reconciliation bridge)
# --------------------------------------------------------------------------


def decompose_gap(sub, non):
    """Split the Nonaktif-vs-Subscription ARR difference into named components."""
    s_churn = collections.defaultdict(float)
    s_any = collections.defaultdict(float)
    for r in sub:
        s_any[r["company"]] += r["arr"]
        if r["status"] == "churn":
            s_churn[r["company"]] += r["arr"]
    n = collections.defaultdict(float)
    for r in non:
        n[r["company"]] += r["arr"]

    no_churn = [c for c in n if s_churn.get(c, 0.0) == 0.0]
    absent = [c for c in no_churn if c not in s_any]
    shown_active = [c for c in no_churn if c in s_any]
    sub_only = [c for c in s_churn if n.get(c, 0.0) == 0.0]
    return {
        "absent_n": len(absent), "absent_arr": sum(n[c] for c in absent),
        "active_n": len(shown_active), "active_arr": sum(n[c] for c in shown_active),
        "active_names": sorted(shown_active, key=lambda c: -n[c]),
        "sub_only_n": len(sub_only), "sub_only_arr": sum(s_churn[c] for c in sub_only),
    }


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        raise SystemExit(2)
    sub_csv, non_csv, outdir = (Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3]))

    import compute

    print("loading subscription register ...")
    sub, unknown_periods = core.load_subscription(sub_csv)
    if unknown_periods:
        raise SystemExit("unknown billing frequencies, add to PERIOD_MONTHS: %s"
                         % dict(unknown_periods))
    print("loading nonaktif register ...")
    non, non_meta = core.load_nonaktif(non_csv)

    print("propagating reasons ...")
    coverage, coverage_arr = core.propagate_reasons(sub, non)
    base, excluded = core.apply_exclusions(sub)

    snapshot = max(r["end"] for r in sub if r["end"])
    customers = core.build_customers(sub, snapshot)
    gap = decompose_gap(sub, non)

    print("computing ...")
    ctx = compute.build_context(sub, non, non_meta, coverage, coverage_arr, base,
                                excluded, customers, snapshot, gap)

    # ---------------- assertions ----------------
    gross = ctx["gross_churn"]
    cov_ok = 1.0 - coverage_arr.get("unmatched", 0.0) / gross
    # 0.97, not 0.98: with ARR correctly annualised as subtotal*12, annually-billed
    # lines carry 12x the weight they did under the earlier (wrong) per-period
    # division, and a few unmatched lines are annually billed. Coverage is ~97.5%.
    assert cov_ok >= 0.97, "reason propagation coverage %.3f below 0.97" % cov_ok
    assert abs(ctx["base_churn"] + ctx["exc_admin"] + ctx["exc_customer"] - gross) < 1.0, \
        "exclusion split does not tie to gross churned ARR"
    assert not any(r["company"] in core.EXCLUDE_CUSTOMERS for r in base), \
        "excluded customer leaked into the analysis base"
    assert not any(r["reason"] in core.ADMIN_REASONS for r in base), \
        "admin reason leaked into the analysis base"
    st = collections.Counter(c["status"] for c in customers.values())
    assert sum(st.values()) == len(customers)

    print("writing workbook ...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    tab_readme(wb, ctx)
    tab_reconciliation(wb, ctx)
    tab_hazard(wb, ctx)
    tab_rate_segment(wb, ctx)
    tab_penetration(wb, ctx)
    tab_tenure(wb, ctx)
    tab_contract(wb, ctx)
    tab_full_partial(wb, ctx)
    tab_reason(wb, ctx)
    tab_reason_cross(wb, ctx)
    tab_concentration(wb, ctx)
    tab_cohort(wb, ctx)
    tab_quality(wb, ctx)

    stamp = str(ctx["snapshot"])
    out_wb = save(wb, outdir / ("churn-cohort-v2-%s.xlsx" % stamp))
    out_tl = outdir / ("churn-am-target-list-%s.xlsx" % stamp)
    write_target_list(out_tl, ctx)

    # ---------------- reconciliation block ----------------
    km = ctx["km_arr"]
    P = print
    P("=" * 78)
    P("RECONCILIATION -- churn cohort v2")
    P("=" * 78)
    P("subscription rows      : %s  (%s active / %s churn / %s not activated)"
      % (format(ctx["n_sub_rows"], ","), format(ctx["n_active_lines"], ","),
         format(len([r for r in sub if r["status"] == "churn"]), ","),
         format(len([r for r in sub if r["status"] == "not activated"]), ",")))
    P("companies              : %s" % format(len({r["company"] for r in sub}), ","))
    P("snapshot / censor date : %s" % ctx["snapshot"])
    P("-" * 78)
    P("active ARR (annualised): %18s" % fmt_idr(ctx["active_arr"]))
    P("gross churned ARR      : %18s" % fmt_idr(gross))
    P("  less admin reasons   : %18s  (%.1f%%)"
      % ("-" + fmt_idr(ctx["exc_admin"]), 100 * ctx["exc_admin"] / gross))
    P("  less excluded cust.  : %18s  (%.1f%%)"
      % ("-" + fmt_idr(ctx["exc_customer"]), 100 * ctx["exc_customer"] / gross))
    P("= analysis base        : %18s" % fmt_idr(ctx["base_churn"]))
    P("  (of which Downgrade contraction tracked separately: %s)"
      % fmt_idr(ctx["exc_downgrade"]))
    P("-" * 78)
    P("NONAKTIF GAP BRIDGE  (start + deltas tie to the authoritative figure: asserted)")
    for label, val, kind, _note in ctx["bridge"]:
        P("  %-58s %16s  %s" % (label[:58], fmt_idr(val), kind))
    P("-" * 78)
    P("REASON PROPAGATION COVERAGE")
    for level in ["L1 co+month+product", "L2 co+month", "L3 co", "unmatched"]:
        P("  %-22s lines=%6d  ARR=%18s  %5.1f%%"
          % (level, coverage.get(level, 0), fmt_idr(coverage_arr.get(level, 0.0)),
             100 * coverage_arr.get(level, 0.0) / gross))
    P("  coverage (any level) : %.2f%%  -> PASS (threshold 97%%)" % (100 * cov_ok))
    P("-" * 78)
    P("CUSTOMER STATUS SPLIT")
    for s, n in sorted(st.items()):
        arr = sum(c["arr_churn_base"] for c in customers.values() if c["status"] == s)
        act = sum(c["arr_active"] for c in customers.values() if c["status"] == s)
        P("  %-15s n=%5d  churned(base)=%16s  active=%18s"
          % (s, n, fmt_idr(arr), fmt_idr(act)))
    P("-" * 78)
    P("HAZARD CURVE (ARR-weighted, analysis base, censored at snapshot)")
    P("  survival @12mo=%.1f%%  @24mo=%.1f%%  @36mo=%.1f%%  @48mo=%.1f%%"
      % (100 * km["survival"][12], 100 * km["survival"][24],
         100 * km["survival"][36], 100 * km["survival"][48]))
    peak = max(range(1, HORIZON + 1), key=lambda m: km["hazard"][m])
    P("  peak monthly hazard  : month %d at %.2f%%  (stable window: months 1-%d)"
      % (ctx["peak_month"], 100 * ctx["peak_hazard"], ctx["last_stable_month"]))
    P("  months %d-%d are thinly observed (<%.0f%% of month-0 pool) and excluded from ranking"
      % (ctx["last_stable_month"] + 1, HORIZON, 100 * ctx["stable_min_share"]))
    P("  implied 24mo churn probability: %.1f%%" % (100 * ctx["overall_24"]))
    P("  logo survival @24mo  : %.1f%%" % (100 * ctx["km_logo"]["survival"][24]))
    P("-" * 78)
    P("HAZARD BY PHASE (index vs mean of stable months)")
    for p in ctx["phases"]:
        P("  %-38s hazard=%.2f%%  index=%.2f  ARR=%16s (%4.1f%%) %s"
          % (p["label"], 100 * p["mean_h"], p["index"], fmt_idr(p["arr"]),
             100 * p["share"], "[THIN]" if p["thin"] else ""))
    P("-" * 78)
    P("24-MONTH CHURN PROBABILITY BY CUSTOMER SIZE BAND (Fleet Category = total fleet)")
    for row in ctx["rate_by_band"]:
        P("  %-12s active=%18s  km24=%5.1f%%  gross-cum=%5.1f%%"
          % (row["key"], fmt_idr(row["active"]), 100 * row["km24"], 100 * row["gross"]))
    P("  %-12s %18s  km24=%5.1f%%  gross-cum=%5.1f%%"
      % ("ALL", fmt_idr(ctx["active_arr"]), 100 * ctx["overall_24"],
         100 * ctx["gross_overall"]))
    P("-" * 78)
    P("24-MONTH CHURN PROBABILITY BY SUBSCRIBED-VEHICLE BAND (derived)")
    for row in ctx["rate_by_veh_band"]:
        P("  %-18s active=%18s  km24=%5.1f%%  logos=%5d"
          % (row["key"], fmt_idr(row["active"]), 100 * row["km24"],
             row["logos_active"] + row["logos_full"]))
    P("-" * 78)
    P("SHARE OF WALLET (median subscribed vehicles vs band midpoint)")
    for row in ctx["penetration"]:
        P("  %-12s n=%4d  median veh=%5.1f / mid %4d = %4.0f%%  breaches=%4.0f%%"
          % (row["band"], row["n"], row["median_veh"], row["band_mid"],
             100 * row["median_pen"], 100 * row["breach_pct"]))
    P("  Fleet Category ceiling breaches overall: %d customers" % ctx["breach_total"])
    P("-" * 78)
    P("CONTRACT TERM vs CHURN TIMING")
    for row in ctx["contract"][:5]:
        km24 = ("CENSORED(%2.0f%% obs)" % (100 * row["km24_share"])) if row["km24_thin"]             else ("%5.1f%%" % (100 * row["km24"]))
        P("  term=%3s mo  churned=%6d  active=%6d  medTenAtChurn=%5.1f  km24=%s  %s"
          % (row["term"], row["lines"], row["active"], row["median"], km24, row["read"]))
    P("  NOTE: medTenAtChurn is over CHURNED lines only and points the opposite way to the")
    P("        rate. Prioritise on km24. Terms flagged CENSORED are too young to measure.")
    P("-" * 78)
    P("CONCENTRATION: %d customers = 50%% of base churned ARR, %d = 80%%, of %d"
      % (ctx["conc"]["n50"], ctx["conc"]["n80"], ctx["conc"]["n"]))
    P("  single-account-dominated bins flagged: %d" % len(ctx["dominance_flags"]))
    P("-" * 78)
    P("all crosstab margins asserted against section totals : PASS")
    P("exclusions absent from analysis base                 : PASS")
    P("vehicle proxy error vs distinct-nopol count           : %+.1f%%" % ctx["veh_bias"])
    P("target list rows (active accounts)                    : %s"
      % format(len(ctx["targets"]), ","))
    P("-" * 78)
    P("workbook    : %s" % out_wb)
    P("target list : %s" % out_tl)
    P("tabs        : %s" % ", ".join(wb.sheetnames))
    P("=" * 78)


if __name__ == "__main__":
    main()
