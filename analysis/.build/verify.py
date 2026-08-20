"""Verification pass for the comparables workbook."""
import sys

import openpyxl

sys.path.insert(0, r"C:\Users\Grady Kusmulyadi\OneDrive - PT. Otto Menara Globalindo"
                   r"\Documents\Claude\Code\grady-agents\analysis")

P = "McEasy_FMS_Comparables_Benchmark.xlsx"
ERR = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")
wf = openpyxl.load_workbook(P)
wv = openpyxl.load_workbook(P, data_only=True)

fails = []
nf = ne = 0
for n in wf.sheetnames:
    for row in wf[n].iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                nf += 1
                v = wv[n][c.coordinate].value
                if isinstance(v, str) and v in ERR:
                    ne += 1
                    print("ERR", n, c.coordinate, c.value[:70], "->", v)
print(f"sheets={len(wf.sheetnames)}  formulas={nf}  errors={ne}")
if ne:
    fails.append("formula errors")

# ---- locate labelled rows on a sheet
def rowmap(sheet, lo=1, hi=140):
    ws = wv[sheet]
    m = {}
    for r in range(lo, hi + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip():
            m.setdefault(v.strip(), r)
    return m


def chk(label, got, exp, tol=0.005, pct=False):
    ok = got is not None and abs(got - exp) <= max(abs(exp) * tol, 1e-9)
    if not ok:
        fails.append(label)
    shown = f"{got:,.4f}" if isinstance(got, float) else str(got)
    print(f"  {'OK  ' if ok else 'FAIL'} {label:56s} {shown:>16s}  expect {exp:,.4f}")


print("\n=== CONSERVATIVE CASES ===")
cc, M = wv["Conservative Cases"], rowmap("Conservative Cases")
COLS = {"base": "B", "v90": "C", "a90": "D", "v80": "E", "a80": "F"}
for lab in ("Total revenue", "EBITDA", "Net profit", "Capex", "EBITDA margin", "Net margin",
            "Sparepart share of revenue (must hold constant)", "Vehicles at Dec-2030",
            "EBITDA per vehicle per year", "Gap to the benchmark",
            "Minimum cash on the whole path", "Funding gap against that buffer"):
    r = M.get(lab)
    if not r:
        print(f"  MISSING ROW: {lab}"); fails.append(lab); continue
    vals = [cc[f"{COLS[k]}{r}"].value for k in COLS]
    fmt = (lambda v: f"{v:>13.2%}") if ("margin" in lab or "share" in lab or "Gap" in lab) \
        else (lambda v: f"{v:>13,.0f}")
    print(f"  {lab:50s}" + "".join(fmt(v) if isinstance(v, (int, float)) else f"{'--':>13s}"
                                   for v in vals))

print("\n=== SELF-CHECKS ===")
# 1. base column must reproduce the plan
mm, MMR = wv["McEasy Model"], rowmap("McEasy Model", 1, 80)
plan = {k: mm.cell(row=MMR[k], column=10).value for k in      # column J = 2030
        ("Total Revenue", "EBITDA", "Net Profit (Loss)",
         "Capex (year-on-year change in gross fixed assets)", "Depreciation")}
chk("base revenue = plan", cc[f"B{M['Total revenue']}"].value, plan["Total Revenue"], 1e-9)
chk("base EBITDA = plan", cc[f"B{M['EBITDA']}"].value, plan["EBITDA"], 1e-9)
chk("base net profit = plan", cc[f"B{M['Net profit']}"].value, plan["Net Profit (Loss)"], 1e-9)
chk("base capex = plan", cc[f"B{M['Capex']}"].value,
    plan["Capex (year-on-year change in gross fixed assets)"], 1e-9)
chk("base depreciation = plan", cc[f"B{M['Depreciation']}"].value, plan["Depreciation"], 1e-9)
chk("base growth haircut h = 1", cc[f"B{M['Growth haircut applied from Jul-2026']}"].value, 1.0,
    1e-6)

# 2. the fixed ratio Grady specified
rr = M["Sparepart share of revenue (must hold constant)"]
for k, col in COLS.items():
    chk(f"spare share {k} = 17.8524%", cc[f"{col}{rr}"].value, 0.178524, 0.001)

# 3. revenue targets hit
rv = M["Total revenue"]
for k, exp in (("v90", 90_000_000), ("a90", 90_000_000),
               ("v80", 80_000_000), ("a80", 80_000_000)):
    chk(f"{k} 2030 revenue on target", cc[f"{COLS[k]}{rv}"].value, exp, 1e-6)

# 4. independent recomputation through the engine
import conservative_engine as CE                                        # noqa: E402
CM = CE.load()
life, cum, tgt = CE.calibrate_opening_block(CM)
REF = CE.base_cohort(CM, life)
plan30 = sum(CM["P"]["rev"][x] for x in CM["keys"] if x.startswith("2030"))
chk("cohort ties to reported accumulated depreciation", cum, tgt, 0.01)
resid = None
for k, target, drv in (("base", plan30, "volume"), ("v90", 90e6, "volume"), ("a90", 90e6, "arpu"),
                       ("v80", 80e6, "volume"), ("a80", 80e6, "arpu")):
    S = CE.build(CM, target, drv, life, REF)
    if k == "base":
        _, resid = CE.cash_path(CM, S, None)
    path, _ = CE.cash_path(CM, S, resid)
    A = {n: CE.annual(S[n], CM["keys"], (2030,)) for n in ("rev", "ebitda", "np")}
    mn = min(path.values())
    chk(f"engine vs sheet: {k} EBITDA", cc[f"{COLS[k]}{M['EBITDA']}"].value, A["ebitda"][2030], 1e-6)
    chk(f"engine vs sheet: {k} net profit", cc[f"{COLS[k]}{M['Net profit']}"].value,
        A["np"][2030], 1e-6)
    chk(f"engine vs sheet: {k} min cash", cc[f"{COLS[k]}{M['Minimum cash on the whole path']}"].value,
        mn, 1e-6)

# 5. Jan-Jun 2026 actuals untouched (engine holds them; confirm against source)
S = CE.build(CM, 80e6, "arpu", life, REF)
for mth in ("2026-01", "2026-06"):
    chk(f"actual {mth} revenue preserved", S["rev"][mth], CM["P"]["rev"][mth], 1e-9)
    chk(f"actual {mth} EBITDA preserved", S["ebitda"][mth], CM["P"]["ebitda"][mth], 1e-6)

print("\n=== REGRESSION ===")
kp = wv["Karooooo Path"]
for coord, exp, lab in (("C23", 110.9854, "FY2019 sub rev USD"),
                        ("C32", 0.4498, "FY2019 EBITDA margin"),
                        ("C34", 0.2996, "FY2019 capex % rev"),
                        ("C33", 0.0216, "FY2019 FCF margin"),
                        ("C39", 9.6261, "FY2019 ARPU/mo")):
    chk(lab, kp[coord].value, exp, 0.002)
chk("listed valuation dispersion", wv["Valuation Signals"]["B20"].value, 14.79, 0.01)
mvb = wv["McEasy vs Benchmark"]
chk("EBITDA/vehicle McEasy", mvb["B20"].value, 56.54, 0.001)
chk("EBITDA/vehicle Cartrack", mvb["C20"].value, 57.85, 0.001)
cx, CXM = wv["Capex Decomposition"], rowmap("Capex Decomposition")
chk("capex per net add McEasy", cx[f"B{CXM['Capex per NET ADDITION']}"].value, 54.22, 0.001)
chk("device+install per net add Cartrack",
    cx[f"C{CXM['Device + installation capex per NET ADDITION']}"].value, 123.27, 0.001)

from collections import defaultdict                                      # noqa: E402
for sheet, lo, hi in (("Revenue by Product", 5, 26), ("Revenue by Country", 5, 33)):
    ws = wv[sheet]
    agg = defaultdict(float)
    for r in range(lo, hi + 1):
        key = (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value)
        v = ws.cell(row=r, column=6).value
        if isinstance(v, (int, float)) and key[0]:
            agg[key] += v
    bad = [(k, round(v, 4)) for k, v in agg.items() if abs(v - 1) > 0.005]
    print(f"  {'OK  ' if not bad else 'FAIL'} {sheet} % groups sum to 100%   "
          f"({len(agg)} groups)" + ("" if not bad else f"  {bad}"))
    if bad:
        fails.append(sheet)

print("\n" + ("ALL CHECKS PASSED" if not fails else f"FAILURES: {fails}"))
sys.exit(1 if fails else 0)
